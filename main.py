import sys
import json
import mysql.connector
from PySide6.QtWidgets import QCompleter,QTabWidget

from PySide6.QtGui import QPainter, QPainterPath, QColor
from PySide6.QtCore import Qt   
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
                              QLabel, QLineEdit, QPushButton, QMessageBox, QStackedWidget,
                              QScrollArea, QCheckBox, QTimeEdit, QComboBox, QFrame,
                              QGridLayout, QDialog, QFileDialog, QButtonGroup, QRadioButton, 
                              QGroupBox, QTextEdit,QDialogButtonBox,QTableWidgetItem,QFormLayout)
from PySide6.QtCore import Qt, QEasingCurve, QPropertyAnimation, QTime, QTimer, QDate
from PySide6.QtGui import QFont, QPixmap, QIcon, QColor, QDoubleValidator
from PySide6.QtWidgets import QTableWidget
from datetime import datetime, timedelta
# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'PythonProject'
}


class DatabaseManager:

    @staticmethod
    def get_member_growth(owner_id, start_date, end_date):
        return DatabaseManager.execute_query("""
            SELECT DATE_FORMAT(s.created_at, '%Y-%m') AS month, COUNT(DISTINCT s.member_id) AS new_members
            FROM subscriptions s
            JOIN gyms g ON s.gym_id = g.id
            WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
            AND s.created_at BETWEEN %s AND %s
            GROUP BY month
            ORDER BY month
        """, (owner_id, start_date, end_date), fetch_all=True)

    @staticmethod
    def get_revenue_trend(owner_id, start_date, end_date):
        return DatabaseManager.execute_query("""
            SELECT DATE_FORMAT(s.created_at, '%Y-%m') AS month,
                   SUM(
                       CASE 
                           WHEN s.subscription_type = 'monthly' THEN 
                               JSON_UNQUOTE(JSON_EXTRACT(g.subscriptions, '$.monthly.price'))
                           WHEN s.subscription_type = 'quarterly' THEN 
                               JSON_UNQUOTE(JSON_EXTRACT(g.subscriptions, '$.quarterly.price'))
                           WHEN s.subscription_type = 'annual' THEN 
                               JSON_UNQUOTE(JSON_EXTRACT(g.subscriptions, '$.annual.price'))
                           ELSE 0
                       END
                   ) AS revenue
            FROM subscriptions s
            JOIN gyms g ON s.gym_id = g.id
            WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
            AND s.created_at BETWEEN %s AND %s
            GROUP BY month
            ORDER BY month
        """, (owner_id, start_date, end_date), fetch_all=True)

    @staticmethod
    def get_subscription_types(owner_id):
        return DatabaseManager.execute_query("""
            SELECT s.subscription_type, COUNT(*) AS count
            FROM subscriptions s
            JOIN gyms g ON s.gym_id = g.id
            WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
            GROUP BY s.subscription_type
        """, (owner_id,), fetch_all=True)

    @staticmethod
    def get_attendance_data(owner_id, start_date, end_date):
        return DatabaseManager.execute_query("""
            SELECT DATE(w.workout_date) AS date, COUNT(*) AS workouts
            FROM workout_tracking w
            JOIN gyms g ON w.gym_id = g.id
            WHERE g.owner_id = %s AND w.workout_date BETWEEN %s AND %s
            GROUP BY date
            ORDER BY date
        """, (owner_id, start_date, end_date), fetch_all=True)

    @staticmethod
    def get_detailed_member_stats(owner_id):
        return DatabaseManager.execute_query("""
            SELECT 
                u.id as member_id,
                u.first_name, 
                u.last_name, 
                u.email,
                u.phone,
                s.subscription_type,
                s.payment_status,
                s.start_date,
                s.end_date,
                g.name AS gym_name,
                g.address AS gym_address,
                TIMESTAMPDIFF(DAY, CURDATE(), s.end_date) AS days_remaining,
                CASE 
                    WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() THEN 'Actif'
                    WHEN s.payment_status = 'confirmed' AND s.end_date < CURDATE() THEN 'Expiré'
                    WHEN s.payment_status = 'pending' THEN 'En attente'
                    ELSE 'Inactif'
                END AS status
            FROM users u
            JOIN subscriptions s ON u.id = s.member_id
            JOIN gyms g ON s.gym_id = g.id
            WHERE g.owner_id = %s
            ORDER BY 
                CASE WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() THEN 0
                    WHEN s.payment_status = 'pending' THEN 1
                    ELSE 2 END,
                s.end_date DESC
        """, (owner_id,), fetch_all=True)

    @staticmethod
    def get_gym_members(gym_id):
        return DatabaseManager.execute_query("""
            SELECT u.id, u.first_name, u.last_name 
            FROM users u
            JOIN subscriptions s ON u.id = s.member_id
            WHERE s.gym_id = %s AND s.payment_status = 'confirmed'
        """, (gym_id,), fetch_all=True)

    @staticmethod
    def save_workout(member_id, gym_id, date, description=None, photo_path=None):
        try:
            photo_data = None
            if photo_path:
                with open(photo_path, 'rb') as f:
                    photo_data = f.read()
            
            return DatabaseManager.execute_query("""
                INSERT INTO workout_tracking (member_id, gym_id, workout_date, description, photo)
                VALUES (%s, %s, %s, %s, %s)
            """, (member_id, gym_id, date, description, photo_data))
        except Exception as e:
            print(f"Error saving workout: {e}")
            return False

    @staticmethod
    def get_workouts(member_id, gym_id, week_start_date):
        return DatabaseManager.execute_query("""
            SELECT workout_date, description, photo IS NOT NULL as has_photo
            FROM workout_tracking
            WHERE member_id = %s AND gym_id = %s 
            AND workout_date >= %s AND workout_date < DATE_ADD(%s, INTERVAL 8 DAY)
            ORDER BY workout_date
        """, (member_id, gym_id, week_start_date, week_start_date), fetch_all=True)

    @staticmethod
    def get_workout_photo(workout_id):
        return DatabaseManager.execute_query(
            "SELECT photo FROM workout_tracking WHERE id = %s",
            (workout_id,),
            fetch_one=True
        )




   

    @staticmethod
    def save_planning(gym_id, planning_data):
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            
            # Delete existing planning for this gym
            cursor.execute("DELETE FROM gym_planning WHERE gym_id = %s", (gym_id,))
            
            # Insert new planning data
            for day, data in planning_data.items():
                cursor.execute("""
                    INSERT INTO gym_planning (gym_id, day, is_open, opening_time, closing_time)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    gym_id,
                    day,
                    data['is_open'],
                    data['opening_time'],
                    data['closing_time']
                ))
            
            conn.commit()
            return True
        except mysql.connector.Error as err:
            print(f"Database error saving planning: {err}")
            return False
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()

    @staticmethod
    def get_planning(gym_id):
        result = DatabaseManager.execute_query(
            "SELECT day, is_open, opening_time, closing_time FROM gym_planning WHERE gym_id = %s",
            (gym_id,),
            fetch_all=True
        )
        
        if not result:
            return None
        
        planning = {}
        for row in result:
            planning[row['day']] = {
                'is_open': bool(row['is_open']),
                'opening_time': str(row['opening_time']) if row['opening_time'] else None,
                'closing_time': str(row['closing_time']) if row['closing_time'] else None
            }
        return planning



    


    

        
    
    @staticmethod
    def initialize_database():
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            
            # Users table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    first_name VARCHAR(100) NOT NULL,
                    last_name VARCHAR(100) NOT NULL,
                    age INT NOT NULL,
                    gender VARCHAR(50),
                    phone VARCHAR(20),
                    email VARCHAR(255) NOT NULL UNIQUE,
                    password VARCHAR(255) NOT NULL,
                    role ENUM('member', 'gym owner') NOT NULL,
                    address VARCHAR(255)
                )
            """)
            # Workout tracking table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS workout_tracking (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    gym_id INT NOT NULL,
                    workout_date DATE NOT NULL,
                    description TEXT,
                    photo LONGBLOB,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES users(id),
                    FOREIGN KEY (gym_id) REFERENCES gyms(id)
                )
            """)
            # Planning table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS gym_planning (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    gym_id INT NOT NULL,
                    day VARCHAR(10) NOT NULL,
                    is_open BOOLEAN DEFAULT FALSE,
                    opening_time TIME,
                    closing_time TIME,
                    FOREIGN KEY (gym_id) REFERENCES gyms(id),
                    UNIQUE KEY (gym_id, day)
                )
            """)
            # Messages table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS gym_messages (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    gym_id INT NOT NULL,
                    sender_id INT NOT NULL,
                    message_type ENUM('text', 'image', 'audio') NOT NULL,
                    content TEXT,
                    media LONGBLOB,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (gym_id) REFERENCES gyms(id),
                    FOREIGN KEY (sender_id) REFERENCES users(id)
                )
            """)
            # Dans DatabaseManager.initialize_database(), ajoutez cette table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS gym_reviews (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    gym_id INT NOT NULL,
                    member_id INT NOT NULL,
                    rating INT NOT NULL CHECK (rating BETWEEN 1 AND 5),
                    comment TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (gym_id) REFERENCES gyms(id),
                    FOREIGN KEY (member_id) REFERENCES users(id)
                )
            """)
            
            # Gyms table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS gyms (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    address VARCHAR(255) NOT NULL,
                    owner_id INT NOT NULL,
                    subscriptions JSON,
                    FOREIGN KEY (owner_id) REFERENCES users(id)
                )
            """)

            cursor.execute("""
                           CREATE TABLE IF NOT EXISTS profile_changes (
    id INT AUTO_INCREMENT PRIMARY KEY,
    user_id INT NOT NULL,
    change_type VARCHAR(50) NOT NULL,
    change_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (user_id) REFERENCES users(id)
)
                           """)

            
            
            # Bank details table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS bank_details (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    owner_id INT NOT NULL,
                    rib VARCHAR(100) NOT NULL,
                    FOREIGN KEY (owner_id) REFERENCES users(id)
                )
            """)
            
            # Subscriptions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS subscriptions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    gym_id INT NOT NULL,
                    subscription_type VARCHAR(20) NOT NULL,
                    payment_method ENUM('espece', 'rib') NOT NULL,
                    payment_proof BLOB,
                    payment_status ENUM('pending', 'confirmed', 'rejected') DEFAULT 'pending',
                    start_date DATE,
                    end_date DATE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES users(id),
                    FOREIGN KEY (gym_id) REFERENCES gyms(id)
                )
            """)

            # Subscription payments history table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS subscription_payments (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    subscription_id INT NOT NULL,
                    payment_method ENUM('espece', 'rib') NOT NULL,
                    payment_proof BLOB,
                    payment_status ENUM('pending', 'confirmed', 'rejected') DEFAULT 'pending',
                    payment_date DATE NOT NULL,
                    subscription_type VARCHAR(20) NOT NULL,
                    extension_months INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (subscription_id) REFERENCES subscriptions(id)
                )
            """)
            
            conn.commit()
        except mysql.connector.Error as err:
            print(f"Error initializing database: {err}")
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()

    @staticmethod
    def add_gym_review(gym_id, member_id, rating, comment):
        return DatabaseManager.execute_query("""
            INSERT INTO gym_reviews (gym_id, member_id, rating, comment)
            VALUES (%s, %s, %s, %s)
        """, (gym_id, member_id, rating, comment))

    @staticmethod
    def get_gym_reviews(gym_id):
        return DatabaseManager.execute_query("""
            SELECT r.*, u.first_name, u.last_name 
            FROM gym_reviews r
            JOIN users u ON r.member_id = u.id
            WHERE r.gym_id = %s
            ORDER BY r.created_at DESC
        """, (gym_id,), fetch_all=True)

    @staticmethod
    def get_gym_avg_rating(gym_id):
        result = DatabaseManager.execute_query("""
            SELECT AVG(rating) as avg_rating, COUNT(*) as review_count
            FROM gym_reviews
            WHERE gym_id = %s
        """, (gym_id,), fetch_one=True)
        
        if result and result['avg_rating']:
            return {
                'avg': round(float(result['avg_rating']), 1),
                'count': result['review_count']
            }
        return {'avg': 0, 'count': 0}

    @staticmethod
    def execute_query(query, params=None, fetch_one=False, fetch_all=False):
        conn = None
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query, params or ())
            
            if fetch_one:
                result = cursor.fetchone()
                # Consume any remaining results to avoid "Unread result found"
                cursor.fetchall()
                return result
            elif fetch_all:
                return cursor.fetchall()
            else:
                conn.commit()
                return True
                
        except mysql.connector.Error as err:
            print(f"Database error: {err}")
            print(f"Query: {query}")
            print(f"Params: {params}")
            return False
                
        finally:
            if conn and conn.is_connected():
                cursor.close()
                conn.close()
    @staticmethod
    def get_gym_stats(owner_id):
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)
            
            # Nombre de salles
            cursor.execute("SELECT COUNT(*) as count FROM gyms WHERE owner_id = %s", (owner_id,))
            gyms_count = cursor.fetchone()['count']
            
            # Nombre de membres (abonnements confirmés)
            cursor.execute("""
                SELECT COUNT(DISTINCT s.member_id) as count 
                FROM subscriptions s
                JOIN gyms g ON s.gym_id = g.id
                WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
            """, (owner_id,))
            members_count = cursor.fetchone()['count']
            
            # Revenus totaux - version simplifiée sans JSON_TABLE
            cursor.execute("""
                SELECT SUM(
                    CASE 
                        WHEN s.subscription_type = 'monthly' THEN 
                            (SELECT JSON_EXTRACT(g.subscriptions, '$.monthly.price') FROM gyms g WHERE g.id = s.gym_id)
                        WHEN s.subscription_type = 'quarterly' THEN 
                            (SELECT JSON_EXTRACT(g.subscriptions, '$.quarterly.price') FROM gyms g WHERE g.id = s.gym_id)
                        WHEN s.subscription_type = 'annual' THEN 
                            (SELECT JSON_EXTRACT(g.subscriptions, '$.annual.price') FROM gyms g WHERE g.id = s.gym_id)
                        ELSE 0
                    END
                ) as total
                FROM subscriptions s
                JOIN gyms g ON s.gym_id = g.id
                WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
            """, (owner_id,))
            
            revenue_result = cursor.fetchone()
            revenue = float(revenue_result['total']) if revenue_result['total'] else 0
            
            return {
                'gyms': gyms_count,
                'members': members_count,
                'revenue': revenue
            }
            
        except mysql.connector.Error as err:
            print(f"Error getting stats: {err}")
            return None
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()

    @staticmethod
    def send_message(gym_id, sender_id, message_type, content=None, media_path=None):
        try:
            media_data = None
            if media_path:
                with open(media_path, 'rb') as f:
                    media_data = f.read()
            
            return DatabaseManager.execute_query("""
                INSERT INTO gym_messages (gym_id, sender_id, message_type, content, media)
                VALUES (%s, %s, %s, %s, %s)
            """, (gym_id, sender_id, message_type, content, media_data))
        except Exception as e:
            print(f"Error sending message: {e}")
            return False

    @staticmethod
    def get_messages(gym_id, limit=50):
        return DatabaseManager.execute_query("""
            SELECT m.*, u.first_name, u.last_name 
            FROM gym_messages m
            JOIN users u ON m.sender_id = u.id
            WHERE m.gym_id = %s
            ORDER BY m.created_at DESC
            LIMIT %s
        """, (gym_id, limit), fetch_all=True)

    @staticmethod
    def get_message_media(message_id):
        return DatabaseManager.execute_query(
            "SELECT media FROM gym_messages WHERE id = %s",
            (message_id,),
            fetch_one=True
        )
class BaseDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_modern_base()
        self.setup_ui()

    def setup_modern_base(self):
        # Modern dialog base styling with gym-appropriate dark colors
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2c2c2c, stop:0.5 #1a1a1a, stop:1 #2c2c2c);
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.3);
            }
            QLabel {
                color: white;
                font-weight: 500;
                background: transparent;
                border: none;
            }
            QLineEdit {
                background: rgba(255, 255, 255, 0.15);
                border: 1px solid rgba(255, 107, 53, 0.3);
                border-radius: 10px;
                padding: 12px 15px;
                font-size: 14px;
                color: white;
                font-weight: 500;
            }
            QLineEdit:focus {
                border: 2px solid rgba(255, 107, 53, 0.6);
                background: rgba(255, 255, 255, 0.2);
            }
            QLineEdit::placeholder {
                color: rgba(255, 255, 255, 0.6);
            }
            QTextEdit {
                background: rgba(255, 255, 255, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 10px;
                padding: 12px;
                font-size: 14px;
                color: white;
                font-weight: 500;
            }
            QComboBox {
                background: rgba(255, 255, 255, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 10px;
                padding: 12px 15px;
                font-size: 14px;
                color: white;
                font-weight: 500;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid white;
                margin-right: 10px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4facfe, stop:1 #00f2fe);
                color: white;
                font-size: 14px;
                font-weight: 600;
                padding: 12px 20px;
                border-radius: 10px;
                border: none;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #06b6d4);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563eb, stop:1 #0891b2);
            }
        """)

    def setup_ui(self):
        raise NotImplementedError("Subclasses must implement setup_ui method")

class BankDetailsDialog(BaseDialog):
    def __init__(self, owner_id, parent=None):
        self.owner_id = owner_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("Ajouter vos coordonnées bancaires")
        self.setFixedSize(400, 200)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # RIB Input
        rib_label = QLabel("RIB:")
        self.rib_input = QLineEdit()
        self.rib_input.setPlaceholderText("Entrez votre RIB")
        self.rib_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
        """)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Enregistrer")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
            }
        """)
        save_btn.clicked.connect(self.save_rib)
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f5f5f5;
                padding: 8px 16px;
                border-radius: 4px;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)
        
        layout.addWidget(rib_label)
        layout.addWidget(self.rib_input)
        layout.addLayout(button_layout)
        
        self.load_existing_rib()
        
    def load_existing_rib(self):
        result = DatabaseManager.execute_query(
            "SELECT rib FROM bank_details WHERE owner_id = %s", 
            (self.owner_id,), 
            fetch_one=True
        )
        if result:
            self.rib_input.setText(result['rib'])
    
    def save_rib(self):
        rib = self.rib_input.text().strip()
        if not rib:
            QMessageBox.warning(self, "Champ requis", "Veuillez entrer votre RIB")
            return
            
        existing = DatabaseManager.execute_query(
            "SELECT id FROM bank_details WHERE owner_id = %s", 
            (self.owner_id,), 
            fetch_one=True
        )
        
        if existing:
            success = DatabaseManager.execute_query(
                "UPDATE bank_details SET rib = %s WHERE owner_id = %s", 
                (rib, self.owner_id)
            )
        else:
            success = DatabaseManager.execute_query(
                "INSERT INTO bank_details (owner_id, rib) VALUES (%s, %s)", 
                (self.owner_id, rib)
            )
            
        if success:
            QMessageBox.information(self, "Succès", "RIB enregistré avec succès!")
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de l'enregistrement")



class SubscriptionDialog(BaseDialog):
    def __init__(self, gym_id, member_id, parent=None, is_renewal=False):
        self.gym_id = gym_id
        self.member_id = member_id
        self.is_renewal = is_renewal
        super().__init__(parent)
        
    def setup_ui(self):
        title = "🔄 Renouveler l'abonnement" if self.is_renewal else "📝 S'abonner"
        self.setWindowTitle(title)
        self.setFixedSize(500, 500)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)

        # Load gym info
        self.gym_info = self.load_gym_info()
        if not self.gym_info:
            self.reject()
            return

        # Load subscription options from gym data
        try:
            self.subscription_options = json.loads(self.gym_info['subscriptions']) if self.gym_info['subscriptions'] else {}
        except:
            self.subscription_options = {}

        if not self.subscription_options:
            QMessageBox.warning(self, "Erreur", "Cette salle n'a pas configuré d'options d'abonnement")
            self.reject()
            return

        # Header section
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(40, 167, 69, 0.1), stop:1 rgba(32, 201, 151, 0.1));
                border-radius: 12px;
                border: 1px solid rgba(40, 167, 69, 0.3);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(15, 15, 15, 15)

        # Gym name
        gym_name = QLabel(self.gym_info['name'])
        gym_name.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(gym_name)

        # Show current subscription info if renewal
        if self.is_renewal:
            current_sub = DatabaseManager.execute_query("""
                SELECT s.*,
                       CASE
                           WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() THEN 'Actif'
                           WHEN s.payment_status = 'confirmed' AND s.end_date < CURDATE() THEN 'Expiré'
                           ELSE 'Inactif'
                       END AS status
                FROM subscriptions s
                WHERE s.member_id = %s AND s.gym_id = %s
                ORDER BY s.end_date DESC
                LIMIT 1
            """, (self.member_id, self.gym_id), fetch_one=True)

            if current_sub:
                status_text = f"📊 Abonnement actuel: {current_sub['subscription_type'].capitalize()}"
                status_text += f"\n📅 Statut: {current_sub['status']}"
                status_text += f"\n⏰ Fin: {current_sub['end_date']}"

                current_info = QLabel(status_text)
                current_info.setStyleSheet("font-size: 14px; color: #34495e; margin-top: 10px;")
                header_layout.addWidget(current_info)

        layout.addWidget(header_container)
        
        # Subscription options
        subs_label = QLabel("Choisissez votre abonnement:")
        subs_label.setStyleSheet("font-weight: bold; margin-top: 15px;")
        layout.addWidget(subs_label)
        
        self.subscription_group = QButtonGroup()
        for sub_type, sub_info in self.subscription_options.items():
            radio = QRadioButton(f"{sub_info['duration']} - {sub_info['price']}€")
            radio.sub_type = sub_type
            radio.sub_info = sub_info
            self.subscription_group.addButton(radio)
            layout.addWidget(radio)
            
        # Payment method
        payment_label = QLabel("Méthode de paiement:")
        payment_label.setStyleSheet("font-weight: bold; margin-top: 15px;")
        layout.addWidget(payment_label)
        
        self.payment_group = QButtonGroup()
        cash_radio = QRadioButton("Paiement en espèces (sur place)")
        cash_radio.payment_method = "espece"
        self.payment_group.addButton(cash_radio)
        layout.addWidget(cash_radio)
        
        rib_radio = QRadioButton("Paiement par virement")
        rib_radio.payment_method = "rib"
        self.payment_group.addButton(rib_radio)
        layout.addWidget(rib_radio)
        
        # RIB display
        self.rib_label = QLabel()
        self.rib_label.setWordWrap(True)
        self.rib_label.setStyleSheet("background-color: #f5f5f5; padding: 10px; border-radius: 5px;")
        self.rib_label.hide()
        layout.addWidget(self.rib_label)
        
        # Payment proof upload
        self.proof_label = QLabel("Preuve de paiement (capture d'écran):")
        self.proof_label.hide()
        layout.addWidget(self.proof_label)
        
        self.proof_layout = QHBoxLayout()
        self.proof_path = QLineEdit()
        self.proof_path.setReadOnly(True)
        browse_btn = QPushButton("Parcourir...")
        browse_btn.clicked.connect(self.browse_proof)
        self.proof_layout.addWidget(self.proof_path)
        self.proof_layout.addWidget(browse_btn)
        self.proof_layout.setContentsMargins(0, 0, 0, 0)
        layout.addLayout(self.proof_layout)
        self.proof_layout.setEnabled(False)
        
        # Connect signals
        self.payment_group.buttonClicked.connect(self.on_payment_method_changed)
        
        # Submit button
        submit_btn = QPushButton("Confirmer l'abonnement")
        submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                padding: 10px;
                border-radius: 5px;
                margin-top: 20px;
            }
        """)
        submit_btn.clicked.connect(self.confirm_subscription)
        layout.addWidget(submit_btn)
        
    def load_gym_info(self):
        return DatabaseManager.execute_query(
            "SELECT * FROM gyms WHERE id = %s", 
            (self.gym_id,), 
            fetch_one=True
        )
    
    def on_payment_method_changed(self, button):
        if button.payment_method == "rib":
            self.load_gym_rib()
            self.proof_label.show()
            self.proof_layout.setEnabled(True)
        else:
            self.rib_label.hide()
            self.proof_label.hide()
            self.proof_layout.setEnabled(False)
    
    def load_gym_rib(self):
        result = DatabaseManager.execute_query("""
            SELECT b.rib 
            FROM bank_details b
            JOIN gyms g ON g.owner_id = b.owner_id
            WHERE g.id = %s
        """, (self.gym_id,), fetch_one=True)
        
        if result:
            self.rib_label.setText(f"Veuillez effectuer le virement à ce RIB:\n{result['rib']}")
            self.rib_label.show()
        else:
            QMessageBox.warning(self, "RIB non disponible", "Le propriétaire n'a pas encore enregistré de RIB")
            self.payment_group.buttons()[0].setChecked(True)
    
    def browse_proof(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Sélectionner une preuve de paiement", 
            "", 
            "Images (*.png *.jpg *.jpeg)"
        )
        if file_path:
            self.proof_path.setText(file_path)
    
    def confirm_subscription(self):
        if not self.subscription_group.checkedButton():
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner un abonnement")
            return
            
        if not self.payment_group.checkedButton():
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner une méthode de paiement")
            return
            
        payment_method = self.payment_group.checkedButton().payment_method
        if payment_method == "rib" and not self.proof_path.text():
            QMessageBox.warning(self, "Erreur", "Veuillez fournir une preuve de paiement")
            return
            
        sub_button = self.subscription_group.checkedButton()
        sub_type = sub_button.sub_type

        # Calculate dates based on renewal or new subscription
        if self.is_renewal:
            # For renewals, get the current subscription to extend
            current_sub = DatabaseManager.execute_query("""
                SELECT id, end_date, payment_status
                FROM subscriptions
                WHERE member_id = %s AND gym_id = %s
                ORDER BY end_date DESC
                LIMIT 1
            """, (self.member_id, self.gym_id), fetch_one=True)

            if current_sub:
                # Always extend from the current end date (whether active or expired)
                current_end_date = current_sub['end_date']

                # If subscription is expired, extend from today
                # If subscription is active, extend from current end date
                if current_sub['payment_status'] == 'confirmed' and current_end_date >= datetime.now().date():
                    # Active subscription - extend from end date
                    extend_from_date = current_end_date
                else:
                    # Expired or inactive - start from today
                    extend_from_date = datetime.now().date()

                # Calculate new end date by extending the period
                if sub_type == "monthly":
                    new_end_date = extend_from_date + timedelta(days=30)
                elif sub_type == "quarterly":
                    new_end_date = extend_from_date + timedelta(days=90)
                else:  # annual
                    new_end_date = extend_from_date + timedelta(days=365)

                start_date = current_sub['start_date'].strftime("yyyy-MM-dd") if hasattr(current_sub['start_date'], 'strftime') else str(current_sub['start_date'])
                end_date = new_end_date.strftime("yyyy-MM-dd")
            else:
                # No existing subscription found - create new one
                start_date = QDate.currentDate().toString("yyyy-MM-dd")
                start_qdate = QDate.currentDate()
                if sub_type == "monthly":
                    end_date = start_qdate.addMonths(1).toString("yyyy-MM-dd")
                elif sub_type == "quarterly":
                    end_date = start_qdate.addMonths(3).toString("yyyy-MM-dd")
                else:
                    end_date = start_qdate.addYears(1).toString("yyyy-MM-dd")
        else:
            # New subscription starts today
            start_date = QDate.currentDate().toString("yyyy-MM-dd")
            start_qdate = QDate.currentDate()
            if sub_type == "monthly":
                end_date = start_qdate.addMonths(1).toString("yyyy-MM-dd")
            elif sub_type == "quarterly":
                end_date = start_qdate.addMonths(3).toString("yyyy-MM-dd")
            else:
                end_date = start_qdate.addYears(1).toString("yyyy-MM-dd")
        
        proof_data = None
        if payment_method == "rib" and self.proof_path.text():
            with open(self.proof_path.text(), "rb") as f:
                proof_data = f.read()
        
        if self.is_renewal:
            # For renewals, update the existing subscription instead of creating new one
            existing_sub = DatabaseManager.execute_query("""
                SELECT id, end_date, payment_status
                FROM subscriptions
                WHERE member_id = %s AND gym_id = %s
                ORDER BY end_date DESC
                LIMIT 1
            """, (self.member_id, self.gym_id), fetch_one=True)

            if existing_sub:
                # Update the existing subscription with new end date
                success = DatabaseManager.execute_query("""
                    UPDATE subscriptions
                    SET subscription_type = %s,
                        end_date = %s,
                        payment_status = %s,
                        payment_method = %s,
                        payment_proof = %s
                    WHERE id = %s
                """, (
                    sub_type,
                    end_date,
                    "pending" if payment_method == "rib" else "confirmed",
                    payment_method,
                    proof_data,
                    existing_sub['id']
                ))

                # Create a payment history record
                if success:
                    DatabaseManager.execute_query("""
                        INSERT INTO subscription_payments
                        (subscription_id, payment_method, payment_proof, payment_status, payment_date, subscription_type, extension_months)
                        VALUES (%s, %s, %s, %s, CURDATE(), %s, %s)
                    """, (
                        existing_sub['id'],
                        payment_method,
                        proof_data,
                        "pending" if payment_method == "rib" else "confirmed",
                        sub_type,
                        1 if sub_type == "monthly" else (3 if sub_type == "quarterly" else 12)
                    ))
            else:
                # No existing subscription found, create new one
                success = DatabaseManager.execute_query("""
                    INSERT INTO subscriptions
                    (member_id, gym_id, subscription_type, payment_method, payment_proof, payment_status, start_date, end_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    self.member_id,
                    self.gym_id,
                    sub_type,
                    payment_method,
                    proof_data,
                    "pending" if payment_method == "rib" else "confirmed",
                    start_date,
                    end_date
                ))
        else:
            # New subscription - create new entry
            success = DatabaseManager.execute_query("""
                INSERT INTO subscriptions
                (member_id, gym_id, subscription_type, payment_method, payment_proof, payment_status, start_date, end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                self.member_id,
                self.gym_id,
                sub_type,
                payment_method,
                proof_data,
                "pending" if payment_method == "rib" else "confirmed",
                start_date,
                end_date
            ))
        
        if success:
            action_text = "renouvellement" if self.is_renewal else "abonnement"
            if payment_method == "rib":
                QMessageBox.information(
                    self,
                    "Succès",
                    f"Votre demande de {action_text} a été envoyée. "
                    "Le propriétaire doit confirmer votre paiement."
                )
            else:
                QMessageBox.information(
                    self,
                    "Succès",
                    f"Votre {action_text} est confirmé. "
                    "Veuillez effectuer le paiement en espèces à la salle."
                )
            
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de l'enregistrement")

class PaymentConfirmationDialog(BaseDialog):
    def __init__(self, subscription_id, parent=None):
        self.subscription_id = subscription_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("💳 Confirmer le paiement")
        self.setFixedSize(800, 700)

        # Modern dark theme styling
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2c2c2c, stop:1 #1a1a1a);
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.3);
            }
            QLabel {
                color: white;
                background: transparent;
            }
            QPushButton {
                font-weight: 600;
                padding: 12px 20px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # Load subscription info
        self.subscription = self.load_subscription()
        if not self.subscription:
            self.reject()
            return

        # Member and gym info
        member = self.load_member_info()
        gym = self.load_gym_info()

        # Modern header
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 140, 0, 0.2),
                    stop:1 rgba(255, 107, 53, 0.2));
                border-radius: 15px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("💳 CONFIRMATION DE PAIEMENT")
        title.setStyleSheet("""
            font-size: 20px;
            font-weight: 800;
            color: white;
            text-align: center;
            letter-spacing: 1px;
            margin-bottom: 15px;
        """)
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)

        # Member info card
        info_container = QWidget()
        info_container.setStyleSheet("""
            background: rgba(255, 255, 255, 0.05);
            border-radius: 10px;
            border: 1px solid rgba(255, 107, 53, 0.2);
        """)
        info_layout = QVBoxLayout(info_container)
        info_layout.setContentsMargins(20, 15, 20, 15)

        member_info = QLabel(f"👤 Membre: {member['first_name']} {member['last_name']}")
        member_info.setStyleSheet("font-size: 16px; font-weight: 600; color: white; margin-bottom: 5px;")

        gym_info = QLabel(f"🏢 Salle: {gym['name']}")
        gym_info.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8); margin-bottom: 5px;")

        sub_type = QLabel(f"📋 Type: {self.subscription['subscription_type']}")
        sub_type.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8); margin-bottom: 5px;")

        payment_method = QLabel(f"💳 Méthode: {'Espèces' if self.subscription['payment_method'] == 'espece' else 'Virement bancaire'}")
        payment_method.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8);")

        info_layout.addWidget(member_info)
        info_layout.addWidget(gym_info)
        info_layout.addWidget(sub_type)
        info_layout.addWidget(payment_method)
        header_layout.addWidget(info_container)
        layout.addWidget(header_container)

        # Payment proof section
        proof_container = QWidget()
        proof_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.8);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        proof_layout = QVBoxLayout(proof_container)
        proof_layout.setContentsMargins(25, 20, 25, 20)

        if self.subscription['payment_method'] == "rib" and self.subscription['payment_proof']:
            proof_header = QLabel("📸 Preuve de paiement")
            proof_header.setStyleSheet("""
                font-size: 18px;
                font-weight: 700;
                color: #ff6b35;
                margin-bottom: 15px;
            """)
            proof_layout.addWidget(proof_header)

            # Create scraollable area for the image
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_area.setStyleSheet("""
                QScrollArea {
                    border: 2px solid rgba(255, 107, 53, 0.3);
                    border-radius: 10px;
                    background: rgba(0, 0, 0, 0.3);
                }
                QScrollBar:vertical, QScrollBar:horizontal {
                    background: rgba(255, 107, 53, 0.1);
                    width: 10px;
                    height: 10px;
                    border-radius: 5px;
                }
                QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
                    background: rgba(255, 107, 53, 0.5);
                    border-radius: 5px;
                    min-height: 20px;
                    min-width: 20px;
                }
            """)

            self.proof_label = QLabel()
            self.proof_label.setAlignment(Qt.AlignCenter)
            self.proof_label.setStyleSheet("""
                background: transparent;
                padding: 10px;
            """)

            try:
                pixmap = QPixmap()
                if pixmap.loadFromData(self.subscription['payment_proof']):
                    # Scale image to fit nicely while maintaining aspect ratio
                    scaled_pixmap = pixmap.scaled(500, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.proof_label.setPixmap(scaled_pixmap)
                else:
                    self.proof_label.setText("❌ Impossible de charger l'image")
                    self.proof_label.setStyleSheet("""
                        color: #ff6b35;
                        font-size: 16px;
                        font-weight: 600;
                        background: transparent;
                        padding: 40px;
                    """)
            except Exception as e:
                print(f"Error loading payment proof: {e}")
                self.proof_label.setText("❌ Erreur lors du chargement de l'image")
                self.proof_label.setStyleSheet("""
                    color: #ff6b35;
                    font-size: 16px;
                    font-weight: 600;
                    background: transparent;
                    padding: 40px;
                """)

            scroll_area.setWidget(self.proof_label)
            proof_layout.addWidget(scroll_area)
        else:
            # No proof case
            no_proof_container = QWidget()
            no_proof_container.setStyleSheet("""
                background: rgba(255, 255, 255, 0.05);
                border-radius: 10px;
                border: 2px dashed rgba(255, 107, 53, 0.3);
            """)
            no_proof_layout = QVBoxLayout(no_proof_container)
            no_proof_layout.setContentsMargins(40, 30, 40, 30)

            no_proof_icon = QLabel("💰")
            no_proof_icon.setStyleSheet("""
                font-size: 48px;
                color: rgba(255, 107, 53, 0.6);
                background: transparent;
            """)
            no_proof_icon.setAlignment(Qt.AlignCenter)

            no_proof_text = QLabel("Paiement en espèces")
            no_proof_text.setStyleSheet("""
                font-size: 18px;
                font-weight: 600;
                color: white;
                background: transparent;
                text-align: center;
            """)
            no_proof_text.setAlignment(Qt.AlignCenter)

            no_proof_subtext = QLabel("À confirmer sur place")
            no_proof_subtext.setStyleSheet("""
                font-size: 14px;
                color: rgba(255, 255, 255, 0.7);
                background: transparent;
                text-align: center;
                margin-top: 5px;
            """)
            no_proof_subtext.setAlignment(Qt.AlignCenter)

            no_proof_layout.addWidget(no_proof_icon)
            no_proof_layout.addWidget(no_proof_text)
            no_proof_layout.addWidget(no_proof_subtext)
            proof_layout.addWidget(no_proof_container)

        layout.addWidget(proof_container)
        
        # Modern action buttons (for owner to confirm/reject)
        if self.subscription['payment_status'] == "pending":
            buttons_container = QWidget()
            buttons_container.setStyleSheet("background: transparent;")
            btn_layout = QHBoxLayout(buttons_container)
            btn_layout.setContentsMargins(0, 0, 0, 0)
            btn_layout.setSpacing(15)

            confirm_btn = QPushButton("✅ CONFIRMER LE PAIEMENT")
            confirm_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #28a745, stop:1 #20c997);
                    color: white;
                    font-weight: 700;
                    padding: 15px 25px;
                    border-radius: 12px;
                    font-size: 14px;
                    border: none;
                    letter-spacing: 0.5px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #218838, stop:1 #1e7e34);
                    transform: translateY(-2px);
                }
            """)
            confirm_btn.clicked.connect(lambda: self.update_status("confirmed"))

            reject_btn = QPushButton("❌ REJETER LE PAIEMENT")
            reject_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #dc3545, stop:1 #c82333);
                    color: white;
                    font-weight: 700;
                    padding: 15px 25px;
                    border-radius: 12px;
                    font-size: 14px;
                    border: none;
                    letter-spacing: 0.5px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #c82333, stop:1 #bd2130);
                    transform: translateY(-2px);
                }
            """)
            reject_btn.clicked.connect(lambda: self.update_status("rejected"))

            btn_layout.addWidget(confirm_btn)
            btn_layout.addWidget(reject_btn)
            layout.addWidget(buttons_container)
        else:
            # Status display for already processed payments
            status_container = QWidget()
            status_container.setStyleSheet("""
                background: rgba(255, 255, 255, 0.05);
                border-radius: 10px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            """)
            status_layout = QVBoxLayout(status_container)
            status_layout.setContentsMargins(20, 15, 20, 15)

            status_text = "✅ Paiement confirmé" if self.subscription['payment_status'] == "confirmed" else "❌ Paiement rejeté"
            status_color = "#28a745" if self.subscription['payment_status'] == "confirmed" else "#dc3545"

            status_label = QLabel(status_text)
            status_label.setStyleSheet(f"""
                font-size: 16px;
                font-weight: 700;
                color: {status_color};
                background: transparent;
                text-align: center;
            """)
            status_label.setAlignment(Qt.AlignCenter)
            status_layout.addWidget(status_label)
            layout.addWidget(status_container)
        
        # Close button
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)
    
    def load_subscription(self):
        return DatabaseManager.execute_query(
            "SELECT * FROM subscriptions WHERE id = %s", 
            (self.subscription_id,), 
            fetch_one=True
        )
    
    def load_member_info(self):
        result = DatabaseManager.execute_query(
            "SELECT first_name, last_name FROM users WHERE id = %s", 
            (self.subscription['member_id'],), 
            fetch_one=True
        )
        return result or {"first_name": "Inconnu", "last_name": ""}
    
    def load_gym_info(self):
        result = DatabaseManager.execute_query(
            "SELECT name FROM gyms WHERE id = %s", 
            (self.subscription['gym_id'],), 
            fetch_one=True
        )
        return result or {"name": "Inconnu"}
    
    def update_status(self, status):
        success = DatabaseManager.execute_query("""
            UPDATE subscriptions 
            SET payment_status = %s 
            WHERE id = %s
        """, (status, self.subscription_id))
        
        if success:
            if status == "confirmed":
                QMessageBox.information(self, "Succès", "Paiement confirmé avec succès!")
            else:
                QMessageBox.information(self, "Succès", "Paiement rejeté.")
            
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de la mise à jour")

class GymSelectionDialog(BaseDialog):
    def __init__(self, user_id, title="Choisir une salle", parent=None):
        self.user_id = user_id
        self.title = title
        self.selected_gym_id = None
        super().__init__(parent)

    def setup_ui(self):
        self.setWindowTitle(self.title)
        self.setFixedSize(500, 350)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # Modern title with icon
        title_label = QLabel(f"🏋️ {self.title}")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 22px;
                font-weight: 700;
                color: white;
                background: transparent;
                border: none;
                letter-spacing: 1px;
                margin-bottom: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Modern description
        desc_label = QLabel("Sélectionnez la salle de sport à laquelle vous souhaitez accéder:")
        desc_label.setStyleSheet("""
            QLabel {
                font-size: 15px;
                color: rgba(255, 255, 255, 0.8);
                background: transparent;
                border: none;
                font-weight: 500;
                margin-bottom: 15px;
            }
        """)
        desc_label.setWordWrap(True)
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)

        # Modern gym selection with enhanced styling
        self.gym_combo = QComboBox()
        self.gym_combo.setMinimumHeight(50)
        layout.addWidget(self.gym_combo)

        # Load gyms
        self.load_user_gyms()

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        cancel_btn = QPushButton("Annuler")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f5f5f5;
                color: #333;
                padding: 10px 20px;
                border-radius: 8px;
                border: 1px solid #ddd;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)
        cancel_btn.clicked.connect(self.reject)

        select_btn = QPushButton("Sélectionner")
        select_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                padding: 10px 20px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        select_btn.clicked.connect(self.select_gym)

        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(select_btn)

        layout.addLayout(button_layout)

    def load_user_gyms(self):
        """Load all gyms the user is subscribed to"""
        gyms = DatabaseManager.execute_query("""
            SELECT g.id, g.name, g.address
            FROM gyms g
            JOIN subscriptions s ON g.id = s.gym_id
            WHERE s.member_id = %s AND s.payment_status = 'confirmed'
            ORDER BY g.name
        """, (self.user_id,), fetch_all=True)

        if not gyms:
            # No gyms found
            self.gym_combo.addItem("Aucune salle disponible", None)
            self.gym_combo.setEnabled(False)
        else:
            for gym in gyms:
                display_text = f"{gym['name']} - {gym['address']}"
                self.gym_combo.addItem(display_text, gym['id'])

    def select_gym(self):
        """Handle gym selection"""
        if self.gym_combo.currentData() is None:
            QMessageBox.warning(self, "Erreur", "Aucune salle disponible")
            return

        self.selected_gym_id = self.gym_combo.currentData()
        self.accept()

    def get_selected_gym_id(self):
        """Return the selected gym ID"""
        return self.selected_gym_id

class AddGymPage(QWidget):
    def __init__(self, owner_id, parent=None):
        super().__init__(parent)
        self.owner_id = owner_id
        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet("background-color: white;")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(40, 30, 40, 30)
        main_layout.setSpacing(20)

        # Scroll Area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        content_widget = QWidget()
        self.layout = QVBoxLayout(content_widget)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(20)
        scroll.setWidget(content_widget)
        main_layout.addWidget(scroll)

        # Title Section
        title_container = QWidget()
        title_container.setStyleSheet("background-color: #002347; border-radius: 10px;")
        title_layout = QVBoxLayout(title_container)
        title_layout.setContentsMargins(20, 15, 20, 15)
        
        title = QLabel("Ajouter votre salle de sport")
        title.setStyleSheet("""
            font-size: 22px;
            color: white;
            font-weight: bold;
        """)
        subtitle = QLabel("Remplissez les détails de votre établissement")
        subtitle.setStyleSheet("font-size: 14px; color: #FFC107;")
        
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        self.layout.addWidget(title_container)

        # Form Container
        form_container = QWidget()
        form_container.setStyleSheet("""
            background-color: #FAFAFA;
            border-radius: 10px;
            border: 1px solid #E0E0E0;
        """)
        form_layout = QVBoxLayout(form_container)
        form_layout.setContentsMargins(25, 25, 25, 25)
        form_layout.setSpacing(20)

        # Gym Name
        name_group = QWidget()
        name_layout = QVBoxLayout(name_group)
        name_layout.setSpacing(8)
        
        name_label = QLabel("Nom de la salle*")
        name_label.setStyleSheet("font-size: 15px; color: #002347; font-weight: 500;")
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Ex: Fitness World")
        self.name_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #DDD;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FFC107;
            }
        """)
        name_layout.addWidget(name_label)
        name_layout.addWidget(self.name_input)
        form_layout.addWidget(name_group)

        # Address
        address_group = QWidget()
        address_layout = QVBoxLayout(address_group)
        address_layout.setSpacing(8)
        
        address_label = QLabel("Adresse complète*")
        address_label.setStyleSheet("font-size: 15px; color: #002347; font-weight: 500;")
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Ex: 123 Rue du Sport, 75000 Paris")
        self.address_input.setStyleSheet(self.name_input.styleSheet())
        address_layout.addWidget(address_label)
        address_layout.addWidget(self.address_input)
        form_layout.addWidget(address_group)

        # Subscription Options
        sub_container = QWidget()
        sub_container.setStyleSheet("""
            background-color: white;
            border-radius: 8px;
            border: 1px solid #E0E0E0;
        """)
        sub_layout = QVBoxLayout(sub_container)
        sub_layout.setContentsMargins(15, 15, 15, 15)
        sub_layout.setSpacing(15)

        sub_title = QLabel("Options d'abonnement*")
        sub_title.setStyleSheet("""
            font-size: 16px;
            color: #002347;
            font-weight: bold;
            margin-bottom: 5px;
        """)
        sub_layout.addWidget(sub_title)

        # Subscription Grid
        grid = QGridLayout()
        grid.setVerticalSpacing(15)
        grid.setHorizontalSpacing(20)

        # Monthly Subscription
        monthly_check = QCheckBox("Mensuel")
        monthly_check.setChecked(True)
        monthly_check.setStyleSheet("font-size: 14px; color: #333;")
        self.monthly_price = QLineEdit()
        self.monthly_price.setPlaceholderText("€")
        self.monthly_price.setValidator(QDoubleValidator(0, 999, 2))
        self.monthly_price.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #DDD;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FFC107;
            }
        """)
        self.monthly_price.setFixedWidth(120)
        
        grid.addWidget(monthly_check, 0, 0)
        grid.addWidget(QLabel("Prix:"), 0, 1, Qt.AlignRight)
        grid.addWidget(self.monthly_price, 0, 2)

        # Quarterly Subscription
        quarterly_check = QCheckBox("Trimestriel")
        quarterly_check.setStyleSheet("font-size: 14px; color: #333;")
        self.quarterly_price = QLineEdit()
        self.quarterly_price.setPlaceholderText("€")
        self.quarterly_price.setValidator(QDoubleValidator(0, 999, 2))
        self.quarterly_price.setStyleSheet(self.monthly_price.styleSheet())
        self.quarterly_price.setFixedWidth(120)
        
        grid.addWidget(quarterly_check, 1, 0)
        grid.addWidget(QLabel("Prix:"), 1, 1, Qt.AlignRight)
        grid.addWidget(self.quarterly_price, 1, 2)

        # Annual Subscription
        annual_check = QCheckBox("Annuel")
        annual_check.setStyleSheet("font-size: 14px; color: #333;")
        self.annual_price = QLineEdit()
        self.annual_price.setPlaceholderText("€")
        self.annual_price.setValidator(QDoubleValidator(0, 999, 2))
        self.annual_price.setStyleSheet(self.monthly_price.styleSheet())
        self.annual_price.setFixedWidth(120)
        
        grid.addWidget(annual_check, 2, 0)
        grid.addWidget(QLabel("Prix:"), 2, 1, Qt.AlignRight)
        grid.addWidget(self.annual_price, 2, 2)

        sub_layout.addLayout(grid)
        form_layout.addWidget(sub_container)

        self.layout.addWidget(form_container)

        # Button Container
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 10, 0, 0)
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #F5F5F5;
                color: #002347;
                font-size: 15px;
                font-weight: bold;
                padding: 12px 25px;
                border-radius: 8px;
                border: 1px solid #DDD;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
            }
        """)
        cancel_btn.clicked.connect(lambda: self.parent().show_dashboard() if self.parent() else None)
        
        submit_btn = QPushButton("Enregistrer")
        submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                font-size: 15px;
                font-weight: bold;
                padding: 12px 25px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        submit_btn.clicked.connect(self.save_gym)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(submit_btn)
        self.layout.addWidget(button_container)

        self.layout.addStretch()

    def save_gym(self):
        name = self.name_input.text().strip()
        address = self.address_input.text().strip()
        
        if not name:
            self.show_error("Le nom de la salle est obligatoire", self.name_input)
            return
            
        if not address:
            self.show_error("L'adresse est obligatoire", self.address_input)
            return
        
        # Collect subscription options
        subscriptions = {}
        try:
            if self.monthly_price.text().strip():
                subscriptions["monthly"] = {
                    "duration": "1 mois",
                    "price": float(self.monthly_price.text().strip())
                }
                
            if self.quarterly_price.text().strip():
                subscriptions["quarterly"] = {
                    "duration": "3 mois",
                    "price": float(self.quarterly_price.text().strip())
                }
                
            if self.annual_price.text().strip():
                subscriptions["annual"] = {
                    "duration": "12 mois",
                    "price": float(self.annual_price.text().strip())
                }
                
            if not subscriptions:
                QMessageBox.warning(self, "Erreur", "Veuillez entrer au moins une option d'abonnement")
                return
                
            # Save to database
            success = DatabaseManager.execute_query("""
                INSERT INTO gyms (name, address, subscriptions, owner_id)
                VALUES (%s, %s, %s, %s)
            """, (
                name,
                address,
                json.dumps(subscriptions),
                self.owner_id
            ))
            
            if success:
                QMessageBox.information(
                    self, 
                    "Succès", 
                    f"Votre salle {name} a été enregistrée avec succès!"
                )
                if self.parent():
                    self.parent().show_dashboard()
            else:
                QMessageBox.critical(self, "Erreur", "Erreur lors de l'enregistrement: Database error")
                
        except ValueError:
            QMessageBox.warning(self, "Erreur", "Veuillez entrer des prix valides")

    def show_error(self, message, widget=None):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Champ requis")
        msg.setText(message)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: white;
            }
            QMessageBox QLabel {
                color: #002347;
                font-size: 14px;
            }
        """)
        msg.exec()
        
        if widget:
            widget.setStyleSheet("""
                QLineEdit {
                    background-color: white;
                    border: 2px solid #FF0000;
                    border-radius: 8px;
                    padding: 12px;
                    font-size: 14px;
                }
            """)
            QTimer.singleShot(2000, lambda: widget.setStyleSheet("""
                QLineEdit {
                    background-color: white;
                    border: 1px solid #DDD;
                    border-radius: 8px;
                    padding: 12px;
                    font-size: 14px;
                }
                QLineEdit:focus {
                    border: 2px solid #FFC107;
                }
            """))

class BaseWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        
    def setup_ui(self):
        raise NotImplementedError("Subclasses must implement setup_ui method")

class SignupWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        
    def setup_ui(self):
        self.setWindowTitle("UaFit Signup")
        self.setFixedSize(800, 400)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Left section
        left_widget = QWidget()
        left_widget.setStyleSheet("background-color: #002347;")
        left_layout = QVBoxLayout(left_widget)
        left_layout.setAlignment(Qt.AlignCenter)

        ua_fit_widget = QWidget()
        ua_fit_layout = QHBoxLayout(ua_fit_widget)
        ua_fit_layout.setAlignment(Qt.AlignCenter)

        ua_label = QLabel("Ua")
        ua_label.setStyleSheet("font-size: 48px; color: #FFC107; font-weight: bold;")
        fit_label = QLabel("Fit")
        fit_label.setStyleSheet("font-size: 48px; color: white; font-weight: bold;")
        ua_fit_layout.addWidget(ua_label)
        ua_fit_layout.addWidget(fit_label)

        slogan_label = QLabel("Votre porte d'entrée vers une gestion sportive simplifiée")
        slogan_label.setStyleSheet("font-size: 14px; color: white; margin-bottom: 40px; margin-top: -5px;")
        
        left_layout.addStretch()
        left_layout.addWidget(ua_fit_widget)
        left_layout.addWidget(slogan_label)
        left_layout.addStretch()
        left_widget.setFixedWidth(400)

        # Right section
        right_widget = QWidget()
        right_widget.setStyleSheet("background-color: white;")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(20, 20, 20, 20)
        
        self.title_widget = QWidget()
        title_layout = QHBoxLayout(self.title_widget)
        title_layout.setContentsMargins(0, 0, 0, 20)
        
        self.signup_title = QLabel("SIGN UP (1/3)")
        self.signup_title.setStyleSheet("font-size: 20px; color: #002347; font-weight: bold;")
        
        self.progress_dots = QWidget()
        dots_layout = QHBoxLayout(self.progress_dots)
        dots_layout.setSpacing(10)
        
        self.dot1 = QLabel("...")
        self.dot1.setStyleSheet("font-size: error; color: #FFC107;")
        self.dot2 = QLabel("...")
        self.dot2.setStyleSheet("font-size: error; color: #CCCCCC;")
        self.dot3 = QLabel("...")
        self.dot3.setStyleSheet("font-size: error; color: #CCCCCC;")
        
        dots_layout.addStretch()
        dots_layout.addWidget(self.dot1)
        dots_layout.addWidget(self.dot2)
        dots_layout.addWidget(self.dot3)
        dots_layout.addStretch()
        
        title_layout.addWidget(self.signup_title)
        title_layout.addStretch()
        title_layout.addWidget(self.progress_dots)
        
        right_layout.addWidget(self.title_widget)
        
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.setFixedHeight(300)
        right_layout.addWidget(self.stacked_widget)
        
        self.nav_widget = QFrame()
        nav_layout = QHBoxLayout(self.nav_widget)
        nav_layout.setContentsMargins(0, 20, 0, 0)
        
        self.back_button = QPushButton("BACK")
        self.back_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #002347;
                border: 2px solid #002347;
                border-radius: 15px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #002347;
                color: white;
            }
        """)
        self.back_button.setFixedWidth(120)
        self.back_button.setVisible(False)
        self.back_button.clicked.connect(self.prev_page)
        
        self.next_button = QPushButton("NEXT")
        self.next_button.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                border-radius: 15px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        self.next_button.setFixedWidth(120)
        self.next_button.clicked.connect(self.next_page)
        
        nav_layout.addWidget(self.back_button)
        nav_layout.addStretch()
        nav_layout.addWidget(self.next_button)
        
        right_layout.addWidget(self.nav_widget)
        
        self.create_personal_info_page()
        self.create_account_info_page()
        self.create_membership_page()
        
        main_layout.addWidget(left_widget)
        main_layout.addWidget(right_widget)
        
        self.current_page = 0
        self.update_navigation_ui()
    
    def create_personal_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        
        title = QLabel("Personal Information")
        title.setStyleSheet("font-size: 18px; color: #002347; margin-bottom: 20px;")
        layout.addWidget(title)
        
        name_layout = QHBoxLayout()
        name_layout.setSpacing(10)
        
        self.first_name_input = QLineEdit()
        self.first_name_input.setPlaceholderText("First Name")
        self.first_name_input.setStyleSheet("""
            QLineEdit {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QLineEdit:focus {
                border-color: #FFC107;
            }
        """)
        self.first_name_input.setFixedHeight(40)
        
        self.last_name_input = QLineEdit()
        self.last_name_input.setPlaceholderText("Last Name")
        self.last_name_input.setStyleSheet(self.first_name_input.styleSheet())
        self.first_name_input.setFixedHeight(40)
        
        name_layout.addWidget(self.first_name_input)
        name_layout.addWidget(self.last_name_input)
        layout.addLayout(name_layout)
        
        age_gender_layout = QHBoxLayout()
        age_gender_layout.setSpacing(10)
        
        self.age_input = QLineEdit()
        self.age_input.setPlaceholderText("Age")
        self.age_input.setStyleSheet(self.first_name_input.styleSheet())
        self.age_input.setFixedHeight(40)
        
        self.gender_input = QComboBox()
        self.gender_input.addItems(["Select Gender", "Male", "Female"])
        self.gender_input.setStyleSheet("""
            QComboBox {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QComboBox:focus {
                border-color: #FFC107;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                width: 0px;
                height: 0px;
            }
        """)
        self.gender_input.setFixedHeight(40)

        age_gender_layout.addWidget(self.age_input)
        age_gender_layout.addWidget(self.gender_input)
        layout.addLayout(age_gender_layout)
        
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Phone Number")
        self.phone_input.setStyleSheet(self.first_name_input.styleSheet())
        self.phone_input.setFixedHeight(40)
        layout.addWidget(self.phone_input)
        
        layout.addStretch()
        self.stacked_widget.addWidget(page)
    
    def create_account_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        
        title = QLabel("Account Information")
        title.setStyleSheet("font-size: 18px; color: #002347; margin-bottom: 20px;")
        layout.addWidget(title)
        
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Email")
        self.email_input.setStyleSheet("""
            QLineEdit {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QLineEdit:focus {
                border-color: #FFC107;
            }
        """)
        self.email_input.setFixedHeight(40)
        layout.addWidget(self.email_input)
        
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Password")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setStyleSheet(self.email_input.styleSheet())
        self.password_input.setFixedHeight(40)
        layout.addWidget(self.password_input)
        
        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setPlaceholderText("Confirm Password")
        self.confirm_password_input.setEchoMode(QLineEdit.Password)
        self.confirm_password_input.setStyleSheet(self.email_input.styleSheet())
        self.confirm_password_input.setFixedHeight(40)
        layout.addWidget(self.confirm_password_input)
        
        layout.addStretch()
        self.stacked_widget.addWidget(page)
    
    def create_membership_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        
        title = QLabel("Membership Details")
        title.setStyleSheet("font-size: 18px; color: #002347; margin-bottom: 20px;")
        layout.addWidget(title)
        
        self.membership_type_input = QComboBox()
        self.membership_type_input.addItems(["Select Role", "Member", "Owner"])
        self.membership_type_input.setStyleSheet("""
            QComboBox {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QComboBox:focus {
                border-color: #FFC107;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                width: 0px;
                height: 0px;
            }
        """)
        self.membership_type_input.setFixedHeight(40)
        layout.addWidget(self.membership_type_input)
        
        self.emergency_phone_input = QLineEdit()
        self.emergency_phone_input.setPlaceholderText("Emergency Contact Phone")
        self.emergency_phone_input.setStyleSheet("""
            QLineEdit {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QLineEdit:focus {
                border-color: #FFC107;
            }
        """)
        self.emergency_phone_input.setFixedHeight(40)
        layout.addWidget(self.emergency_phone_input)

        self.address = QLineEdit()
        self.address.setPlaceholderText("Address")
        self.address.setStyleSheet("""
            QLineEdit {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QLineEdit:focus {
                border-color: #FFC107;
            }
        """)
        self.address.setFixedHeight(40)
        layout.addWidget(self.address)
        
        layout.addStretch()
        self.stacked_widget.addWidget(page)
    
    def next_page(self):
        if self.current_page == 2:
            self.submit_form()
            return
            
        if not self.validate_current_page():
            return
            
        animation = QPropertyAnimation(self.stacked_widget, b"pos")
        animation.setDuration(300)
        animation.setEasingCurve(QEasingCurve.OutCubic)
        animation.setStartValue(self.stacked_widget.pos())
        animation.setEndValue(self.stacked_widget.pos() - self.stacked_widget.rect().topRight())
        animation.start()
        
        self.current_page += 1
        self.stacked_widget.setCurrentIndex(self.current_page)
        
        self.update_navigation_ui()
        
        self.stacked_widget.move(self.stacked_widget.pos() + self.stacked_widget.rect().topRight())
    
    def prev_page(self):
        animation = QPropertyAnimation(self.stacked_widget, b"pos")
        animation.setDuration(300)
        animation.setEasingCurve(QEasingCurve.OutCubic)
        animation.setStartValue(self.stacked_widget.pos())
        animation.setEndValue(self.stacked_widget.pos() + self.stacked_widget.rect().topRight())
        animation.start()
        
        self.current_page -= 1
        self.stacked_widget.setCurrentIndex(self.current_page)
        
        self.update_navigation_ui()
        
        self.stacked_widget.move(self.stacked_widget.pos() - self.stacked_widget.rect().topRight())
    
    def update_navigation_ui(self):
        self.signup_title.setText(f"SIGN UP ({self.current_page + 1}/3)")
        
        self.dot1.setText("○")
        self.dot2.setText("○")
        self.dot3.setText("○")
        
        if self.current_page == 0:
            self.dot1.setText("●")
            self.back_button.setVisible(False)
            self.next_button.setText("NEXT")
        elif self.current_page == 1:
            self.dot2.setText("●")
            self.back_button.setVisible(True)
            self.next_button.setText("NEXT")
        else:
            self.dot3.setText("●")
            self.back_button.setVisible(True)
            self.next_button.setText("SIGN UP")
    
    def validate_current_page(self):
        if self.current_page == 0:
            if not self.first_name_input.text() or not self.last_name_input.text():
                QMessageBox.warning(self, "Validation Error", "Please enter your first and last name")
                return False
            if not self.age_input.text().isdigit() or int(self.age_input.text()) < 16:
                QMessageBox.warning(self, "Validation Error", "Please enter a valid age (16+)")
                return False
        elif self.current_page == 1:
            if not self.email_input.text() or "@" not in self.email_input.text():
                QMessageBox.warning(self, "Validation Error", "Please enter a valid email address")
                return False
            if not self.password_input.text() or len(self.password_input.text()) < 6:
                QMessageBox.warning(self, "Validation Error", "Password must be at least 6 characters")
                return False
            if self.password_input.text() != self.confirm_password_input.text():
                QMessageBox.warning(self, "Validation Error", "Passwords do not match")
                return False
        return True
    
    def submit_form(self):
        if not self.validate_current_page():
            return
            
        if self.membership_type_input.currentText() == "Select Role":
            QMessageBox.warning(self, "Validation Error", "Please select a membership type")
            return
            
        role = "gym owner" if self.membership_type_input.currentText().lower() == "owner" else "member"
        
        success = DatabaseManager.execute_query("""
            INSERT INTO users (first_name, last_name, age, gender, phone, email, password, role, address)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            self.first_name_input.text(),
            self.last_name_input.text(),
            self.age_input.text(),
            self.gender_input.currentText(),
            self.phone_input.text(),
            self.email_input.text(),
            self.password_input.text(),  # In production, hash this password
            role,
            self.address.text()
        ))
        
        if success:
            QMessageBox.information(self, "Success", "Account created successfully!")
            
            # Redirect to login page with pre-filled credentials
            self.login_window = LoginWindow()
            self.login_window.email_input.setText(self.email_input.text())
            self.login_window.password_input.setText(self.password_input.text())
            self.login_window.show()
            self.close()
        else:
            QMessageBox.critical(self, "Database Error", "Failed to create account")
class MemberWorkoutHistoryDialog(BaseDialog):
    def __init__(self, owner_id, parent=None):
        self.owner_id = owner_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("👥 Historique des membres")
        self.setFixedSize(1000, 700)

        # Modern dark theme styling
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2c2c2c, stop:1 #1a1a1a);
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.3);
            }
            QLabel {
                color: white;
                background: transparent;
                font-weight: 500;
            }
            QComboBox {
                background: rgba(35, 35, 35, 0.8);
                border: 2px solid rgba(255, 107, 53, 0.3);
                border-radius: 10px;
                padding: 12px 15px;
                font-size: 14px;
                color: white;
                font-weight: 500;
                min-height: 20px;
            }
            QComboBox:focus {
                border: 2px solid rgba(255, 107, 53, 0.6);
                background: rgba(35, 35, 35, 0.9);
            }
            QComboBox::drop-down {
                border: none;
                width: 25px;
                background: rgba(255, 107, 53, 0.2);
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 2px solid white;
                border-top: none;
                border-right: none;
                width: 8px;
                height: 8px;
                transform: rotate(-45deg);
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background: rgba(35, 35, 35, 0.95);
                border: 2px solid rgba(255, 107, 53, 0.4);
                border-radius: 8px;
                color: white;
                selection-background-color: rgba(255, 107, 53, 0.3);
                padding: 5px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # Modern title
        title = QLabel("👥 HISTORIQUE DES MEMBRES")
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: 800;
            color: #ff6b35;
            text-align: center;
            margin-bottom: 20px;
            letter-spacing: 1px;
        """)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Selection container
        selection_container = QWidget()
        selection_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.8);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        selection_layout = QVBoxLayout(selection_container)
        selection_layout.setContentsMargins(25, 20, 25, 20)
        selection_layout.setSpacing(20)

        # Gym selection
        gym_section = QWidget()
        gym_section.setStyleSheet("background: transparent;")
        gym_section_layout = QVBoxLayout(gym_section)
        gym_section_layout.setContentsMargins(0, 0, 0, 0)
        gym_section_layout.setSpacing(8)

        gym_label = QLabel("🏢 Sélectionnez une salle:")
        gym_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #ff6b35;
            background: transparent;
        """)
        gym_section_layout.addWidget(gym_label)

        self.gym_combo = QComboBox()
        self.gym_combo.currentIndexChanged.connect(self.load_members)
        gym_section_layout.addWidget(self.gym_combo)

        selection_layout.addWidget(gym_section)

        # Member selection
        member_section = QWidget()
        member_section.setStyleSheet("background: transparent;")
        member_section_layout = QVBoxLayout(member_section)
        member_section_layout.setContentsMargins(0, 0, 0, 0)
        member_section_layout.setSpacing(8)

        member_label = QLabel("👤 Sélectionnez un membre:")
        member_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #ff6b35;
            background: transparent;
        """)
        member_section_layout.addWidget(member_label)

        self.member_combo = QComboBox()
        self.member_combo.currentIndexChanged.connect(self.show_history)
        member_section_layout.addWidget(self.member_combo)

        selection_layout.addWidget(member_section)
        layout.addWidget(selection_container)

        # History container
        history_container = QWidget()
        history_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.6);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        history_layout = QVBoxLayout(history_container)
        history_layout.setContentsMargins(20, 15, 20, 15)

        # Placeholder for history
        self.history_placeholder = QLabel("📊 Sélectionnez un membre pour voir son historique d'entraînement")
        self.history_placeholder.setStyleSheet("""
            font-size: 16px;
            color: rgba(255, 255, 255, 0.6);
            background: transparent;
            padding: 40px;
            text-align: center;
        """)
        self.history_placeholder.setAlignment(Qt.AlignCenter)
        history_layout.addWidget(self.history_placeholder)

        # History widget (initially hidden)
        self.history_widget = WorkoutHistoryDialog(None, None, self)
        self.history_widget.hide()
        history_layout.addWidget(self.history_widget)

        layout.addWidget(history_container)

        # Load gyms
        self.load_gyms()
    
    def load_gyms(self):
        self.gym_combo.clear()
        self.gym_combo.addItem("-- Choisissez une salle --", None)

        gyms = DatabaseManager.execute_query(
            "SELECT id, name FROM gyms WHERE owner_id = %s",
            (self.owner_id,),
            fetch_all=True
        )

        if gyms:
            for gym in gyms:
                self.gym_combo.addItem(f"🏢 {gym['name']}", gym['id'])
        else:
            self.gym_combo.addItem("Aucune salle trouvée", None)

    def load_members(self):
        self.member_combo.clear()
        self.member_combo.addItem("-- Choisissez un membre --", None)

        gym_id = self.gym_combo.currentData()

        if gym_id:
            members = DatabaseManager.get_gym_members(gym_id)

            if members:
                for member in members:
                    self.member_combo.addItem(
                        f"👤 {member['first_name']} {member['last_name']}",
                        (member['id'], gym_id)
                    )
            else:
                self.member_combo.addItem("Aucun membre trouvé", None)

        # Hide history when gym changes
        self.history_widget.hide()
        self.history_placeholder.show()

    def show_history(self):
        print(f"🔍 show_history called - combo index: {self.member_combo.currentIndex()}")
        if self.member_combo.currentIndex() > 0 and self.member_combo.currentData():
            try:
                member_data = self.member_combo.currentData()
                print(f"📊 Member data: {member_data}")
                if isinstance(member_data, tuple) and len(member_data) == 2:
                    member_id, gym_id = member_data
                    print(f"👤 Loading history for member_id: {member_id}, gym_id: {gym_id}")

                    # Update history widget
                    self.history_widget.member_id = member_id
                    self.history_widget.gym_id = gym_id
                    self.history_widget.current_week = QDate.currentDate()
                    self.history_widget.update_week_display()

                    # Show history widget and hide placeholder
                    self.history_placeholder.hide()
                    self.history_widget.show()
                    print("✅ History widget shown successfully")
                else:
                    print("❌ Invalid member data format")
                    # Invalid data, hide history
                    self.history_widget.hide()
                    self.history_placeholder.show()
            except Exception as e:
                print(f"❌ Error showing history: {e}")
                import traceback
                traceback.print_exc()
                self.history_widget.hide()
                self.history_placeholder.show()
        else:
            print("❌ No valid member selection")
            # No valid selection, hide history
            self.history_widget.hide()
            self.history_placeholder.show()

class DashboardWindow(BaseWindow):

    
    

    def track_workout(self):
        """Allow members to track workouts at their subscribed gyms"""
        # Use gym selection dialog for consistency
        gym_dialog = GymSelectionDialog(self.user_id, "Choisir une salle pour l'entraînement", self)
        if gym_dialog.exec() == QDialog.Accepted:
            gym_id = gym_dialog.get_selected_gym_id()
            if gym_id:
                dialog = WorkoutTrackingDialog(self.user_id, gym_id, self)
                dialog.exec()

    def show_workout_history(self):
        # Get all gyms the member is subscribed to
        gyms = DatabaseManager.execute_query("""
            SELECT g.id, g.name FROM gyms g
            JOIN subscriptions s ON g.id = s.gym_id
            WHERE s.member_id = %s AND s.payment_status = 'confirmed'
            ORDER BY g.name
        """, (self.user_id,), fetch_all=True)

        if not gyms:
            QMessageBox.warning(self, "Erreur", "Vous n'avez pas d'abonnement actif")
            return
        elif len(gyms) == 1:
            # If only one gym, use it directly
            dialog = WorkoutHistoryDialog(self.user_id, gyms[0]['id'], self)
            dialog.exec()
        else:
            # Multiple gyms - let user choose
            selection_dialog = GymSelectionDialog(
                self.user_id,
                "Choisir une salle pour l'historique",
                self
            )
            if selection_dialog.exec() == QDialog.Accepted:
                selected_gym_id = selection_dialog.get_selected_gym_id()
                if selected_gym_id:
                    dialog = WorkoutHistoryDialog(self.user_id, selected_gym_id, self)
                    dialog.exec()

    def rate_gym(self):
        """Allow members to rate gyms they are subscribed to"""
        # Use gym selection dialog to choose which gym to rate
        gym_dialog = GymSelectionDialog(self.user_id, "Choisir une salle à évaluer", self)
        if gym_dialog.exec() == QDialog.Accepted:
            gym_id = gym_dialog.get_selected_gym_id()
            if gym_id:
                # Check if user has already rated this gym
                existing_review = DatabaseManager.execute_query("""
                    SELECT id FROM gym_reviews
                    WHERE gym_id = %s AND member_id = %s
                """, (gym_id, self.user_id), fetch_one=True)

                if existing_review:
                    QMessageBox.information(self, "Information", "Vous avez déjà évalué cette salle")
                    return

                # Open rating dialog
                dialog = AddReviewDialog(gym_id, self.user_id, self)
                dialog.exec()

    def rate_specific_gym(self, gym_id):
        """Rate a specific gym directly (called from gym card)"""
        # Check if user has already rated this gym
        existing_review = DatabaseManager.execute_query("""
            SELECT id FROM gym_reviews
            WHERE gym_id = %s AND member_id = %s
        """, (gym_id, self.user_id), fetch_one=True)

        if existing_review:
            QMessageBox.information(self, "Information", "Vous avez déjà évalué cette salle")
            return

        # Open rating dialog
        dialog = AddReviewDialog(gym_id, self.user_id, self)
        if dialog.exec() == QDialog.Accepted:
            # Refresh the gym cards to update the rating button
            self.load_nearby_gyms()

    def show_planning(self):
        # Get the first gym (for simplicity - you might want to let the owner choose)
        gym = DatabaseManager.execute_query(
            "SELECT id FROM gyms WHERE owner_id = %s LIMIT 1", 
            (self.user_id,), 
            fetch_one=True
        )
        
        if gym:
            dialog = PlanningDialog(gym['id'], self)
            dialog.exec()
        else:
            QMessageBox.warning(self, "Erreur", "Vous devez d'abord créer une salle")

    def __init__(self, user_id, user_name, user_role, address=None):
        self.user_id = user_id
        self.user_name = user_name
        self.user_role = user_role
        self.address = address
        super().__init__()
        
    def setup_ui(self):
        # Modern window title with emoji and role
        role_emoji = "🏢" if self.user_role == "gym owner" else "💪"
        self.setWindowTitle(f"UaFit 2025 {role_emoji} {self.user_name}'s Dashboard")
        self.setFixedSize(1200, 700)  # Better size for all screens

        # Set modern gym-style background with dark theme
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2d2d2d, stop:1 #1a1a1a);
            }
        """)

        # Add modern status bar
        status_bar = self.statusBar()
        status_bar.setStyleSheet("""
            QStatusBar {
                background: rgba(255, 255, 255, 0.1);
                color: rgba(255, 255, 255, 0.8);
                border-top: 1px solid rgba(255, 255, 255, 0.2);
                font-size: 12px;
                font-weight: 500;
            }
        """)
        status_bar.showMessage(f"✨ Bienvenue dans UaFit 2025 - Interface moderne activée | Rôle: {self.user_role.title()}")

        # Set window icon from online source
        try:
            # Use a fitness/gym icon from online source
            import urllib.request
            icon_url = "https://cdn-icons-png.flaticon.com/512/2936/2936719.png"  # Gym dumbbell icon
            icon_data = urllib.request.urlopen(icon_url).read()
            pixmap = QPixmap()
            pixmap.loadFromData(icon_data)
            self.setWindowIcon(QIcon(pixmap))
        except:
            # Fallback to default icon if online loading fails
            pass

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)  # Small margin for modern look
        main_layout.setSpacing(8)

        self.create_modern_sidebar(main_layout)

        # Modern content area with glassmorphism
        self.content_area = QWidget()
        self.content_area.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.05);
                border-radius: 20px;
                border: 1px solid rgba(255, 255, 255, 0.1);
            }
        """)
        self.content_layout = QVBoxLayout(self.content_area)
        self.content_layout.setContentsMargins(25, 25, 25, 25)
        self.content_layout.setSpacing(20)

        self.create_modern_header()

        self.stacked_widget = QStackedWidget()
        self.stacked_widget.setStyleSheet("background: transparent; border: none;")
        self.content_layout.addWidget(self.stacked_widget)

        self.create_dashboard_views()

        main_layout.addWidget(self.content_area, stretch=5)

        self.show_initial_view()

        # Show modern welcome message
        self.show_welcome_message()

    def show_welcome_message(self):
        """Show a modern welcome message when the app starts"""
        from PySide6.QtCore import QTimer

        # Create a timer to show the message after a short delay
        QTimer.singleShot(500, lambda: self.statusBar().showMessage(
            f"🎉 Interface moderne UaFit 2025 chargée avec succès! | Utilisateur: {self.user_name} ({self.user_role.title()})",
            5000  # Show for 5 seconds
        ))

    def get_tooltip_for_button(self, button_text):
        """Get modern tooltip text for navigation buttons"""
        tooltips = {
            "🏢 Mes Salles": "Gérez vos salles de sport et leurs paramètres",
            "💳 Paiements": "Vérifiez et confirmez les paiements en attente",
            "👥 Membres": "Consultez la liste de tous vos membres",
            "📊 Statistiques": "Analysez les performances de vos salles",
            "📋 Historique Membres": "Consultez l'historique d'entraînement des membres",
            "🏋️ Salles à proximité": "Découvrez les salles de sport près de chez vous",
            "💳 Mes Abonnements": "Gérez vos abonnements actifs",
            "📊 Historique": "Consultez votre historique d'entraînement",
            "➕ Pointer entraînement": "Enregistrez votre séance d'entraînement",
            "⭐ Évaluer une salle": "Donnez votre avis sur les salles fréquentées",
            "💬 Chat de la salle": "Discutez avec les autres membres",
            "👤 Profil": "Modifiez vos informations personnelles",
            "🏦 RIB": "Gérez vos coordonnées bancaires",
            "🚪 Déconnexion": "Quittez votre session en toute sécurité"
        }
        return tooltips.get(button_text, "")

    def create_modern_sidebar(self, main_layout):
        # Modern gym-style sidebar with dark theme
        sidebar = QWidget()
        sidebar.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(45, 45, 45, 0.95),
                    stop:0.5 rgba(35, 35, 35, 0.90),
                    stop:1 rgba(25, 25, 25, 0.95));
                border-radius: 15px;
                border: 1px solid rgba(255, 140, 0, 0.3);
                backdrop-filter: blur(10px);
                transition: all 0.3s ease-in-out;
            }
            QWidget:hover {
                border: 1px solid rgba(255, 140, 0, 0.5);
            }
        """)
        sidebar.setFixedWidth(320)  # Wider for modern feel
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setAlignment(Qt.AlignTop)
        sidebar_layout.setContentsMargins(15, 20, 15, 20)
        sidebar_layout.setSpacing(10)

        # Modern Logo with gradient text effect
        logo_container = QWidget()
        logo_container.setStyleSheet("""
            QWidget {
                background: transparent;
                padding: 20px 0px;
            }
        """)
        logo_layout = QVBoxLayout(logo_container)
        logo_layout.setContentsMargins(0, 0, 0, 0)
        logo_layout.setAlignment(Qt.AlignCenter)

        # Modern gym logo with orange accent
        logo_text = QLabel("UaFit")
        logo_text.setStyleSheet("""
            QLabel {
                font-size: 28px;
                font-weight: 800;
                color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff8c00, stop:0.5 #ff6b35, stop:1 #ff8c00);
                background: transparent;
                letter-spacing: 2px;
                transition: all 0.3s ease-in-out;
            }
            QLabel:hover {
                transform: scale(1.05);
                text-shadow: 0 0 15px rgba(255, 140, 0, 0.6);
            }
        """)
        logo_text.setAlignment(Qt.AlignCenter)
        logo_text.setCursor(Qt.PointingHandCursor)

        # Subtitle
        subtitle = QLabel("Fitness Management")
        subtitle.setStyleSheet("""
            QLabel {
                font-size: 11px;
                color: rgba(255, 255, 255, 0.7);
                background: transparent;
                letter-spacing: 1px;
                font-weight: 500;
                margin-top: 5px;
            }
        """)
        subtitle.setAlignment(Qt.AlignCenter)

        logo_layout.addWidget(logo_text)
        logo_layout.addWidget(subtitle)
        sidebar_layout.addWidget(logo_container)

        # Modern Navigation Menu with scroll area
        menu_scroll = QScrollArea()
        menu_scroll.setWidgetResizable(True)
        menu_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 140, 0, 0.1);
                width: 6px;
                border-radius: 3px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 140, 0, 0.5);
                border-radius: 3px;
                min-height: 20px;
            }
        """)

        menu_container = QWidget()
        menu_container.setStyleSheet("background: transparent;")
        menu_layout = QVBoxLayout(menu_container)
        menu_layout.setContentsMargins(0, 5, 0, 5)
        menu_layout.setSpacing(6)

        nav_buttons = self.get_nav_buttons()

        for text, callback, is_active in nav_buttons:
            btn = QPushButton(text)
            btn.setCursor(Qt.PointingHandCursor)

            # Add modern tooltips
            tooltip_text = self.get_tooltip_for_button(text)
            if tooltip_text:
                btn.setToolTip(tooltip_text)

            if is_active:
                btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #ff8c00, stop:0.5 #ff6b35, stop:1 #ff8c00);
                        color: white;
                        font-size: 14px;
                        font-weight: 700;
                        text-align: left;
                        padding: 12px 16px;
                        border-radius: 10px;
                        border: none;
                        letter-spacing: 0.5px;
                        min-height: 40px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #ff7700, stop:0.5 #ff5722, stop:1 #ff7700);
                        transform: translateY(-2px);
                    }
                    QPushButton:pressed {
                        transform: translateY(0px);
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: rgba(45, 45, 45, 0.8);
                        color: white;
                        font-size: 14px;
                        font-weight: 600;
                        text-align: left;
                        padding: 12px 16px;
                        border-radius: 10px;
                        border: 1px solid rgba(255, 140, 0, 0.2);
                        letter-spacing: 0.5px;
                        min-height: 40px;
                    }
                    QPushButton:hover {
                        background: rgba(255, 140, 0, 0.2);
                        border: 1px solid rgba(255, 140, 0, 0.4);
                        color: white;
                        transform: translateX(5px);
                    }
                """)

            btn.clicked.connect(callback)
            menu_layout.addWidget(btn)

        # Add menu container to scroll area
        menu_scroll.setWidget(menu_container)
        sidebar_layout.addWidget(menu_scroll)
        sidebar_layout.addStretch()

        # User Profile Section
        profile_container = QWidget()
        profile_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.08);
                border-radius: 15px;
                border: 1px solid rgba(255, 255, 255, 0.15);
                padding: 15px;
            }
        """)
        profile_layout = QHBoxLayout(profile_container)
        profile_layout.setContentsMargins(0, 0, 0, 0)
        profile_layout.setSpacing(15)

        # Profile picture
        profile_pic_container = QWidget()
        profile_pic_container.setFixedSize(60, 60)
        profile_pic_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #ff8c00, stop:1 #ff6b35);
                border-radius: 30px;
                border: 2px solid rgba(255, 255, 255, 0.3);
            }
        """)

        # Try to load user's profile picture, fallback to icon
        profile_pic_layout = QVBoxLayout(profile_pic_container)
        profile_pic_layout.setContentsMargins(0, 0, 0, 0)
        profile_pic_layout.setAlignment(Qt.AlignCenter)

        try:
            # Try to get user's profile picture from database
            print(f"🖼️ Loading profile picture for user ID: {self.user_id}")
            user_data = DatabaseManager.execute_query("""
                SELECT profile_picture FROM users WHERE id = %s
            """, (self.user_id,), fetch_one=True)

            print(f"🖼️ User data retrieved: {user_data is not None}")
            if user_data:
                print(f"🖼️ Profile picture exists: {user_data['profile_picture'] is not None}")

            if user_data and user_data['profile_picture'] is not None:
                print("🖼️ Loading profile picture from database...")
                # Load profile picture
                profile_pic_label = QLabel()
                pixmap = QPixmap()

                # Check if profile_picture is bytes or string path
                if isinstance(user_data['profile_picture'], bytes):
                    # Load from binary data
                    success = pixmap.loadFromData(user_data['profile_picture'])
                    print(f"🖼️ Loaded from bytes: {success}")
                elif isinstance(user_data['profile_picture'], str) and user_data['profile_picture'].strip():
                    # Load from file path
                    success = pixmap.load(user_data['profile_picture'])
                    print(f"🖼️ Loaded from path '{user_data['profile_picture']}': {success}")
                else:
                    success = False
                    print("🖼️ Profile picture data is neither bytes nor valid path")

                if success and not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    profile_pic_label.setPixmap(scaled_pixmap)
                    profile_pic_label.setStyleSheet("border-radius: 28px; background: transparent;")
                    profile_pic_layout.addWidget(profile_pic_label)
                    print("✅ Profile picture loaded successfully")
                else:
                    print("❌ Failed to load profile picture, using default icon")
                    # Default user icon
                    default_icon = QLabel("👤")
                    default_icon.setStyleSheet("""
                        font-size: 24px;
                        color: white;
                        background: transparent;
                        border: none;
                    """)
                    default_icon.setAlignment(Qt.AlignCenter)
                    profile_pic_layout.addWidget(default_icon)
            else:
                print("🖼️ No profile picture found, using default icon")
                # Default user icon
                default_icon = QLabel("👤")
                default_icon.setStyleSheet("""
                    font-size: 24px;
                    color: white;
                    background: transparent;
                    border: none;
                """)
                default_icon.setAlignment(Qt.AlignCenter)
                profile_pic_layout.addWidget(default_icon)
        except Exception as e:
            print(f"❌ Error loading profile picture: {e}")
            import traceback
            traceback.print_exc()
            # Default user icon
            default_icon = QLabel("👤")
            default_icon.setStyleSheet("""
                font-size: 24px;
                color: white;
                background: transparent;
                border: none;
            """)
            default_icon.setAlignment(Qt.AlignCenter)
            profile_pic_layout.addWidget(default_icon)

        profile_layout.addWidget(profile_pic_container)

        # User info section
        user_info_layout = QVBoxLayout()
        user_info_layout.setSpacing(5)

        # User name
        user_name_label = QLabel(f"👋 {self.user_name}")
        user_name_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 16px;
                font-weight: 600;
                background: transparent;
                border: none;
            }
        """)

        # User role
        role_label = QLabel(f"{'🏢 Propriétaire' if self.user_role == 'gym owner' else '💪 Membre'}")
        role_label.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.7);
                font-size: 12px;
                background: transparent;
                border: none;
            }
        """)

        user_info_layout.addWidget(user_name_label)
        user_info_layout.addWidget(role_label)
        profile_layout.addLayout(user_info_layout)
        sidebar_layout.addWidget(profile_container)

        # Modern Logout button
        logout_btn = QPushButton("🚪 Déconnexion")
        logout_btn.setCursor(Qt.PointingHandCursor)
        logout_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255, 107, 107, 0.2);
                color: #ff6b6b;
                font-size: 14px;
                font-weight: 600;
                padding: 14px 20px;
                border-radius: 12px;
                border: 1px solid rgba(255, 107, 107, 0.3);
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: rgba(255, 107, 107, 0.3);
                border: 1px solid rgba(255, 107, 107, 0.5);
                color: #ff5252;
            }
        """)
        logout_btn.clicked.connect(self.logout)
        sidebar_layout.addWidget(logout_btn)

        main_layout.addWidget(sidebar)
    
    def get_nav_buttons(self):
        if self.user_role == "gym owner":
            return [
                ("🏠 Dashboard", self.show_dashboard, True),
                ("🏋️ Mes Salles", self.show_gyms, False),
                ("💰 Paiements", self.show_payments, False),
                ("👥 Membres", self.show_members, False),
                ("📊 Statistiques", self.show_statistics, False),
                ("📅 Planning", self.show_planning, False),
                ("📝 Historique Membres", self.show_member_workouts, False),
                ("🏦 RIB", self.show_rib_management, False),
                ("👤 Mon Profil", self.show_profile, False),  # Nouveau bouton
                ("⚙️ Paramètres", lambda: self.show_coming_soon("Paramètres"), False)
            ]
        else:  # Member
            return [
                ("🏠 Accueil", self.show_nearby_gyms, True),
                ("💳 Mes Abonnements", self.show_my_subscriptions, False),
                ("➕ Pointer entraînement", self.track_workout, False),
                ("📊 Historique", self.show_workout_history, False),
                ("💬 Chat de la salle", self.open_gym_chat, False),
                ("⭐ Évaluer une salle", self.rate_gym, False),
                ("👤 Mon Profil", self.show_profile, False),  # Nouveau bouton
                ("⚙️ Paramètres", lambda: self.show_coming_soon("Paramètres"), False)
            ]
    def open_gym_chat(self):
        """Open chat for a selected gym"""
        # Use gym selection dialog for consistency
        gym_dialog = GymSelectionDialog(self.user_id, "Choisir une salle pour le chat", self)
        if gym_dialog.exec() == QDialog.Accepted:
            gym_id = gym_dialog.get_selected_gym_id()
            if gym_id:
                dialog = GymChatDialog(gym_id, self.user_id, self)
                dialog.exec()
    
    def create_gym_card(self, gym):
        # Ultra-modern gym card with glassmorphism and animations
        card = QFrame()
        card.setFixedHeight(220)  # Taller for better proportions
        card.setCursor(Qt.PointingHandCursor)
        card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(45, 45, 45, 0.95),
                    stop:0.5 rgba(35, 35, 35, 0.90),
                    stop:1 rgba(25, 25, 25, 0.95));
                border-radius: 15px;
                border: 1px solid rgba(255, 140, 0, 0.3);
                backdrop-filter: blur(10px);
                transition: all 0.3s ease-in-out;
                box-shadow: 0 4px 15px rgba(0, 0, 0, 0.3);
            }
            QFrame:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(55, 55, 55, 0.95),
                    stop:0.5 rgba(45, 45, 45, 0.90),
                    stop:1 rgba(35, 35, 35, 0.95));
                border: 1px solid rgba(255, 140, 0, 0.6);
                transform: translateY(-5px);
                box-shadow: 0 8px 25px rgba(255, 140, 0, 0.2);
            }
        """)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(25, 20, 25, 20)
        layout.setSpacing(15)

        # Modern header section
        header_container = QWidget()
        header_container.setStyleSheet("background: transparent; border: none;")
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(8)

        # Top row with name and rating
        top_row = QWidget()
        top_row.setStyleSheet("background: transparent; border: none;")
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Modern gym name with gradient text effect
        name = QLabel(f"🏋️ {gym['name']}")
        name.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: 700;
                color: white;
                background: transparent;
                border: none;
                letter-spacing: 0.5px;
            }
        """)

        # Modern rating display
        avg_rating = DatabaseManager.get_gym_avg_rating(gym['id'])
        rating_container = QWidget()
        rating_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 193, 7, 0.2);
                border-radius: 12px;
                border: 1px solid rgba(255, 193, 7, 0.3);
                padding: 5px 10px;
            }
        """)
        rating_layout = QHBoxLayout(rating_container)
        rating_layout.setContentsMargins(8, 4, 8, 4)
        rating_layout.setSpacing(5)

        # Modern star display
        stars = QLabel("★" * int(round(avg_rating['avg'])) + "☆" * (5 - int(round(avg_rating['avg']))))
        stars.setStyleSheet("""
            QLabel {
                color: #FFC107;
                font-size: 14px;
                background: transparent;
                border: none;
            }
        """)

        rating_text = QLabel(f"{avg_rating['avg']}")
        rating_text.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: white;
                font-weight: 600;
                background: transparent;
                border: none;
            }
        """)

        rating_layout.addWidget(stars)
        rating_layout.addWidget(rating_text)

        top_layout.addWidget(name)
        top_layout.addStretch()
        top_layout.addWidget(rating_container)
        header_layout.addWidget(top_row)

        # Modern address with icon
        address = QLabel(f"📍 {gym['address']}")
        address.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: rgba(255, 255, 255, 0.8);
                background: transparent;
                border: none;
                font-weight: 500;
                margin-top: 5px;
            }
        """)
        header_layout.addWidget(address)
        layout.addWidget(header_container)

        # Modern button section
        btn_container = QWidget()
        btn_container.setStyleSheet("background: transparent; border: none;")
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.setSpacing(10)

        # Modern reviews button
        reviews_btn = QPushButton("👁️ Avis")
        reviews_btn.setCursor(Qt.PointingHandCursor)
        reviews_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255, 255, 255, 0.1);
                color: rgba(255, 255, 255, 0.9);
                font-size: 12px;
                font-weight: 600;
                padding: 10px 15px;
                border-radius: 10px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: rgba(255, 255, 255, 0.2);
                border: 1px solid rgba(255, 255, 255, 0.4);
                color: white;
            }
        """)
        reviews_btn.clicked.connect(lambda _, gid=gym['id']: self.show_reviews(gid))

        # Check subscription status
        is_subscribed = DatabaseManager.execute_query("""
            SELECT id FROM subscriptions
            WHERE member_id = %s AND gym_id = %s AND payment_status = 'confirmed'
        """, (self.user_id, gym['id']), fetch_one=True)

        btn_layout.addWidget(reviews_btn)

        # Modern rating button for subscribed users
        if is_subscribed:
            has_rated = DatabaseManager.execute_query("""
                SELECT id FROM gym_reviews
                WHERE gym_id = %s AND member_id = %s
            """, (gym['id'], self.user_id), fetch_one=True)

            if not has_rated:
                rate_btn = QPushButton("⭐ Évaluer")
                rate_btn.setCursor(Qt.PointingHandCursor)
                rate_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #f093fb, stop:1 #f5576c);
                        color: white;
                        font-size: 12px;
                        font-weight: 600;
                        padding: 10px 15px;
                        border-radius: 10px;
                        border: none;
                        letter-spacing: 0.5px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #e91e63, stop:1 #f44336);
                    }
                """)
                rate_btn.clicked.connect(lambda _, gid=gym['id']: self.rate_specific_gym(gid))
                btn_layout.addWidget(rate_btn)

        btn_layout.addStretch()

        # Modern action button
        if not is_subscribed:
            subscribe_btn = QPushButton("✨ S'abonner")
            subscribe_btn.setCursor(Qt.PointingHandCursor)
            subscribe_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #ff6b35, stop:1 #ff8c00);
                    color: white;
                    font-size: 12px;
                    font-weight: 700;
                    padding: 12px 20px;
                    border-radius: 12px;
                    border: none;
                    letter-spacing: 0.5px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #ff5722, stop:1 #ff7700);
                    transform: translateY(-2px);
                }
            """)
            subscribe_btn.clicked.connect(lambda _, gid=gym['id']: self.subscribe_to_gym(gid))
            btn_layout.addWidget(subscribe_btn)
        else:
            # Modern subscribed status
            status_btn = QPushButton("✅ Abonné")
            status_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4facfe, stop:1 #00f2fe);
                    color: white;
                    font-size: 12px;
                    font-weight: 600;
                    padding: 12px 20px;
                    border-radius: 12px;
                    border: none;
                    letter-spacing: 0.5px;
                }
            """)
            status_btn.setEnabled(False)
            btn_layout.addWidget(status_btn)

        # Assemble modern card
        layout.addStretch()
        layout.addWidget(btn_container)

        return card
    


    def show_profile(self):
        # 1. Récupération des données utilisateur EN PREMIER
        user_info = DatabaseManager.execute_query(
            "SELECT first_name, last_name, email, phone, address, profile_pic FROM users WHERE id = %s",
            (self.user_id,),
            fetch_one=True
        )
        
        if not user_info:
            QMessageBox.warning(self, "Erreur", "Impossible de charger les informations du profil")
            return
        
        # 2. Stockage de l'email original
        self.original_email = user_info['email']
        
        # 3. Création de l'interface
        profile_page = QWidget()
        layout = QVBoxLayout(profile_page)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Titre
        title = QLabel("Mon Profil")
        title.setStyleSheet("""
            font-size: 22px;
            font-weight: bold;
            color: #002347;
        """)
        layout.addWidget(title)

        # Photo de profil
        self.profile_pic_label = QLabel()
        self.profile_pic_label.setAlignment(Qt.AlignCenter)
        self.profile_pic_label.setFixedSize(180, 180)
        self.profile_pic_label.setStyleSheet("""
            QLabel {
                border: 2px solid #E0E0E0;
                border-radius: 8px;
                background-color: #FAFAFA;
            }
        """)
        
        # Chargement de la photo
        if user_info['profile_pic']:
            pixmap = QPixmap()
            pixmap.loadFromData(user_info['profile_pic'])
            self.profile_pic_label.setPixmap(pixmap.scaled(
                180, 180, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation
            ))
        else:
            default_icon = QIcon.fromTheme("user").pixmap(120, 120)
            self.profile_pic_label.setPixmap(default_icon)

        # Bouton changement photo
        change_btn = QPushButton("Changer la photo")
        change_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                padding: 8px;
                border-radius: 5px;
                margin-top: 10px;
            }
        """)
        change_btn.clicked.connect(self.change_profile_picture)

        # Layout photo
        photo_layout = QVBoxLayout()
        photo_layout.addWidget(self.profile_pic_label)
        photo_layout.addWidget(change_btn, 0, Qt.AlignCenter)
        
        # Layout principal
        main_layout = QHBoxLayout()
        main_layout.addLayout(photo_layout)
        
        # Formulaire
        form_layout = QFormLayout()
        form_layout.setVerticalSpacing(15)
        form_layout.setHorizontalSpacing(20)

        # Création des champs
        self.first_name_edit = QLineEdit(user_info['first_name'])
        self.last_name_edit = QLineEdit(user_info['last_name'])
        self.email_edit = QLineEdit(user_info['email'])
        self.phone_edit = QLineEdit(user_info.get('phone', ''))
        self.address_edit = QLineEdit(user_info.get('address', ''))

        # Style des champs
        field_style = """
            QLineEdit {
                padding: 10px;
                border: 1px solid #DDD;
                border-radius: 5px;
                min-width: 250px;
            }
            QLineEdit:focus {
                border: 2px solid #002347;
            }
        """

        # Ajout des champs au formulaire
        form_layout.addRow("Prénom:", self.first_name_edit)
        form_layout.addRow("Nom:", self.last_name_edit)
        form_layout.addRow("Email:", self.email_edit)
        form_layout.addRow("Téléphone:", self.phone_edit)
        form_layout.addRow("Adresse:", self.address_edit)

        # Appliquer le style
        for edit in [self.first_name_edit, self.last_name_edit, 
                    self.email_edit, self.phone_edit, self.address_edit]:
            edit.setStyleSheet(field_style)

        main_layout.addLayout(form_layout, 1)
        layout.addLayout(main_layout)

        # Boutons
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("Enregistrer les modifications")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: 500;
            }
        """)
        save_btn.clicked.connect(self.save_profile)

        pwd_btn = QPushButton("Changer le mot de passe")
        pwd_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: 500;
            }
        """)
        pwd_btn.clicked.connect(self.show_change_password_dialog)

        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(pwd_btn)
        layout.addLayout(btn_layout)

        self.stacked_widget.addWidget(profile_page)
        self.stacked_widget.setCurrentIndex(self.stacked_widget.count() - 1)

    def get_remaining_changes(self):
        # Vérifier combien de modifications l'utilisateur a fait ce mois-ci
        changes = DatabaseManager.execute_query(
            "SELECT COUNT(*) as count FROM profile_changes "
            "WHERE user_id = %s AND change_date >= DATE_FORMAT(NOW(), '%Y-%m-01')",
            (self.user_id,),
            fetch_one=True
        )
        max_changes = 3  # Limite mensuelle
        return max(0, max_changes - changes['count'])

    def change_profile_picture(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Choisir une photo", "", "Images (*.png *.jpg *.jpeg)"
        )
        if file_path:
            pixmap = QPixmap(file_path)
            if not pixmap.isNull():
                self.profile_pic_label.setPixmap(pixmap.scaled(
                    180, 180, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation
                ))
                # Stocker le chemin pour sauvegarde ultérieure
                self.new_profile_pic_path = file_path
                # Pas de comptage pour les changements de photo
            else:
                QMessageBox.warning(self, "Erreur", "Impossible de charger l'image")

    def save_profile(self):
        # Liste des champs à vérifier (exclure la photo)
        limited_fields = [
            'first_name',
            'last_name', 
            'email',
            'phone',
            'address'
        ]
        
        # Vérifier si au moins un champ limité a été modifié
        fields_changed = False
        for field in limited_fields:
            current_value = getattr(self, f"{field}_edit").text().strip()
            original_value = DatabaseManager.execute_query(
                f"SELECT {field} FROM users WHERE id = %s",
                (self.user_id,),
                fetch_one=True
            ).get(field, '')
            
            if current_value != original_value:
                fields_changed = True
                break
        
        # Si modification des champs limités, vérifier le quota
        if fields_changed:
            changes_count = DatabaseManager.execute_query(
                "SELECT COUNT(*) as count FROM profile_changes "
                "WHERE user_id = %s AND change_type = 'profile_update' "
                "AND change_date >= DATE_FORMAT(NOW(), '%Y-%m-01')",
                (self.user_id,),
                fetch_one=True
            )['count']
            
            if changes_count >= 5:  # Limite de 5 modifications/mois
                QMessageBox.warning(
                    self, 
                    "Limite atteinte", 
                    "Vous avez atteint votre limite de 5 modifications ce mois-ci.\n"
                    "Vous pouvez toujours changer votre photo de profil."
                )
                return
        
        # Enregistrement des modifications
        try:
            new_first = self.first_name_edit.text().strip()
            new_last = self.last_name_edit.text().strip()
            new_email = self.email_edit.text().strip()
            new_phone = self.phone_edit.text().strip()
            new_address = self.address_edit.text().strip()
            
            # Vérification email unique
            if new_email != self.original_email:
                existing = DatabaseManager.execute_query(
                    "SELECT id FROM users WHERE email = %s AND id != %s",
                    (new_email, self.user_id),
                    fetch_one=True
                )
                if existing:
                    QMessageBox.warning(self, "Email existant", "Cet email est déjà utilisé")
                    return
            
            # Sauvegarde photo si modifiée
            profile_pic_data = None
            if hasattr(self, 'new_profile_pic_path'):
                with open(self.new_profile_pic_path, 'rb') as f:
                    profile_pic_data = f.read()
            
            # Mise à jour en base
            success = DatabaseManager.execute_query(
                "UPDATE users SET first_name = %s, last_name = %s, email = %s, "
                "phone = %s, address = %s, profile_pic = %s WHERE id = %s",
                (new_first, new_last, new_email, new_phone, new_address, profile_pic_data, self.user_id)
            )
            
            if success:
                if fields_changed:  # Seulement enregistrer si modification des champs limités
                    DatabaseManager.execute_query(
                        "INSERT INTO profile_changes (user_id, change_type) VALUES (%s, 'profile_update')",
                        (self.user_id,)
                    )
                QMessageBox.information(self, "Succès", "Profil mis à jour!")
                self.show_profile()  # Recharger
            else:
                QMessageBox.critical(self, "Erreur", "Échec de la mise à jour")
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Une erreur est survenue: {str(e)}")
    def show_change_password_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Changer le mot de passe")
        dialog.setFixedSize(400, 300)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        
        form = QFormLayout()
        form.setSpacing(15)
        
        current_pwd = QLineEdit()
        current_pwd.setEchoMode(QLineEdit.Password)
        new_pwd = QLineEdit()
        new_pwd.setEchoMode(QLineEdit.Password)
        confirm_pwd = QLineEdit()
        confirm_pwd.setEchoMode(QLineEdit.Password)
        
        for edit in [current_pwd, new_pwd, confirm_pwd]:
            edit.setStyleSheet("padding: 8px; border: 1px solid #ddd; border-radius: 5px;")
        
        form.addRow("Mot de passe actuel:", current_pwd)
        form.addRow("Nouveau mot de passe:", new_pwd)
        form.addRow("Confirmer le nouveau:", confirm_pwd)
        
        layout.addLayout(form)
        
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Enregistrer")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        save_btn.clicked.connect(lambda: self.change_password(
            current_pwd.text(), new_pwd.text(), confirm_pwd.text(), dialog
        ))
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setStyleSheet("background-color: #f5f5f5; padding: 8px;")
        cancel_btn.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
        
        dialog.exec()

    def change_password(self, current_pwd, new_pwd, confirm_pwd, dialog):
        # Validation
        if not all([current_pwd, new_pwd, confirm_pwd]):
            QMessageBox.warning(self, "Champs requis", "Veuillez remplir tous les champs.")
            return
        
        if new_pwd != confirm_pwd:
            QMessageBox.warning(self, "Erreur", "Les nouveaux mots de passe ne correspondent pas.")
            return
        
        if len(new_pwd) < 6:
            QMessageBox.warning(self, "Erreur", "Le mot de passe doit contenir au moins 6 caractères.")
            return
        
        # Vérifier l'ancien mot de passe
        user = DatabaseManager.execute_query(
            "SELECT password FROM users WHERE id = %s",
            (self.user_id,),
            fetch_one=True
        )
        
        if not user or user['password'] != current_pwd:  # En production, utiliser un hash!
            QMessageBox.warning(self, "Erreur", "Mot de passe actuel incorrect.")
            return
        
        # Mettre à jour le mot de passe
        success = DatabaseManager.execute_query(
            "UPDATE users SET password = %s WHERE id = %s",
            (new_pwd, self.user_id)  # En production, hasher le mot de passe!
        )
        
        if success:
            QMessageBox.information(self, "Succès", "Mot de passe changé avec succès!")
            dialog.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Échec de la mise à jour du mot de passe.")
    def create_modern_header(self):
        header = QWidget()
        header.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 255, 255, 0.15),
                    stop:1 rgba(255, 255, 255, 0.08));
                border-radius: 18px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(25, 20, 25, 20)

        welcome_container = QWidget()
        welcome_container.setStyleSheet("background: transparent; border: none;")
        welcome_layout = QVBoxLayout(welcome_container)
        welcome_layout.setContentsMargins(0, 0, 0, 0)
        welcome_layout.setSpacing(5)

        # Modern welcome message with gradient text
        welcome = QLabel(f"Bonjour, {self.user_name}! ✨")
        welcome.setStyleSheet("""
            QLabel {
                font-size: 28px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)

        # Subtitle with modern styling
        if self.user_role == "member":
            subtitle = QLabel(f"📍 {self.address or 'Adresse non spécifiée'}")
        else:
            subtitle = QLabel(QDate.currentDate().toString("dddd, d MMMM yyyy"))

        subtitle.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: rgba(255, 255, 255, 0.8);
                background: transparent;
                border: none;
                font-weight: 500;
            }
        """)

        welcome_layout.addWidget(welcome)
        welcome_layout.addWidget(subtitle)
        header_layout.addWidget(welcome_container)

        if self.user_role == "gym owner":
            header_layout.addStretch()
            self.create_modern_stats_widget(header_layout)

        self.content_layout.addWidget(header)
    
    def create_modern_stats_widget(self, layout):
        stats_container = QWidget()
        stats_container.setStyleSheet("background: transparent; border: none;")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(12)

        # Get real statistics
        stats_data = DatabaseManager.get_gym_stats(self.user_id)

        if stats_data is None:
            stats_data = {
                'gyms': 0,
                'members': 0,
                'revenue': 0.0
            }

        # Modern gym-style stats with appropriate colors
        stats = [
            ("🏢", "Salles", str(stats_data['gyms']), ["#ff8c00", "#ff6b35"]),
            ("👥", "Membres", str(stats_data['members']), ["#ff6b35", "#ff8c00"]),
            ("💰", "Revenus", f"€{stats_data['revenue']:.2f}", ["#28a745", "#20c997"])
        ]

        for icon, label, value, gradient in stats:
            stat_widget = QWidget()
            stat_widget.setStyleSheet(f"""
                QWidget {{
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 {gradient[0]}, stop:1 {gradient[1]});
                    border-radius: 15px;
                    border: 1px solid rgba(255, 255, 255, 0.3);
                    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
                    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
                }}
                QWidget:hover {{
                    transform: translateY(-3px) scale(1.05);
                    box-shadow: 0 8px 25px rgba(0, 0, 0, 0.2);
                    border: 1px solid rgba(255, 255, 255, 0.5);
                }}
            """)
            stat_widget.setFixedSize(120, 80)

            stat_layout = QVBoxLayout(stat_widget)
            stat_layout.setContentsMargins(12, 8, 12, 8)
            stat_layout.setSpacing(2)
            stat_layout.setAlignment(Qt.AlignCenter)

            # Icon and value in same line
            top_container = QWidget()
            top_container.setStyleSheet("background: transparent; border: none;")
            top_layout = QHBoxLayout(top_container)
            top_layout.setContentsMargins(0, 0, 0, 0)
            top_layout.setSpacing(5)

            icon_label = QLabel(icon)
            icon_label.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    background: transparent;
                    border: none;
                }
            """)

            value_label = QLabel(value)
            value_label.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    font-weight: 700;
                    color: white;
                    background: transparent;
                    border: none;
                }
            """)

            top_layout.addWidget(icon_label)
            top_layout.addWidget(value_label)
            top_layout.addStretch()

            # Label
            label_label = QLabel(label)
            label_label.setStyleSheet("""
                QLabel {
                    font-size: 11px;
                    color: rgba(255, 255, 255, 0.9);
                    background: transparent;
                    border: none;
                    font-weight: 500;
                }
            """)
            label_label.setAlignment(Qt.AlignCenter)

            stat_layout.addWidget(top_container)
            stat_layout.addWidget(label_label)
            stats_layout.addWidget(stat_widget)

        layout.addWidget(stats_container)
    
    def create_dashboard_views(self):
        if self.user_role == "gym owner":
            self.create_gym_owner_views()
        else:
            self.create_member_views()
    
    def create_gym_owner_views(self):
        # Modern Dashboard view
        view = QWidget()
        view.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(view)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(25)

        # Modern title section with glassmorphism
        title_container = QWidget()
        title_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.1);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("🏢 Mes Salles de Sport")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)

        # Modern add button with gradient
        add_btn = QPushButton("✨ Ajouter une salle")
        add_btn.setCursor(Qt.PointingHandCursor)
        add_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff6b35, stop:1 #ff8c00);
                color: white;
                font-size: 14px;
                font-weight: 600;
                padding: 14px 20px;
                border-radius: 12px;
                border: none;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff5722, stop:1 #ff7700);
                transform: translateY(-2px);
            }
        """)
        add_btn.clicked.connect(self.show_add_gym)

        title_layout.addWidget(title)
        title_layout.addStretch()
        title_layout.addWidget(add_btn)
        layout.addWidget(title_container)

        # Modern scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 255, 255, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 255, 255, 0.3);
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(255, 255, 255, 0.5);
            }
        """)

        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        self.gyms_layout = QVBoxLayout(scroll_content)
        self.gyms_layout.setContentsMargins(0, 0, 10, 0)
        self.gyms_layout.setSpacing(20)

        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        self.stacked_widget.addWidget(view)
        
        # Add gym view
        self.add_gym_page = AddGymPage(self.user_id, self)
        self.stacked_widget.addWidget(self.add_gym_page)
        
        # Modern Payments view
        payments_view = QWidget()
        payments_view.setStyleSheet("background: transparent;")
        payments_layout = QVBoxLayout(payments_view)
        payments_layout.setContentsMargins(0, 0, 0, 0)
        payments_layout.setSpacing(25)

        # Modern title section
        title_container = QWidget()
        title_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.1);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("💳 Paiements en attente")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)
        title_layout.addWidget(title)
        payments_layout.addWidget(title_container)

        # Modern scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 255, 255, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 255, 255, 0.3);
                border-radius: 4px;
                min-height: 20px;
            }
        """)
        content = QWidget()
        content.setStyleSheet("background: transparent;")
        self.payments_layout = QVBoxLayout(content)
        self.payments_layout.setSpacing(15)
        scroll.setWidget(content)
        payments_layout.addWidget(scroll)

        self.stacked_widget.addWidget(payments_view)
        
        # Modern Members view
        members_view = QWidget()
        members_view.setStyleSheet("background: transparent;")
        members_layout = QVBoxLayout(members_view)
        members_layout.setContentsMargins(0, 0, 0, 0)
        members_layout.setSpacing(25)

        # Modern title and search section
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.1);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(25, 20, 25, 20)
        header_layout.setSpacing(15)

        title = QLabel("👥 Liste des Membres")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)
        header_layout.addWidget(title)

        # Modern search bar
        search_container = QWidget()
        search_container.setStyleSheet("background: transparent; border: none;")
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(10)

        self.member_search_input = QLineEdit()
        self.member_search_input.setPlaceholderText("🔍 Rechercher un membre...")
        self.member_search_input.setStyleSheet("""
            QLineEdit {
                background: rgba(255, 255, 255, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 12px;
                padding: 14px 18px;
                font-size: 14px;
                color: white;
                font-weight: 500;
            }
            QLineEdit:focus {
                border: 2px solid rgba(102, 126, 234, 0.8);
                background: rgba(255, 255, 255, 0.2);
            }
            QLineEdit::placeholder {
                color: rgba(255, 255, 255, 0.6);
            }
        """)
        self.member_search_input.textChanged.connect(self.search_members)

        search_btn = QPushButton("🔍 Rechercher")
        search_btn.setCursor(Qt.PointingHandCursor)
        search_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4facfe, stop:1 #00f2fe);
                color: white;
                font-size: 14px;
                font-weight: 600;
                padding: 14px 20px;
                border-radius: 12px;
                border: none;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #06b6d4);
            }
        """)
        search_btn.clicked.connect(self.search_members)
        
        search_layout.addWidget(self.member_search_input)
        search_layout.addWidget(search_btn)
        members_layout.addWidget(search_container)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none; background: transparent;")
        
        scroll_content = QWidget()
        self.members_layout = QVBoxLayout(scroll_content)
        self.members_layout.setContentsMargins(0, 0, 10, 0)
        self.members_layout.setSpacing(15)
        
        scroll.setWidget(scroll_content)
        members_layout.addWidget(scroll)
        
        self.stacked_widget.addWidget(members_view)


        # Modern Statistics view - Create empty placeholder that will be populated by load_statistics
        stats_view = QWidget()
        stats_view.setStyleSheet("background: transparent;")
        stats_layout = QVBoxLayout(stats_view)
        stats_layout.setContentsMargins(0, 0, 0, 0)
        stats_layout.setSpacing(0)

        # Store reference to the main layout for load_statistics to use
        self.main_stats_layout = stats_layout

        self.stacked_widget.addWidget(stats_view)
    
    def create_member_views(self):
        # Modern Member dashboard view
        view = QWidget()
        view.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(view)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(25)

        # Modern title section
        title_container = QWidget()
        title_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.1);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("🏋️ Salles à proximité")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)
        title_layout.addWidget(title)
        layout.addWidget(title_container)

        # Modern scroll area with custom styling
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 255, 255, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 255, 255, 0.3);
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(255, 255, 255, 0.5);
            }
        """)

        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        self.nearby_gyms_layout = QVBoxLayout(scroll_content)
        self.nearby_gyms_layout.setContentsMargins(0, 0, 10, 0)
        self.nearby_gyms_layout.setSpacing(20)

        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        self.stacked_widget.addWidget(view)
        
        # Modern Subscriptions view
        subscriptions_view = QWidget()
        subscriptions_view.setStyleSheet("background: transparent;")
        subscriptions_layout = QVBoxLayout(subscriptions_view)
        subscriptions_layout.setContentsMargins(0, 0, 0, 0)
        subscriptions_layout.setSpacing(25)

        # Modern title section
        title_container = QWidget()
        title_container.setStyleSheet("""
            QWidget {
                background: rgba(255, 255, 255, 0.1);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                backdrop-filter: blur(20px);
            }
        """)
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("💳 Mes Abonnements")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: white;
                font-weight: 700;
                background: transparent;
                border: none;
                letter-spacing: 1px;
            }
        """)
        title_layout.addWidget(title)
        subscriptions_layout.addWidget(title_container)

        # Modern scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 255, 255, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 255, 255, 0.3);
                border-radius: 4px;
                min-height: 20px;
            }
        """)
        content = QWidget()
        content.setStyleSheet("background: transparent;")
        self.subscriptions_layout = QVBoxLayout(content)
        self.subscriptions_layout.setSpacing(20)
        scroll.setWidget(content)
        subscriptions_layout.addWidget(scroll)

        self.stacked_widget.addWidget(subscriptions_view)
    
    def show_initial_view(self):
        self.show_dashboard() if self.user_role == "gym owner" else self.show_nearby_gyms()
    
    def show_dashboard(self):
        self.stacked_widget.setCurrentIndex(0)
        self.load_gyms()

    def show_gyms(self):
        self.stacked_widget.setCurrentIndex(0)
        self.load_gyms()

    def show_add_gym(self):
        self.stacked_widget.setCurrentIndex(1)

    def show_payments(self):
        self.stacked_widget.setCurrentIndex(2)  # Correct: Payments view is at index 2
        self.load_payments()

    def show_members(self):
        self.stacked_widget.setCurrentIndex(3)  # Correct: Members view is at index 3
        self.load_members()
    
    def show_nearby_gyms(self):
        self.stacked_widget.setCurrentIndex(0)
        self.load_nearby_gyms()
    
    def show_my_subscriptions(self):
        self.stacked_widget.setCurrentIndex(1)
        self.load_subscriptions()
    
    def show_coming_soon(self, feature):
        QMessageBox.information(self, "À venir", f"La fonctionnalité '{feature}' sera bientôt disponible!")
    
    def show_rib_management(self):
        dialog = BankDetailsDialog(self.user_id, self)
        dialog.exec()
    
    def load_gyms(self):
        for i in reversed(range(self.gyms_layout.count())):
            widget = self.gyms_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        gyms = DatabaseManager.execute_query(
            "SELECT * FROM gyms WHERE owner_id = %s", 
            (self.user_id,), 
            fetch_all=True
        )
        
        if not gyms:
            label = QLabel("Vous n'avez pas encore de salle enregistrée")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.gyms_layout.addWidget(label)
        else:
            for gym in gyms:
                # Modern gym owner card
                card = QFrame()
                card.setCursor(Qt.PointingHandCursor)
                card.setStyleSheet("""
                    QFrame {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(255, 255, 255, 0.15),
                            stop:0.5 rgba(255, 255, 255, 0.10),
                            stop:1 rgba(255, 255, 255, 0.08));
                        border-radius: 18px;
                        border: 1px solid rgba(255, 255, 255, 0.2);
                        backdrop-filter: blur(20px);
                    }
                    QFrame:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(255, 255, 255, 0.25),
                            stop:0.5 rgba(255, 255, 255, 0.18),
                            stop:1 rgba(255, 255, 255, 0.15));
                        border: 1px solid rgba(102, 126, 234, 0.4);
                        transform: translateY(-3px);
                    }
                """)
                card_layout = QVBoxLayout(card)
                card_layout.setContentsMargins(25, 20, 25, 20)
                card_layout.setSpacing(15)

                # Header with name and modify button
                header_container = QWidget()
                header_container.setStyleSheet("background: transparent; border: none;")
                header_layout = QHBoxLayout(header_container)
                header_layout.setContentsMargins(0, 0, 0, 0)

                name = QLabel(f"🏢 {gym['name']}")
                name.setStyleSheet("""
                    QLabel {
                        font-size: 20px;
                        font-weight: 700;
                        color: white;
                        background: transparent;
                        border: none;
                        letter-spacing: 0.5px;
                    }
                """)

                # Modern modify button
                modify_btn = QPushButton("⚙️ Modifier")
                modify_btn.setCursor(Qt.PointingHandCursor)
                modify_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #f093fb, stop:1 #f5576c);
                        color: white;
                        font-size: 12px;
                        font-weight: 600;
                        padding: 10px 15px;
                        border-radius: 10px;
                        border: none;
                        letter-spacing: 0.5px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #e91e63, stop:1 #f44336);
                    }
                """)
                modify_btn.clicked.connect(lambda _, gid=gym['id']: self.modify_gym(gid))

                header_layout.addWidget(name)
                header_layout.addStretch()
                header_layout.addWidget(modify_btn)

                # Address
                address = QLabel(f"📍 {gym['address']}")
                address.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: rgba(255, 255, 255, 0.8);
                        background: transparent;
                        border: none;
                        font-weight: 500;
                    }
                """)

                card_layout.addWidget(header_container)
                card_layout.addWidget(address)
                self.gyms_layout.addWidget(card)

    def modify_gym(self, gym_id):
        dialog = ModifyGymDialog(gym_id, self)
        dialog.exec()
        self.load_gyms()  # Rafraîchir la liste après modification
    
    def load_payments(self):
        for i in reversed(range(self.payments_layout.count())):
            widget = self.payments_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        payments = DatabaseManager.execute_query("""
            SELECT s.*, u.first_name, u.last_name, g.name as gym_name
            FROM subscriptions s
            JOIN users u ON s.member_id = u.id
            JOIN gyms g ON s.gym_id = g.id
            WHERE g.owner_id = %s AND s.payment_status = 'pending'
        """, (self.user_id,), fetch_all=True)
        
        if not payments:
            # Modern empty state
            empty_container = QWidget()
            empty_container.setStyleSheet("""
                QWidget {
                    background: rgba(255, 255, 255, 0.1);
                    border-radius: 16px;
                    border: 1px solid rgba(255, 255, 255, 0.2);
                    backdrop-filter: blur(20px);
                }
            """)
            empty_layout = QVBoxLayout(empty_container)
            empty_layout.setContentsMargins(40, 30, 40, 30)

            label = QLabel("💳 Aucun paiement en attente")
            label.setStyleSheet("""
                QLabel {
                    font-size: 18px;
                    color: rgba(255, 255, 255, 0.8);
                    background: transparent;
                    border: none;
                    font-weight: 600;
                }
            """)
            label.setAlignment(Qt.AlignCenter)
            empty_layout.addWidget(label)
            self.payments_layout.addWidget(empty_container)
        else:
            for payment in payments:
                # Modern payment card
                card = QFrame()
                card.setCursor(Qt.PointingHandCursor)
                card.setStyleSheet("""
                    QFrame {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(255, 255, 255, 0.15),
                            stop:0.5 rgba(255, 255, 255, 0.10),
                            stop:1 rgba(255, 255, 255, 0.08));
                        border-radius: 18px;
                        border: 1px solid rgba(255, 255, 255, 0.2);
                        backdrop-filter: blur(20px);
                    }
                    QFrame:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(255, 255, 255, 0.25),
                            stop:0.5 rgba(255, 255, 255, 0.18),
                            stop:1 rgba(255, 255, 255, 0.15));
                        border: 1px solid rgba(102, 126, 234, 0.4);
                    }
                """)
                card_layout = QVBoxLayout(card)
                card_layout.setContentsMargins(25, 20, 25, 20)
                card_layout.setSpacing(12)

                # Header with member name and verify button
                header_container = QWidget()
                header_container.setStyleSheet("background: transparent; border: none;")
                header_layout = QHBoxLayout(header_container)
                header_layout.setContentsMargins(0, 0, 0, 0)

                member = QLabel(f"👤 {payment['first_name']} {payment['last_name']}")
                member.setStyleSheet("""
                    QLabel {
                        font-size: 18px;
                        font-weight: 700;
                        color: white;
                        background: transparent;
                        border: none;
                        letter-spacing: 0.5px;
                    }
                """)

                # Modern verify button
                btn = QPushButton("✅ Vérifier")
                btn.setCursor(Qt.PointingHandCursor)
                btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #4facfe, stop:1 #00f2fe);
                        color: white;
                        font-size: 12px;
                        font-weight: 600;
                        padding: 10px 15px;
                        border-radius: 10px;
                        border: none;
                        letter-spacing: 0.5px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #3b82f6, stop:1 #06b6d4);
                    }
                """)
                btn.clicked.connect(lambda _, sid=payment['id']: self.verify_payment(sid))

                header_layout.addWidget(member)
                header_layout.addStretch()
                header_layout.addWidget(btn)

                # Payment details
                gym = QLabel(f"🏢 {payment['gym_name']}")
                gym.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: rgba(255, 255, 255, 0.8);
                        background: transparent;
                        border: none;
                        font-weight: 500;
                    }
                """)

                sub_type = QLabel(f"📋 Abonnement: {payment['subscription_type']}")
                sub_type.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: rgba(255, 255, 255, 0.8);
                        background: transparent;
                        border: none;
                        font-weight: 500;
                    }
                """)

                method = QLabel(f"💰 Méthode: {'Espèces' if payment['payment_method'] == 'espece' else 'Virement'}")
                method.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        color: rgba(255, 255, 255, 0.8);
                        background: transparent;
                        border: none;
                        font-weight: 500;
                    }
                """)

                card_layout.addWidget(header_container)
                card_layout.addWidget(gym)
                card_layout.addWidget(sub_type)
                card_layout.addWidget(method)

                self.payments_layout.addWidget(card)
    
    def verify_payment(self, subscription_id):
        dialog = PaymentConfirmationDialog(subscription_id, self)
        dialog.exec()
        self.load_payments()
    
    def load_members(self):
        for i in reversed(range(self.members_layout.count())):
            widget = self.members_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()
        
        gyms = DatabaseManager.execute_query(
            "SELECT id FROM gyms WHERE owner_id = %s", 
            (self.user_id,), 
            fetch_all=True
        )
        
        if not gyms:
            label = QLabel("Vous n'avez pas encore de salle enregistrée")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.members_layout.addWidget(label)
            return
        
        gym_ids = [str(gym['id']) for gym in gyms]
        
        members = DatabaseManager.execute_query(f"""
            SELECT DISTINCT u.id, u.first_name, u.last_name, u.email, u.phone,
                   s.payment_status, s.subscription_type, s.end_date, g.name AS gym_name,
                   CASE
                       WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() THEN 'Actif'
                       WHEN s.payment_status = 'confirmed' AND s.end_date < CURDATE() THEN 'Expiré'
                       WHEN s.payment_status = 'pending' THEN 'En attente'
                       ELSE 'Inactif'
                   END AS status,
                   TIMESTAMPDIFF(DAY, CURDATE(), s.end_date) AS days_remaining
            FROM users u
            JOIN subscriptions s ON u.id = s.member_id
            JOIN gyms g ON s.gym_id = g.id
            WHERE s.gym_id IN ({','.join(gym_ids)})
            ORDER BY
                CASE WHEN s.payment_status = 'confirmed' AND s.end_date < CURDATE() THEN 0
                     WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() AND TIMESTAMPDIFF(DAY, CURDATE(), s.end_date) <= 7 THEN 1
                     ELSE 2 END,
                u.last_name, u.first_name
        """, fetch_all=True)
        
        if not members:
            label = QLabel("Aucun membre trouvé")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.members_layout.addWidget(label)
        else:
            for member in members:
                self.add_member_card(member)
    
    def search_members(self):
        search_text = self.member_search_input.text().strip().lower()
        
        for i in reversed(range(self.members_layout.count())):
            widget = self.members_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()
        
        gyms = DatabaseManager.execute_query(
            "SELECT id FROM gyms WHERE owner_id = %s", 
            (self.user_id,), 
            fetch_all=True
        )
        
        if not gyms:
            label = QLabel("Vous n'avez pas encore de salle enregistrée")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.members_layout.addWidget(label)
            return
        
        gym_ids = [str(gym['id']) for gym in gyms]
        
        members = DatabaseManager.execute_query(f"""
            SELECT DISTINCT u.id, u.first_name, u.last_name, u.email, u.phone,
                   s.payment_status, s.subscription_type, s.end_date, g.name AS gym_name
            FROM users u
            JOIN subscriptions s ON u.id = s.member_id
            JOIN gyms g ON s.gym_id = g.id
            WHERE s.gym_id IN ({','.join(gym_ids)})
            AND (LOWER(u.first_name) LIKE %s OR LOWER(u.last_name) LIKE %s OR LOWER(u.email) LIKE %s)
            ORDER BY u.last_name, u.first_name
        """, (f'%{search_text}%', f'%{search_text}%', f'%{search_text}%'), fetch_all=True)
        
        if not members:
            label = QLabel("Aucun membre trouvé")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.members_layout.addWidget(label)
        else:
            for member in members:
                self.add_member_card(member)
    
    def add_member_card(self, member):
        card = QFrame()

        # Dynamic styling based on subscription status
        if member['status'] == 'Actif':
            if member['days_remaining'] is not None and member['days_remaining'] <= 7:
                border_color = "rgba(255, 193, 7, 0.6)"  # Warning yellow for expiring soon
                hover_border = "rgba(255, 193, 7, 0.8)"
            else:
                border_color = "rgba(40, 167, 69, 0.4)"  # Green for active
                hover_border = "rgba(40, 167, 69, 0.6)"
        elif member['status'] == 'Expiré':
            border_color = "rgba(220, 53, 69, 0.6)"  # Red for expired
            hover_border = "rgba(220, 53, 69, 0.8)"
        elif member['status'] == 'En attente':
            border_color = "rgba(255, 193, 7, 0.4)"  # Yellow for pending
            hover_border = "rgba(255, 193, 7, 0.6)"
        else:
            border_color = "rgba(108, 117, 125, 0.4)"  # Gray for inactive
            hover_border = "rgba(108, 117, 125, 0.6)"

        card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(45, 45, 45, 0.9),
                    stop:1 rgba(35, 35, 35, 0.9));
                border-radius: 12px;
                border: 2px solid {border_color};
            }}
            QFrame:hover {{
                border: 2px solid {hover_border};
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(50, 50, 50, 0.9),
                    stop:1 rgba(40, 40, 40, 0.9));
            }}
        """)
        card.setFixedHeight(200)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(8)

        # Member info header
        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        name = QLabel(f"{member['first_name']} {member['last_name']}")
        name.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: white;
            background: transparent;
        """)

        # Enhanced status badge
        status_badge = QLabel(member['status'])
        if member['status'] == 'Actif':
            if member['days_remaining'] is not None and member['days_remaining'] <= 7:
                status_badge.setStyleSheet("""
                    background: rgba(255, 193, 7, 0.9);
                    color: black;
                    padding: 4px 8px;
                    border-radius: 12px;
                    font-size: 12px;
                    font-weight: bold;
                """)
            else:
                status_badge.setStyleSheet("""
                    background: rgba(40, 167, 69, 0.9);
                    color: white;
                    padding: 4px 8px;
                    border-radius: 12px;
                    font-size: 12px;
                    font-weight: bold;
                """)
        elif member['status'] == 'Expiré':
            status_badge.setStyleSheet("""
                background: rgba(220, 53, 69, 0.9);
                color: white;
                padding: 4px 8px;
                border-radius: 12px;
                font-size: 12px;
                font-weight: bold;
            """)
        elif member['status'] == 'En attente':
            status_badge.setStyleSheet("""
                background: rgba(255, 193, 7, 0.9);
                color: black;
                padding: 4px 8px;
                border-radius: 12px;
                font-size: 12px;
                font-weight: bold;
            """)
        else:
            status_badge.setStyleSheet("""
                background: rgba(108, 117, 125, 0.9);
                color: white;
                padding: 4px 8px;
                border-radius: 12px;
                font-size: 12px;
                font-weight: bold;
            """)

        header_layout.addWidget(name)
        header_layout.addStretch()
        header_layout.addWidget(status_badge)
        layout.addWidget(header)

        # Contact info
        contact = QLabel(f"📞 {member['phone']} | ✉️ {member['email']}")
        contact.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8); background: transparent;")
        layout.addWidget(contact)

        # Enhanced subscription info with expiration warnings
        if member['status'] == 'Actif' and member['days_remaining'] is not None:
            if member['days_remaining'] <= 7:
                date_icon = "⚠️"
                date_text = f"Expire dans {member['days_remaining']} jour(s)"
                date_color = "rgba(255, 193, 7, 1)"
            else:
                date_icon = "✅"
                date_text = f"{member['days_remaining']} jours restants"
                date_color = "rgba(40, 167, 69, 1)"
        elif member['status'] == 'Expiré':
            date_icon = "❌"
            expired_days = abs(member['days_remaining']) if member['days_remaining'] else 0
            date_text = f"Expiré depuis {expired_days} jour(s)"
            date_color = "rgba(220, 53, 69, 1)"
        else:
            date_icon = "📅"
            date_text = f"Fin: {member['end_date']}"
            date_color = "rgba(255, 255, 255, 0.8)"

        sub_info = QLabel(
            f"📍 {member['gym_name']}\n"
            f"🏋️ Abonnement: {member['subscription_type'].capitalize()}\n"
            f"{date_icon} {date_text}"
        )
        sub_info.setStyleSheet(f"font-size: 14px; color: {date_color}; margin-top: 5px; background: transparent; font-weight: 500;")
        layout.addWidget(sub_info)

        # Action buttons for expired subscriptions
        if member['status'] == 'Expiré':
            action_container = QWidget()
            action_layout = QHBoxLayout(action_container)
            action_layout.setContentsMargins(0, 5, 0, 0)

            notify_btn = QPushButton("📧 Notifier le membre")
            notify_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 rgba(255, 193, 7, 0.8), stop:1 rgba(255, 152, 0, 0.8));
                    color: black;
                    padding: 6px 12px;
                    border-radius: 6px;
                    border: none;
                    font-weight: 600;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 rgba(255, 193, 7, 1), stop:1 rgba(255, 152, 0, 1));
                }
            """)
            notify_btn.clicked.connect(lambda: self.notify_expired_member(member))

            action_layout.addWidget(notify_btn)
            action_layout.addStretch()
            layout.addWidget(action_container)

        self.members_layout.addWidget(card)

    def notify_expired_member(self, member):
        """Notify member about expired subscription"""
        QMessageBox.information(
            self,
            "Notification envoyée",
            f"Une notification de renouvellement a été envoyée à {member['first_name']} {member['last_name']} "
            f"({member['email']}) concernant l'expiration de son abonnement."
        )
    
    def load_nearby_gyms(self):
        for i in reversed(range(self.nearby_gyms_layout.count())):
            widget = self.nearby_gyms_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        gyms = DatabaseManager.execute_query("SELECT * FROM gyms", fetch_all=True)
        
        if not gyms:
            label = QLabel("Aucune salle trouvée à proximité")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.nearby_gyms_layout.addWidget(label)
        else:
            for gym in gyms:
                card = QFrame()
                card.setStyleSheet("""
                    QFrame {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(45, 45, 45, 0.9),
                            stop:1 rgba(35, 35, 35, 0.9));
                        border-radius: 12px;
                        border: 1px solid rgba(255, 107, 53, 0.2);
                    }
                    QFrame:hover {
                        border: 1px solid rgba(255, 107, 53, 0.4);
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(50, 50, 50, 0.9),
                            stop:1 rgba(40, 40, 40, 0.9));
                    }
                    QLabel {
                        color: white;
                        background: transparent;
                    }
                """)
                card_layout = QVBoxLayout(card)
                card_layout.setContentsMargins(20, 15, 20, 15)
                
                name = QLabel(gym['name'])
                name.setStyleSheet("font-size: 16px; font-weight: bold; color: white; background: transparent;")

                address = QLabel(gym['address'])
                address.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8); background: transparent;")
                
                btn = QPushButton("S'abonner")
                btn.setStyleSheet("""
                    background-color: #002347;
                    color: white;
                    padding: 8px;
                    border-radius: 5px;
                """)
                btn.clicked.connect(lambda _, gid=gym['id']: self.subscribe_to_gym(gid))
                
                card_layout.addWidget(name)
                card_layout.addWidget(address)
                card_layout.addWidget(btn)
                
                self.nearby_gyms_layout.addWidget(card)
    
    def subscribe_to_gym(self, gym_id):
        dialog = SubscriptionDialog(gym_id, self.user_id, self)
        dialog.exec()
        self.load_nearby_gyms()
    
    def load_subscriptions(self):
        for i in reversed(range(self.subscriptions_layout.count())):
            widget = self.subscriptions_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        subscriptions = DatabaseManager.execute_query("""
            SELECT s.*, g.name as gym_name,
                   CASE
                       WHEN s.payment_status = 'confirmed' AND s.end_date >= CURDATE() THEN 'Actif'
                       WHEN s.payment_status = 'confirmed' AND s.end_date < CURDATE() THEN 'Expiré'
                       WHEN s.payment_status = 'pending' THEN 'En attente'
                       ELSE 'Inactif'
                   END AS status,
                   TIMESTAMPDIFF(DAY, CURDATE(), s.end_date) AS days_remaining
            FROM subscriptions s
            JOIN gyms g ON s.gym_id = g.id
            WHERE s.member_id = %s
            ORDER BY s.end_date DESC
        """, (self.user_id,), fetch_all=True)

        if not subscriptions:
            label = QLabel("Vous n'avez aucun abonnement")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.subscriptions_layout.addWidget(label)
        else:
            for sub in subscriptions:
                card = QFrame()

                # Dynamic styling based on subscription status
                if sub['status'] == 'Actif':
                    border_color = "rgba(40, 167, 69, 0.4)"  # Green for active
                    hover_border = "rgba(40, 167, 69, 0.6)"
                elif sub['status'] == 'Expiré':
                    border_color = "rgba(220, 53, 69, 0.4)"  # Red for expired
                    hover_border = "rgba(220, 53, 69, 0.6)"
                elif sub['status'] == 'En attente':
                    border_color = "rgba(255, 193, 7, 0.4)"  # Yellow for pending
                    hover_border = "rgba(255, 193, 7, 0.6)"
                else:
                    border_color = "rgba(108, 117, 125, 0.4)"  # Gray for inactive
                    hover_border = "rgba(108, 117, 125, 0.6)"

                card.setStyleSheet(f"""
                    QFrame {{
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(45, 45, 45, 0.9),
                            stop:1 rgba(35, 35, 35, 0.9));
                        border-radius: 12px;
                        border: 2px solid {border_color};
                    }}
                    QFrame:hover {{
                        border: 2px solid {hover_border};
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(50, 50, 50, 0.9),
                            stop:1 rgba(40, 40, 40, 0.9));
                    }}
                    QLabel {{
                        color: white;
                        background: transparent;
                    }}
                """)
                card_layout = QVBoxLayout(card)
                card_layout.setContentsMargins(20, 15, 20, 15)

                # Header with gym name and status
                header_layout = QHBoxLayout()
                gym = QLabel(sub['gym_name'])
                gym.setStyleSheet("font-size: 16px; font-weight: bold; color: white; background: transparent;")

                # Status badge
                status_badge = QLabel(sub['status'])
                if sub['status'] == 'Actif':
                    status_badge.setStyleSheet("""
                        background: rgba(40, 167, 69, 0.8);
                        color: white;
                        padding: 4px 8px;
                        border-radius: 12px;
                        font-size: 12px;
                        font-weight: bold;
                    """)
                elif sub['status'] == 'Expiré':
                    status_badge.setStyleSheet("""
                        background: rgba(220, 53, 69, 0.8);
                        color: white;
                        padding: 4px 8px;
                        border-radius: 12px;
                        font-size: 12px;
                        font-weight: bold;
                    """)
                elif sub['status'] == 'En attente':
                    status_badge.setStyleSheet("""
                        background: rgba(255, 193, 7, 0.8);
                        color: black;
                        padding: 4px 8px;
                        border-radius: 12px;
                        font-size: 12px;
                        font-weight: bold;
                    """)
                else:
                    status_badge.setStyleSheet("""
                        background: rgba(108, 117, 125, 0.8);
                        color: white;
                        padding: 4px 8px;
                        border-radius: 12px;
                        font-size: 12px;
                        font-weight: bold;
                    """)

                header_layout.addWidget(gym)
                header_layout.addStretch()
                header_layout.addWidget(status_badge)

                sub_type = QLabel(f"Abonnement: {sub['subscription_type'].capitalize()}")
                sub_type.setStyleSheet("font-size: 14px; color: rgba(255, 255, 255, 0.8); background: transparent;")

                # Enhanced date information
                if sub['status'] == 'Actif' and sub['days_remaining'] is not None:
                    if sub['days_remaining'] <= 7:
                        date_color = "rgba(255, 193, 7, 1)"  # Warning yellow
                        date_text = f"⚠️ Expire dans {sub['days_remaining']} jour(s) - {sub['end_date']}"
                    else:
                        date_color = "rgba(40, 167, 69, 1)"  # Success green
                        date_text = f"✅ Expire le {sub['end_date']} ({sub['days_remaining']} jours restants)"
                elif sub['status'] == 'Expiré':
                    date_color = "rgba(220, 53, 69, 1)"  # Danger red
                    expired_days = abs(sub['days_remaining']) if sub['days_remaining'] else 0
                    date_text = f"❌ Expiré depuis {expired_days} jour(s) - {sub['end_date']}"
                else:
                    date_color = "rgba(255, 255, 255, 0.8)"
                    date_text = f"Fin: {sub['end_date']}"

                end_date = QLabel(date_text)
                end_date.setStyleSheet(f"font-size: 14px; color: {date_color}; background: transparent; font-weight: 500;")

                # Action buttons
                button_layout = QHBoxLayout()

                # Renew/Subscribe button for expired or active subscriptions
                if sub['status'] in ['Expiré', 'Actif']:
                    renew_btn = QPushButton("🔄 Renouveler" if sub['status'] == 'Expiré' else "➕ Prolonger")
                    renew_btn.setStyleSheet("""
                        QPushButton {
                            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 rgba(40, 167, 69, 0.8), stop:1 rgba(32, 201, 151, 0.8));
                            color: white;
                            padding: 8px 16px;
                            border-radius: 8px;
                            border: none;
                            font-weight: 600;
                        }
                        QPushButton:hover {
                            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 rgba(40, 167, 69, 1), stop:1 rgba(32, 201, 151, 1));
                        }
                    """)
                    renew_btn.clicked.connect(lambda _, gid=sub['gym_id']: self.renew_subscription(gid))
                    button_layout.addWidget(renew_btn)

                button_layout.addStretch()

                card_layout.addLayout(header_layout)
                card_layout.addWidget(sub_type)
                card_layout.addWidget(end_date)
                card_layout.addLayout(button_layout)

                self.subscriptions_layout.addWidget(card)

    def renew_subscription(self, gym_id):
        """Allow members to renew or extend their subscription"""
        dialog = SubscriptionDialog(gym_id, self.user_id, self, is_renewal=True)
        if dialog.exec() == QDialog.Accepted:
            self.load_subscriptions()  # Refresh the subscription list

    def logout(self):
        self.close()
        self.login_window = LoginWindow()
        self.login_window.show()

    def show_member_workouts(self):
        dialog = MemberWorkoutHistoryDialog(self.user_id, self)
        dialog.exec()

    def load_nearby_gyms(self):
        for i in reversed(range(self.nearby_gyms_layout.count())):
            widget = self.nearby_gyms_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        # Add search bar at the top (keeping existing style)
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 15)
        
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Rechercher par adresse (ex: Sidi Bernoussi)")
        if self.address:
            self.address_input.setText(self.address)
        self.address_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #DDD;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FFC107;
            }
        """)
        
        search_btn = QPushButton("Rechercher")
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                font-size: 14px;
                padding: 12px 20px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        search_btn.clicked.connect(self.search_gyms_by_address)
        search_layout.addWidget(self.address_input)
        search_layout.addWidget(search_btn)
        self.nearby_gyms_layout.addWidget(search_container)
        
        # Load gyms based on current address
        self.search_gyms_by_address()

    def load_nearby_gyms(self):
        for i in reversed(range(self.nearby_gyms_layout.count())):
            widget = self.nearby_gyms_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        # Barre de recherche avec autocomplétion
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 15)
        
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Entrez une adresse (ex: Sidi Bernoussi)")
        if self.address:
            self.address_input.setText(self.address)
        
        # Configuration de l'autocomplétion dynamique
        self.setup_address_autocompletion()
        
        self.address_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #DDD;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FFC107;
            }
        """)
        
        search_btn = QPushButton("Rechercher")
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                font-size: 14px;
                padding: 12px 20px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        search_btn.clicked.connect(self.search_gyms_by_address)
        search_layout.addWidget(self.address_input)
        search_layout.addWidget(search_btn)
        self.nearby_gyms_layout.addWidget(search_container)
        
        # Chargement initial avec l'adresse du membre
        self.search_gyms_by_address()

    def setup_address_autocompletion(self):
        # Récupérer toutes les adresses uniques des salles
        addresses = DatabaseManager.execute_query(
            "SELECT DISTINCT address FROM gyms", 
            fetch_all=True
        )
        
        if addresses:
            address_list = [addr['address'] for addr in addresses]
            self.completer = QCompleter(address_list, self)
            self.completer.setCaseSensitivity(Qt.CaseInsensitive)
            self.completer.setFilterMode(Qt.MatchContains)
            self.completer.setCompletionMode(QCompleter.PopupCompletion)
            self.address_input.setCompleter(self.completer)

    def search_gyms_by_address(self):
        # Nettoyer les résultats précédents (sauf la barre de recherche)
        for i in reversed(range(self.nearby_gyms_layout.count())):
            widget = self.nearby_gyms_layout.itemAt(i).widget()
            if widget and widget != self.nearby_gyms_layout.itemAt(0).widget():
                widget.deleteLater()
        
        search_text = self.address_input.text().strip().lower()
        gyms = DatabaseManager.execute_query("SELECT * FROM gyms", fetch_all=True)
        
        if not gyms:
            label = QLabel("Aucune salle de sport disponible")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.nearby_gyms_layout.addWidget(label)
            return
        
        # Filtrage par adresse
        if search_text:
            matched_gyms = []
            search_terms = search_text.split()
            
            for gym in gyms:
                gym_address = gym['address'].lower()
                # Vérifie que tous les termes de recherche sont dans l'adresse
                if all(term in gym_address for term in search_terms):
                    matched_gyms.append(gym)
            
            gyms = matched_gyms
        
        if not gyms:
            label = QLabel(f"Aucune salle trouvée pour '{search_text}'")
            label.setStyleSheet("font-size: 16px; color: #666;")
            label.setAlignment(Qt.AlignCenter)
            self.nearby_gyms_layout.addWidget(label)
        else:
            for gym in gyms:
                # Création de la carte pour chaque salle
                card = QFrame()
                card.setStyleSheet("""
                    QFrame {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(45, 45, 45, 0.9),
                            stop:1 rgba(35, 35, 35, 0.9));
                        border-radius: 12px;
                        border: 1px solid rgba(255, 107, 53, 0.2);
                    }
                    QFrame:hover {
                        border: 1px solid rgba(255, 107, 53, 0.4);
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(50, 50, 50, 0.9),
                            stop:1 rgba(40, 40, 40, 0.9));
                    }
                    QLabel {
                        color: white;
                        background: transparent;
                    }
                """)
                card.setFixedHeight(180)
                
                layout = QVBoxLayout(card)
                layout.setContentsMargins(20, 15, 20, 15)
                layout.setSpacing(10)

                # En-tête avec nom et note
                header = QWidget()
                header_layout = QHBoxLayout(header)
                header_layout.setContentsMargins(0, 0, 0, 0)
                
                name = QLabel(gym['name'])
                name.setStyleSheet("""
                    font-size: 18px;
                    font-weight: 600;
                    color: white;
                    background: transparent;
                """)

                # Note moyenne
                avg_rating = DatabaseManager.get_gym_avg_rating(gym['id'])
                rating_text = QLabel(f"{avg_rating['avg']} ★ ({avg_rating['count']})")
                rating_text.setStyleSheet("color: #FFC107; font-weight: bold; background: transparent;")

                header_layout.addWidget(name)
                header_layout.addStretch()
                header_layout.addWidget(rating_text)
                layout.addWidget(header)

                # Adresse
                address = QLabel(gym['address'])
                address.setStyleSheet("""
                    font-size: 14px;
                    color: rgba(255, 255, 255, 0.8);
                    margin-top: 5px;
                    background: transparent;
                """)
                layout.addWidget(address)
                layout.addStretch()

                # Boutons
                btn_container = QWidget()
                btn_layout = QHBoxLayout(btn_container)
                btn_layout.setContentsMargins(0, 0, 0, 0)

                # Bouton avis
                reviews_btn = QPushButton("Avis")
                reviews_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #f5f5f5;
                        color: #002347;
                        font-size: 13px;
                        padding: 8px 12px;
                        border-radius: 5px;
                        border: 1px solid #e0e0e0;
                    }
                    QPushButton:hover {
                        background-color: #e0e0e0;
                    }
                """)
                reviews_btn.clicked.connect(lambda _, gid=gym['id']: self.show_reviews(gid))

                # Bouton s'abonner
                subscribe_btn = QPushButton("S'abonner")
                subscribe_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #002347;
                        color: white;
                        font-size: 13px;
                        padding: 8px 12px;
                        border-radius: 5px;
                    }
                    QPushButton:hover {
                        background-color: #003366;
                    }
                """)
                subscribe_btn.clicked.connect(lambda _, gid=gym['id']: self.subscribe_to_gym(gid))

                btn_layout.addWidget(reviews_btn)
                btn_layout.addStretch()
                btn_layout.addWidget(subscribe_btn)
                layout.addWidget(btn_container)

                self.nearby_gyms_layout.addWidget(card)

    def show_reviews(self, gym_id):
        dialog = ReviewsDialog(gym_id, self)
        dialog.exec()

    def add_review(self, gym_id):
        # Check if user already reviewed this gym
        existing_review = DatabaseManager.execute_query(
            "SELECT id FROM gym_reviews WHERE gym_id = %s AND member_id = %s",
            (gym_id, self.user_id),
            fetch_one=True
        )
        
        if existing_review:
            QMessageBox.information(self, "Information", "Vous avez déjà évalué cette salle")
            return
        
        dialog = AddReviewDialog(gym_id, self.user_id, self)
        dialog.exec()
        self.load_nearby_gyms()  # Refresh to show updated ratings
    
    def subscribe_to_gym(self, gym_id):
            # Check for active subscriptions
            active_sub = DatabaseManager.execute_query("""
                SELECT id, end_date, subscription_type
                FROM subscriptions
                WHERE member_id = %s AND gym_id = %s
                AND payment_status = 'confirmed'
                AND end_date >= CURDATE()
                ORDER BY end_date DESC
                LIMIT 1
            """, (self.user_id, gym_id), fetch_one=True)

            if active_sub:
                # Show option to extend/renew instead of blocking
                reply = QMessageBox.question(
                    self,
                    "Abonnement existant",
                    f"Vous avez déjà un abonnement {active_sub['subscription_type']} actif jusqu'au {active_sub['end_date']}.\n\n"
                    "Voulez-vous le prolonger ou souscrire un nouvel abonnement ?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )

                if reply == QMessageBox.No:
                    return

                # Open renewal dialog
                dialog = SubscriptionDialog(gym_id, self.user_id, self, is_renewal=True)
            else:
                # No active subscription, open regular subscription dialog
                dialog = SubscriptionDialog(gym_id, self.user_id, self)

            if dialog.exec() == QDialog.Accepted:
                self.load_nearby_gyms()

    def show_statistics(self):
        try:
            print("📊 show_statistics called")

            # Check if we're a gym owner
            if self.user_role != "gym owner":
                print(f"❌ User role is {self.user_role}, not gym owner")
                QMessageBox.information(self, "Information", "Les statistiques ne sont disponibles que pour les propriétaires de salles.")
                return

            print(f"✅ User is gym owner, switching to index 4")
            print(f"📋 Current stacked widget count: {self.stacked_widget.count()}")

            # Ensure the statistics view exists
            if self.stacked_widget.count() <= 4:  # Correct: Statistics view is at index 4
                print("❌ Statistics view not found, creating views...")
                self.create_dashboard_views()

            self.stacked_widget.setCurrentIndex(4)  # Correct: Statistics view is at index 4
            print(f"✅ Switched to statistics page (index 4)")

            # Load statistics immediately without timer to prevent multiple calls
            self.load_statistics()
            print("✅ Statistics loaded successfully")

        except Exception as e:
            print(f"❌ Error in show_statistics: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement des statistiques: {str(e)}")
            # Fallback to dashboard
            try:
                self.show_dashboard()
            except:
                pass

    def load_statistics(self):
        try:
            print("📈 load_statistics called")

            # Clear existing content from main stats layout
            if hasattr(self, 'main_stats_layout') and self.main_stats_layout is not None:
                print("🧹 Clearing existing statistics content...")
                for i in reversed(range(self.main_stats_layout.count())):
                    child = self.main_stats_layout.itemAt(i).widget()
                    if child:
                        child.setParent(None)
            else:
                print("❌ main_stats_layout not found")
                return

            print("📊 Creating unified statistics content...")
            # Create a modern scroll area for the statistics page
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setStyleSheet("""
                QScrollArea {
                    border: none;
                    background: transparent;
                }
                QScrollBar:vertical {
                    background: rgba(255, 107, 53, 0.1);
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: rgba(255, 107, 53, 0.5);
                    border-radius: 6px;
                    min-height: 20px;
                }
                QScrollBar::handle:vertical:hover {
                    background: rgba(255, 107, 53, 0.7);
                }
            """)

            content = QWidget()
            content.setStyleSheet("background: transparent;")
            content_layout = QVBoxLayout(content)
            content_layout.setContentsMargins(20, 20, 20, 20)
            content_layout.setSpacing(25)

            # Modern header section
            header_container = QWidget()
            header_container.setStyleSheet("""
                QWidget {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 rgba(255, 140, 0, 0.2),
                        stop:1 rgba(255, 107, 53, 0.2));
                    border-radius: 15px;
                    border: 1px solid rgba(255, 140, 0, 0.3);
                }
            """)
            header_layout = QVBoxLayout(header_container)
            header_layout.setContentsMargins(25, 20, 25, 20)

            # Main title
            title = QLabel("📊 STATISTIQUES DE PERFORMANCE")
            title.setStyleSheet("""
                font-size: 24px;
                font-weight: 800;
                color: white;
                text-align: center;
                letter-spacing: 1px;
                margin-bottom: 10px;
            """)
            title.setAlignment(Qt.AlignCenter)
            header_layout.addWidget(title)

            # Action buttons container
            actions_container = QWidget()
            actions_container.setStyleSheet("background: transparent;")
            actions_layout = QHBoxLayout(actions_container)
            actions_layout.setContentsMargins(0, 0, 0, 0)

            # Date range selector
            date_range_label = QLabel("📅 Période:")
            date_range_label.setStyleSheet("""
                font-size: 14px;
                font-weight: 600;
                color: white;
                background: transparent;
            """)

            # Don't recreate the combo box if it already exists - this prevents recursive calls
            if not hasattr(self, 'date_range_combo') or self.date_range_combo is None:
                self.date_range_combo = QComboBox()
                self.date_range_combo.addItems(["Dernier mois", "3 derniers mois", "6 derniers mois", "Cette année"])
                # Use a separate method to handle date range changes to prevent recursion
                self.date_range_combo.currentTextChanged.connect(self.on_date_range_changed)

            self.date_range_combo.setStyleSheet("""
                QComboBox {
                    background: rgba(35, 35, 35, 0.8);
                    border: 2px solid rgba(255, 107, 53, 0.3);
                    border-radius: 8px;
                    padding: 8px 12px;
                    font-size: 13px;
                    color: white;
                    font-weight: 500;
                    min-width: 120px;
                }
                QComboBox:focus {
                    border: 2px solid rgba(255, 107, 53, 0.6);
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                    background: rgba(255, 107, 53, 0.2);
                    border-top-right-radius: 6px;
                    border-bottom-right-radius: 6px;
                }
                QComboBox QAbstractItemView {
                    background: rgba(35, 35, 35, 0.95);
                    border: 2px solid rgba(255, 107, 53, 0.4);
                    border-radius: 6px;
                    color: white;
                    selection-background-color: rgba(255, 107, 53, 0.3);
                }
            """)

            # Modern export button
            export_btn = QPushButton("📊 EXPORTER EXCEL")
            export_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #28a745, stop:1 #20c997);
                    color: white;
                    padding: 10px 20px;
                    border-radius: 10px;
                    font-size: 13px;
                    font-weight: 700;
                    border: none;
                    letter-spacing: 0.5px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #218838, stop:1 #1e7e34);
                    transform: translateY(-2px);
                }
            """)
            export_btn.clicked.connect(lambda: self.export_stats_to_excel(start_date_str, end_date_str))

            actions_layout.addWidget(date_range_label)
            actions_layout.addWidget(self.date_range_combo)
            actions_layout.addStretch()
            actions_layout.addWidget(export_btn)
            header_layout.addWidget(actions_container)
            content_layout.addWidget(header_container)

            # Gym Filter Section
            filter_container = QWidget()
            filter_container.setStyleSheet("""
                background: rgba(25, 25, 25, 0.9);
                border-radius: 16px;
                border: 2px solid rgba(255, 107, 53, 0.3);
                backdrop-filter: blur(10px);
            """)
            filter_layout = QHBoxLayout(filter_container)
            filter_layout.setContentsMargins(25, 20, 25, 20)
            filter_layout.setSpacing(20)

            # Gym selector
            gym_label = QLabel("🏢 Sélectionner une salle:")
            gym_label.setStyleSheet("""
                font-size: 16px;
                font-weight: 600;
                color: white;
                background: transparent;
            """)

            # Create gym filter combo only if it doesn't exist
            if not hasattr(self, 'gym_filter_combo'):
                self.gym_filter_combo = QComboBox()
                self.gym_filter_combo.setMinimumWidth(200)
                self.gym_filter_combo.setStyleSheet("""
                    QComboBox {
                        background: rgba(35, 35, 35, 0.9);
                        border: 2px solid rgba(255, 107, 53, 0.4);
                        border-radius: 12px;
                        padding: 12px 15px;
                        font-size: 14px;
                        color: white;
                        font-weight: 500;
                    }
                    QComboBox:focus {
                        border: 2px solid rgba(255, 107, 53, 0.8);
                        background: rgba(45, 45, 45, 0.9);
                    }
                    QComboBox::drop-down {
                        border: none;
                        width: 30px;
                        background: rgba(255, 107, 53, 0.3);
                        border-top-right-radius: 10px;
                        border-bottom-right-radius: 10px;
                    }
                    QComboBox::down-arrow {
                        image: none;
                        border-left: 6px solid transparent;
                        border-right: 6px solid transparent;
                        border-top: 6px solid white;
                        margin-right: 10px;
                    }
                    QComboBox QAbstractItemView {
                        background: rgba(35, 35, 35, 0.95);
                        border: 2px solid rgba(255, 107, 53, 0.4);
                        border-radius: 8px;
                        color: white;
                        selection-background-color: rgba(255, 107, 53, 0.4);
                        padding: 5px;
                    }
                """)

                # Load gyms for filter and connect signal
                self.load_gym_filter()
                self.gym_filter_combo.currentTextChanged.connect(self.on_gym_filter_changed)

            filter_layout.addWidget(gym_label)
            filter_layout.addWidget(self.gym_filter_combo)
            filter_layout.addStretch()
            content_layout.addWidget(filter_container)



            # Statistics Charts Section
            charts_container = QWidget()
            charts_container.setStyleSheet("""
                background: rgba(25, 25, 25, 0.9);
                border-radius: 16px;
                border: 2px solid rgba(255, 107, 53, 0.3);
                backdrop-filter: blur(10px);
            """)
            charts_layout = QVBoxLayout(charts_container)
            charts_layout.setContentsMargins(30, 25, 30, 25)
            charts_layout.setSpacing(30)

            # Title
            charts_title = QLabel("📊 ANALYSE DES DONNÉES")
            charts_title.setStyleSheet("""
                font-size: 20px;
                font-weight: 800;
                color: white;
                background: transparent;
                letter-spacing: 1px;
                margin-bottom: 10px;
            """)
            charts_title.setAlignment(Qt.AlignCenter)
            charts_layout.addWidget(charts_title)

            # Clear existing charts first
            for i in reversed(range(charts_layout.count())):
                child = charts_layout.itemAt(i)
                if child and child.widget() and child.widget() != charts_title:
                    child.widget().setParent(None)

            # Charts Grid
            charts_grid = QHBoxLayout()
            charts_grid.setSpacing(40)

            try:
                print("📊 Creating circle charts...")

                # Get selected gym ID
                selected_gym_id = self.get_selected_gym_id()
                print(f"🏢 Selected gym ID: {selected_gym_id}")

                # Subscription types circle chart
                try:
                    print(f"📊 Loading subscription types data for gym: {selected_gym_id}")
                    if selected_gym_id == "all":
                        sub_types = DatabaseManager.execute_query("""
                            SELECT s.subscription_type, COUNT(*) AS count
                            FROM subscriptions s
                            JOIN gyms g ON s.gym_id = g.id
                            WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
                            GROUP BY s.subscription_type
                        """, (self.user_id,), fetch_all=True)
                    else:
                        sub_types = DatabaseManager.execute_query("""
                            SELECT s.subscription_type, COUNT(*) AS count
                            FROM subscriptions s
                            WHERE s.gym_id = %s AND s.payment_status = 'confirmed'
                            GROUP BY s.subscription_type
                        """, (selected_gym_id,), fetch_all=True)

                    if sub_types:
                        labels = [row['subscription_type'].capitalize() for row in sub_types]
                        data = [row['count'] for row in sub_types]
                        colors = ["#ff6b35", "#28a745", "#20c997", "#17a2b8", "#6f42c1"]
                        print(f"📊 Subscription types data: {len(labels)} entries - {data}")
                        subscription_chart = self.create_smooth_circle_chart("📊 Types d'Abonnements", labels, data, colors)
                    else:
                        print("📊 No subscription types data found")
                        subscription_chart = self.create_smooth_circle_chart("📊 Types d'Abonnements", [], [], ["#ff6b35"])

                    charts_grid.addWidget(subscription_chart)
                except Exception as e:
                    print(f"Error loading subscription types: {e}")
                    import traceback
                    traceback.print_exc()
                    subscription_chart = self.create_smooth_circle_chart("📊 Types d'Abonnements", [], [], ["#ff6b35"])
                    charts_grid.addWidget(subscription_chart)

                # Gender distribution circle chart
                try:
                    print(f"👥 Loading gender distribution data for gym: {selected_gym_id}")
                    if selected_gym_id == "all":
                        gender_data = DatabaseManager.execute_query("""
                            SELECT u.gender, COUNT(DISTINCT u.id) AS count
                            FROM users u
                            JOIN subscriptions s ON u.id = s.member_id
                            JOIN gyms g ON s.gym_id = g.id
                            WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
                            AND u.gender IS NOT NULL AND u.gender != ''
                            GROUP BY u.gender
                        """, (self.user_id,), fetch_all=True)
                    else:
                        gender_data = DatabaseManager.execute_query("""
                            SELECT u.gender, COUNT(DISTINCT u.id) AS count
                            FROM users u
                            JOIN subscriptions s ON u.id = s.member_id
                            WHERE s.gym_id = %s AND s.payment_status = 'confirmed'
                            AND u.gender IS NOT NULL AND u.gender != ''
                            GROUP BY u.gender
                        """, (selected_gym_id,), fetch_all=True)

                    if gender_data:
                        labels = [row['gender'].capitalize() for row in gender_data]
                        data = [row['count'] for row in gender_data]
                        colors = ["#ff6b35", "#28a745", "#20c997"]
                        print(f"👥 Gender distribution data: {len(labels)} entries - {data}")
                        gender_chart = self.create_smooth_circle_chart("👥 Répartition par Genre", labels, data, colors)
                    else:
                        print("👥 No gender distribution data found")
                        gender_chart = self.create_smooth_circle_chart("👥 Répartition par Genre", [], [], ["#ff6b35"])

                    charts_grid.addWidget(gender_chart)
                except Exception as e:
                    print(f"Error loading gender distribution: {e}")
                    import traceback
                    traceback.print_exc()
                    gender_chart = self.create_smooth_circle_chart("👥 Répartition par Genre", [], [], ["#ff6b35"])
                    charts_grid.addWidget(gender_chart)

                charts_layout.addLayout(charts_grid)
                print("✅ Circle charts created successfully")
            except Exception as e:
                print(f"❌ Critical error creating charts: {e}")
                import traceback
                traceback.print_exc()
                error_label = QLabel(f"Erreur lors de la création des graphiques: {str(e)}")
                error_label.setStyleSheet("color: #ff6b35; font-size: 14px; padding: 20px;")
                charts_layout.addWidget(error_label)

            content_layout.addWidget(charts_container)

            # Revenue Summary Section
            revenue_container = QWidget()
            revenue_container.setStyleSheet("""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(40, 167, 69, 0.9), stop:1 rgba(32, 201, 151, 0.9));
                border-radius: 16px;
                border: 2px solid rgba(40, 167, 69, 0.4);
                backdrop-filter: blur(10px);
            """)
            revenue_layout = QVBoxLayout(revenue_container)
            revenue_layout.setContentsMargins(30, 20, 30, 20)
            revenue_layout.setSpacing(10)

            # Get revenue data
            selected_gym_id = self.get_selected_gym_id()
            try:
                if selected_gym_id == "all":
                    revenue_result = DatabaseManager.execute_query("""
                        SELECT SUM(
                            CASE
                                WHEN s.subscription_type = 'monthly' THEN
                                    (SELECT JSON_EXTRACT(g.subscriptions, '$.monthly.price') FROM gyms g WHERE g.id = s.gym_id)
                                WHEN s.subscription_type = 'quarterly' THEN
                                    (SELECT JSON_EXTRACT(g.subscriptions, '$.quarterly.price') FROM gyms g WHERE g.id = s.gym_id)
                                WHEN s.subscription_type = 'annual' THEN
                                    (SELECT JSON_EXTRACT(g.subscriptions, '$.annual.price') FROM gyms g WHERE g.id = s.gym_id)
                                ELSE 0
                            END
                        ) as total_revenue
                        FROM subscriptions s
                        JOIN gyms g ON s.gym_id = g.id
                        WHERE g.owner_id = %s AND s.payment_status = 'confirmed'
                    """, (self.user_id,), fetch_one=True)
                else:
                    revenue_result = DatabaseManager.execute_query("""
                        SELECT SUM(
                            CASE
                                WHEN s.subscription_type = 'monthly' THEN
                                    JSON_EXTRACT(g.subscriptions, '$.monthly.price')
                                WHEN s.subscription_type = 'quarterly' THEN
                                    JSON_EXTRACT(g.subscriptions, '$.quarterly.price')
                                WHEN s.subscription_type = 'annual' THEN
                                    JSON_EXTRACT(g.subscriptions, '$.annual.price')
                                ELSE 0
                            END
                        ) as total_revenue
                        FROM subscriptions s
                        JOIN gyms g ON s.gym_id = g.id
                        WHERE s.gym_id = %s AND s.payment_status = 'confirmed'
                    """, (selected_gym_id,), fetch_one=True)

                total_revenue = float(revenue_result['total_revenue']) if revenue_result and revenue_result['total_revenue'] else 0.0
            except Exception as e:
                print(f"Error calculating revenue: {e}")
                total_revenue = 0.0

            # Revenue title
            revenue_title = QLabel("💰 REVENUS TOTAUX")
            revenue_title.setStyleSheet("""
                font-size: 18px;
                font-weight: 800;
                color: white;
                background: transparent;
                letter-spacing: 1px;
                margin-bottom: 5px;
            """)
            revenue_title.setAlignment(Qt.AlignCenter)
            revenue_layout.addWidget(revenue_title)

            # Revenue amount
            revenue_amount = QLabel(f"€{total_revenue:,.2f}")
            revenue_amount.setStyleSheet("""
                font-size: 32px;
                font-weight: 900;
                color: white;
                background: transparent;
                letter-spacing: 2px;
            """)
            revenue_amount.setAlignment(Qt.AlignCenter)
            revenue_layout.addWidget(revenue_amount)

            # Revenue subtitle
            gym_text = "toutes les salles" if selected_gym_id == "all" else "cette salle"
            revenue_subtitle = QLabel(f"Revenus générés par {gym_text}")
            revenue_subtitle.setStyleSheet("""
                font-size: 14px;
                font-weight: 500;
                color: rgba(255, 255, 255, 0.9);
                background: transparent;
                letter-spacing: 0.5px;
            """)
            revenue_subtitle.setAlignment(Qt.AlignCenter)
            revenue_layout.addWidget(revenue_subtitle)

            content_layout.addWidget(revenue_container)

            scroll.setWidget(content)

            # Add to main stats layout instead of charts_layout
            if hasattr(self, 'main_stats_layout') and self.main_stats_layout is not None:
                self.main_stats_layout.addWidget(scroll)
                print("✅ Statistics content added to main layout")
            else:
                print("❌ main_stats_layout not available")

        except Exception as e:
            print(f"❌ Critical error in load_statistics: {e}")
            import traceback
            traceback.print_exc()
            # Create a simple error message widget
            error_widget = QLabel(f"Erreur lors du chargement des statistiques:\n{str(e)}")
            error_widget.setStyleSheet("""
                color: #ff6b35;
                font-size: 16px;
                padding: 30px;
                background: rgba(45, 45, 45, 0.8);
                border-radius: 10px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            """)
            error_widget.setAlignment(Qt.AlignCenter)
            error_widget.setWordWrap(True)
            if hasattr(self, 'main_stats_layout') and self.main_stats_layout is not None:
                self.main_stats_layout.addWidget(error_widget)



    def load_gym_filter(self):
        """Load gyms for the filter dropdown"""
        try:
            self.gym_filter_combo.clear()
            self.gym_filter_combo.addItem("🏢 Toutes les salles", "all")

            gyms = DatabaseManager.execute_query("""
                SELECT id, name FROM gyms WHERE owner_id = %s ORDER BY name
            """, (self.user_id,), fetch_all=True)

            if gyms:
                for gym in gyms:
                    self.gym_filter_combo.addItem(f"🏋️ {gym['name']}", gym['id'])
        except Exception as e:
            print(f"Error loading gym filter: {e}")

    def get_selected_gym_id(self):
        """Get the currently selected gym ID from filter"""
        try:
            return self.gym_filter_combo.currentData()
        except:
            return "all"

    def on_gym_filter_changed(self):
        """Handle gym filter changes"""
        try:
            selected_gym = self.gym_filter_combo.currentText()
            selected_gym_id = self.gym_filter_combo.currentData()
            print(f"🏢 Gym filter changed to: {selected_gym} (ID: {selected_gym_id})")

            # Store current selection to prevent reset
            self._current_gym_selection = (selected_gym, selected_gym_id)

            if hasattr(self, '_stats_reload_timer'):
                self._stats_reload_timer.stop()

            self._stats_reload_timer = QTimer()
            self._stats_reload_timer.setSingleShot(True)
            self._stats_reload_timer.timeout.connect(self.reload_statistics_only)
            self._stats_reload_timer.start(300)  # 300ms delay to prevent rapid calls
        except Exception as e:
            print(f"Error in on_gym_filter_changed: {e}")

    def reload_statistics_only(self):
        """Reload only the statistics charts without recreating the filter"""
        try:
            print("🔄 Reloading statistics charts only...")
            # Just reload the full statistics - the gym filter will be preserved
            self.load_statistics()
        except Exception as e:
            print(f"Error in reload_statistics_only: {e}")
            # Fallback to full reload if needed
            self.load_statistics()



    def on_date_range_changed(self):
        """Handle date range changes without causing recursive calls"""
        try:
            print("📅 Date range changed, reloading statistics...")
            # Use a timer to prevent rapid successive calls
            if hasattr(self, '_stats_reload_timer'):
                self._stats_reload_timer.stop()

            self._stats_reload_timer = QTimer()
            self._stats_reload_timer.setSingleShot(True)
            self._stats_reload_timer.timeout.connect(self.load_statistics)
            self._stats_reload_timer.start(300)  # 300ms delay to prevent rapid calls
        except Exception as e:
            print(f"Error in on_date_range_changed: {e}")

    def export_stats_to_excel(self, start_date, end_date):
        # Récupérer les données du tableau détaillé
        member_details = DatabaseManager.get_detailed_member_stats(self.user_id)
        
        if not member_details:
            QMessageBox.warning(self, "Aucune donnée", "Aucune donnée membre à exporter")
            return

        # Demander où sauvegarder le fichier
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer les statistiques membres",
            "",
            "Fichiers Excel (*.xlsx);;Tous les fichiers (*)"
        )
        
        if not file_path:
            return
        
        # Ajouter l'extension si nécessaire
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Détails Membres"
            
            # En-têtes
            headers = ["Nom", "Email", "Type Abonnement", "Salle", "Date Début", "Date Fin"]
            ws.append(headers)
            
            # Données
            for member in member_details:
                ws.append([
                    f"{member['first_name']} {member['last_name']}",
                    member['email'],
                    member['subscription_type'].capitalize(),
                    member['gym_name'],
                    str(member['start_date']),
                    str(member['end_date'])
                ])
            
            # Ajuster la largeur des colonnes
            from openpyxl.utils import get_column_letter
            for i, header in enumerate(headers, 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max(15, len(header) + 5)
            
            # Sauvegarder
            wb.save(file_path)
            QMessageBox.information(self, "Succès", f"Export réussi vers:\n{file_path}")
            
        except ImportError:
            QMessageBox.critical(
                self, 
                "Erreur", 
                "La fonction d'export Excel nécessite openpyxl.\n"
                "Installez-le avec: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export:\n{str(e)}")




    
    def create_professional_chart(self, title, chart_type, labels, data, colors, row, col):
        """Create a professional-looking chart widget with visual bars/indicators"""
        chart_container = QWidget()
        chart_container.setFixedHeight(320)
        chart_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(25, 25, 25, 0.95),
                    stop:0.5 rgba(35, 35, 35, 0.90),
                    stop:1 rgba(25, 25, 25, 0.95));
                border-radius: 16px;
                border: 2px solid rgba(255, 107, 53, 0.3);
                backdrop-filter: blur(10px);
            }
        """)
        chart_layout = QVBoxLayout(chart_container)
        chart_layout.setContentsMargins(20, 20, 20, 20)
        chart_layout.setSpacing(15)

        # Professional title with icon
        title_container = QWidget()
        title_container.setStyleSheet("background: transparent;")
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 0)

        # Icon based on chart type
        icon_map = {
            "line": "📈", "bar": "📊", "pie": "🥧", "area": "📉"
        }
        icon = icon_map.get(chart_type, "📊")

        icon_label = QLabel(icon)
        icon_label.setStyleSheet("""
            font-size: 24px;
            background: transparent;
            margin-right: 10px;
        """)

        title_label = QLabel(title.replace("📈 ", "").replace("💰 ", "").replace("📊 ", "").replace("🏋️ ", ""))
        title_label.setStyleSheet("""
            font-size: 18px;
            font-weight: 700;
            color: white;
            background: transparent;
            letter-spacing: 0.5px;
        """)

        title_layout.addWidget(icon_label)
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        chart_layout.addWidget(title_container)

        # Chart area
        chart_area = QWidget()
        chart_area.setStyleSheet("""
            background: rgba(15, 15, 15, 0.8);
            border-radius: 12px;
            border: 1px solid rgba(255, 107, 53, 0.2);
        """)
        chart_area_layout = QVBoxLayout(chart_area)
        chart_area_layout.setContentsMargins(15, 15, 15, 15)

        if labels and data and len(labels) > 0:
            if chart_type in ["bar", "line"]:
                self.create_bar_chart(chart_area_layout, labels, data, colors)
            elif chart_type == "pie":
                self.create_pie_chart(chart_area_layout, labels, data, colors)
            else:
                self.create_bar_chart(chart_area_layout, labels, data, colors)
        else:
            # No data available
            no_data_container = QWidget()
            no_data_container.setStyleSheet("background: transparent;")
            no_data_layout = QVBoxLayout(no_data_container)

            no_data_icon = QLabel("📊")
            no_data_icon.setStyleSheet("""
                font-size: 48px;
                color: rgba(255, 107, 53, 0.3);
                background: transparent;
            """)
            no_data_icon.setAlignment(Qt.AlignCenter)

            no_data_text = QLabel("Aucune donnée disponible")
            no_data_text.setStyleSheet("""
                color: rgba(255, 255, 255, 0.5);
                font-size: 16px;
                font-weight: 500;
                background: transparent;
            """)
            no_data_text.setAlignment(Qt.AlignCenter)

            no_data_layout.addWidget(no_data_icon)
            no_data_layout.addWidget(no_data_text)
            chart_area_layout.addWidget(no_data_container)

        chart_layout.addWidget(chart_area)
        return chart_container

    def create_bar_chart(self, layout, labels, data, color):
        """Create a visual bar chart using Qt widgets"""
        if not data:
            return

        max_value = max(data) if data else 1

        # Create scrollable area for many bars
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:horizontal {
                background: rgba(255, 107, 53, 0.1);
                height: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:horizontal {
                background: rgba(255, 107, 53, 0.5);
                border-radius: 4px;
                min-width: 20px;
            }
        """)

        bars_widget = QWidget()
        bars_layout = QHBoxLayout(bars_widget)
        bars_layout.setSpacing(8)
        bars_layout.setContentsMargins(10, 10, 10, 10)

        # Show max 12 bars to keep it readable
        display_labels = labels[:12]
        display_data = data[:12]

        for i, (label, value) in enumerate(zip(display_labels, display_data)):
            bar_container = QWidget()
            bar_container.setFixedWidth(60)
            bar_container.setStyleSheet("background: transparent;")
            bar_layout = QVBoxLayout(bar_container)
            bar_layout.setContentsMargins(0, 0, 0, 0)
            bar_layout.setSpacing(5)

            # Value label on top
            value_label = QLabel(str(value))
            value_label.setStyleSheet("""
                color: white;
                font-size: 11px;
                font-weight: 600;
                background: transparent;
            """)
            value_label.setAlignment(Qt.AlignCenter)
            bar_layout.addWidget(value_label)

            # Bar visualization
            bar_height = int((value / max_value) * 120) if max_value > 0 else 0
            bar_widget = QWidget()
            bar_widget.setFixedSize(40, 120)
            bar_widget.setStyleSheet("background: transparent;")

            # Create the actual bar
            bar = QWidget(bar_widget)
            bar.setFixedSize(40, max(bar_height, 5))  # Minimum height of 5
            bar.move(0, 120 - bar_height)

            # Color gradient for bars
            if isinstance(color, list):
                bar_color = color[i % len(color)]
            else:
                bar_color = color

            bar.setStyleSheet(f"""
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {bar_color},
                    stop:1 {bar_color}80);
                border-radius: 6px;
                border: 1px solid {bar_color};
            """)

            bar_layout.addWidget(bar_widget)

            # Label at bottom
            label_text = str(label)[:8] + "..." if len(str(label)) > 8 else str(label)
            label_widget = QLabel(label_text)
            label_widget.setStyleSheet("""
                color: rgba(255, 255, 255, 0.8);
                font-size: 10px;
                font-weight: 500;
                background: transparent;
            """)
            label_widget.setAlignment(Qt.AlignCenter)
            bar_layout.addWidget(label_widget)

            bars_layout.addWidget(bar_container)

        if len(labels) > 12:
            more_label = QLabel(f"+{len(labels) - 12}")
            more_label.setStyleSheet("""
                color: rgba(255, 107, 53, 0.7);
                font-size: 12px;
                font-weight: 600;
                background: transparent;
                padding: 10px;
            """)
            more_label.setAlignment(Qt.AlignCenter)
            bars_layout.addWidget(more_label)

        bars_layout.addStretch()
        scroll.setWidget(bars_widget)
        layout.addWidget(scroll)

    def create_pie_chart(self, layout, labels, data, colors):
        """Create a visual pie chart representation"""
        if not data:
            return

        total = sum(data)
        if total == 0:
            return

        pie_container = QWidget()
        pie_container.setFixedHeight(180)
        pie_layout = QHBoxLayout(pie_container)
        pie_layout.setContentsMargins(0, 0, 0, 0)

        # Visual pie representation (simplified as progress bars)
        chart_area = QWidget()
        chart_area.setFixedSize(160, 160)
        chart_layout = QVBoxLayout(chart_area)
        chart_layout.setContentsMargins(10, 10, 10, 10)

        # Create circular segments as horizontal bars
        for i, (label, value) in enumerate(zip(labels[:5], data[:5])):  # Max 5 segments
            percentage = (value / total) * 100

            segment_container = QWidget()
            segment_layout = QHBoxLayout(segment_container)
            segment_layout.setContentsMargins(0, 2, 0, 2)

            # Color indicator
            color_indicator = QWidget()
            color_indicator.setFixedSize(20, 20)
            if isinstance(colors, list):
                segment_color = colors[i % len(colors)]
            else:
                segment_color = colors
            color_indicator.setStyleSheet(f"""
                background: {segment_color};
                border-radius: 10px;
                border: 2px solid white;
            """)

            # Progress bar
            progress_bar = QWidget()
            progress_bar.setFixedHeight(20)
            bar_width = int((percentage / 100) * 120)
            progress_bar.setFixedWidth(max(bar_width, 10))
            progress_bar.setStyleSheet(f"""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {segment_color},
                    stop:1 {segment_color}60);
                border-radius: 10px;
                border: 1px solid {segment_color};
            """)

            segment_layout.addWidget(color_indicator)
            segment_layout.addWidget(progress_bar)
            segment_layout.addStretch()
            chart_layout.addWidget(segment_container)

        # Legend
        legend_area = QWidget()
        legend_layout = QVBoxLayout(legend_area)
        legend_layout.setContentsMargins(20, 10, 10, 10)

        for i, (label, value) in enumerate(zip(labels[:5], data[:5])):
            percentage = (value / total) * 100

            legend_item = QWidget()
            legend_item_layout = QHBoxLayout(legend_item)
            legend_item_layout.setContentsMargins(0, 2, 0, 2)

            # Color dot
            dot = QWidget()
            dot.setFixedSize(12, 12)
            if isinstance(colors, list):
                dot_color = colors[i % len(colors)]
            else:
                dot_color = colors
            dot.setStyleSheet(f"""
                background: {dot_color};
                border-radius: 6px;
            """)

            # Label and percentage
            text = QLabel(f"{label}: {percentage:.1f}%")
            text.setStyleSheet("""
                color: white;
                font-size: 11px;
                font-weight: 500;
                background: transparent;
            """)

            legend_item_layout.addWidget(dot)
            legend_item_layout.addWidget(text)
            legend_item_layout.addStretch()
            legend_layout.addWidget(legend_item)

        legend_layout.addStretch()

        pie_layout.addWidget(chart_area)
        pie_layout.addWidget(legend_area)
        layout.addWidget(pie_container)

    def create_smooth_circle_chart(self, title, labels, data, colors):
        """Create a proper circular pie chart with percentages"""
        chart_container = QWidget()
        chart_container.setFixedSize(450, 500)
        chart_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(15, 15, 15, 0.95),
                    stop:0.5 rgba(25, 25, 25, 0.90),
                    stop:1 rgba(15, 15, 15, 0.95));
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.4);
                backdrop-filter: blur(15px);
            }
        """)
        chart_layout = QVBoxLayout(chart_container)
        chart_layout.setContentsMargins(25, 25, 25, 25)
        chart_layout.setSpacing(20)

        # Professional title
        title_label = QLabel(title.replace("📊 ", "").replace("👥 ", ""))
        title_label.setStyleSheet("""
            font-size: 18px;
            font-weight: 700;
            color: white;
            background: transparent;
            letter-spacing: 0.8px;
            margin-bottom: 10px;
        """)
        title_label.setAlignment(Qt.AlignCenter)
        chart_layout.addWidget(title_label)

        if not labels or not data or sum(data) == 0:
            # No data state
            no_data_container = QWidget()
            no_data_container.setStyleSheet("background: transparent;")
            no_data_layout = QVBoxLayout(no_data_container)

            no_data_circle = QWidget()
            no_data_circle.setFixedSize(250, 250)
            no_data_circle.setStyleSheet("""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(255, 107, 53, 0.1),
                    stop:1 rgba(255, 107, 53, 0.05));
                border: 3px dashed rgba(255, 107, 53, 0.3);
                border-radius: 125px;
            """)

            no_data_text = QLabel("Aucune donnée\ndisponible")
            no_data_text.setStyleSheet("""
                color: rgba(255, 255, 255, 0.5);
                font-size: 16px;
                font-weight: 500;
                background: transparent;
                text-align: center;
            """)
            no_data_text.setAlignment(Qt.AlignCenter)
            no_data_text.setParent(no_data_circle)
            no_data_text.move(75, 110)

            no_data_layout.addWidget(no_data_circle, 0, Qt.AlignCenter)
            chart_layout.addWidget(no_data_container)
            chart_layout.addStretch()
            return chart_container

        # Calculate percentages
        total = sum(data)
        percentages = [(value / total) * 100 for value in data]

        # Create the actual circular pie chart
        pie_widget = CircularPieChart(labels, data, colors, percentages)
        pie_widget.setFixedSize(280, 280)
        chart_layout.addWidget(pie_widget, 0, Qt.AlignCenter)

        # Legend with percentages
        legend_container = QWidget()
        legend_container.setStyleSheet("background: transparent;")
        legend_layout = QVBoxLayout(legend_container)
        legend_layout.setContentsMargins(0, 0, 0, 0)
        legend_layout.setSpacing(10)

        for i, (label, percentage, value) in enumerate(zip(labels, percentages, data)):
            legend_item = QWidget()
            legend_item.setStyleSheet("""
                background: rgba(35, 35, 35, 0.7);
                border-radius: 10px;
                border: 1px solid rgba(255, 255, 255, 0.1);
                padding: 8px;
            """)
            legend_item_layout = QHBoxLayout(legend_item)
            legend_item_layout.setContentsMargins(15, 10, 15, 10)

            # Color indicator
            color_dot = QWidget()
            color_dot.setFixedSize(20, 20)
            segment_color = colors[i % len(colors)] if colors else "#ff6b35"
            color_dot.setStyleSheet(f"""
                background: {segment_color};
                border-radius: 10px;
                border: 2px solid rgba(255, 255, 255, 0.3);
            """)

            # Label
            label_text = QLabel(f"{label}")
            label_text.setStyleSheet("""
                color: white;
                font-size: 14px;
                font-weight: 600;
                background: transparent;
            """)

            # Percentage (large and prominent)
            percentage_text = QLabel(f"{percentage:.1f}%")
            percentage_text.setStyleSheet(f"""
                color: {segment_color};
                font-size: 18px;
                font-weight: 800;
                background: transparent;
                letter-spacing: 0.5px;
            """)

            # Count
            count_text = QLabel(f"({value})")
            count_text.setStyleSheet("""
                color: rgba(255, 255, 255, 0.7);
                font-size: 12px;
                font-weight: 500;
                background: transparent;
            """)

            legend_item_layout.addWidget(color_dot)
            legend_item_layout.addWidget(label_text)
            legend_item_layout.addStretch()
            legend_item_layout.addWidget(count_text)
            legend_item_layout.addWidget(percentage_text)

            legend_layout.addWidget(legend_item)

        chart_layout.addWidget(legend_container)
        chart_layout.addStretch()

        return chart_container


class CircularPieChart(QWidget):
    """Custom widget to draw a proper circular pie chart with percentages"""

    def __init__(self, labels, data, colors, percentages):
        super().__init__()
        self.labels = labels
        self.data = data
        self.colors = colors
        self.percentages = percentages
        self.setStyleSheet("background: transparent;")

    def paintEvent(self, event):
        from PySide6.QtGui import QPainter, QPen, QBrush, QFont, QColor
        from PySide6.QtCore import QRect
        import math

        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Chart dimensions
        rect = QRect(20, 20, 240, 240)
        center_x, center_y = rect.center().x(), rect.center().y()
        radius = 120

        if not self.data or sum(self.data) == 0:
            # Draw empty circle
            painter.setPen(QPen(QColor(255, 107, 53, 80), 3))
            painter.setBrush(QBrush(QColor(45, 45, 45, 100)))
            painter.drawEllipse(rect)
            return

        # Draw pie segments
        start_angle = 0
        total = sum(self.data)

        for i, (value, percentage) in enumerate(zip(self.data, self.percentages)):
            if value == 0:
                continue

            # Calculate angle for this segment
            angle = int((value / total) * 360 * 16)  # Qt uses 16ths of degrees

            # Get color
            color = QColor(self.colors[i % len(self.colors)] if self.colors else "#ff6b35")

            # Draw segment
            painter.setPen(QPen(QColor(255, 255, 255, 50), 2))
            painter.setBrush(QBrush(color))
            painter.drawPie(rect, start_angle, angle)

            # Draw percentage text on segment
            if percentage > 5:  # Only show text for segments > 5%
                # Calculate text position
                mid_angle = start_angle + angle / 2
                text_angle = mid_angle / 16.0 * math.pi / 180.0  # Convert to radians
                text_radius = radius * 0.7
                text_x = center_x + text_radius * math.cos(text_angle)
                text_y = center_y + text_radius * math.sin(text_angle)

                # Draw percentage text
                painter.setPen(QPen(QColor(255, 255, 255), 1))
                painter.setFont(QFont("Arial", 12, QFont.Bold))
                text_rect = QRect(int(text_x - 25), int(text_y - 10), 50, 20)
                painter.drawText(text_rect, Qt.AlignCenter, f"{percentage:.1f}%")

            start_angle += angle

        # Draw center circle
        center_rect = QRect(center_x - 40, center_y - 40, 80, 80)
        painter.setPen(QPen(QColor(255, 107, 53, 100), 2))
        painter.setBrush(QBrush(QColor(25, 25, 25, 200)))
        painter.drawEllipse(center_rect)

        # Draw total in center
        painter.setPen(QPen(QColor(255, 107, 53), 1))
        painter.setFont(QFont("Arial", 16, QFont.Bold))
        painter.drawText(center_rect, Qt.AlignCenter, str(sum(self.data)))


class LoginWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        
    def setup_ui(self):
        self.setWindowTitle("UaFit Login")
        self.setFixedSize(800, 400)

        self.stats_timer = QTimer(self)
        self.stats_timer.timeout.connect(self.update_stats)
        self.stats_timer.start(30000)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Left section
        left_widget = QWidget()
        left_widget.setStyleSheet("background-color: #002347;")
        left_layout = QVBoxLayout(left_widget)
        left_layout.setAlignment(Qt.AlignCenter)
        
        # Logo
        ua_fit_widget = QWidget()
        ua_fit_layout = QHBoxLayout(ua_fit_widget)
        ua_fit_layout.setAlignment(Qt.AlignCenter)
        
        ua_label = QLabel("Ua")
        ua_label.setStyleSheet("font-size: 48px; color: #FFC107; font-weight: bold;")
        fit_label = QLabel("Fit")
        fit_label.setStyleSheet("font-size: 48px; color: white; font-weight: bold;")
        ua_fit_layout.addWidget(ua_label)
        ua_fit_layout.addWidget(fit_label)
        
        slogan_label = QLabel("Votre porte d'entrée vers une gestion sportive simplifiée")
        slogan_label.setStyleSheet("font-size: 14px; color: white; margin-bottom: 40px; margin-top: -5px;")
        
        left_layout.addStretch()
        left_layout.addWidget(ua_fit_widget)
        left_layout.addWidget(slogan_label)
        left_layout.addStretch()
        left_widget.setFixedWidth(400)
        
        # Right section
        right_widget = QWidget()
        right_widget.setStyleSheet("background-color: white;")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(40, 40, 40, 40)
        right_layout.setSpacing(20)
        
        # Title
        title = QLabel("Connexion")
        title.setStyleSheet("""
            font-size: 24px;
            color: #002347;
            font-weight: bold;
            margin-bottom: 20px;
        """)
        right_layout.addWidget(title)
        
        # Email input
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Email")
        self.email_input.setStyleSheet("""
            QLineEdit {
                background-color: #FFC107;
                border: 2px solid #FFC107;
                border-radius: 15px;
                padding: 10px;
                color: #333;
            }
            QLineEdit:focus {
                border-color: #FFC107;
            }
        """)
        self.email_input.setFixedHeight(40)
        right_layout.addWidget(self.email_input)
        
        # Password input
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Mot de passe")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setStyleSheet(self.email_input.styleSheet())
        self.password_input.setFixedHeight(40)
        right_layout.addWidget(self.password_input)
        
        # Login button
        login_btn = QPushButton("Se connecter")
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #002347;
                color: white;
                border-radius: 15px;
                padding: 10px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #003366;
            }
        """)
        login_btn.setFixedHeight(40)
        login_btn.clicked.connect(self.login)
        right_layout.addWidget(login_btn)
        
        # Signup link
        signup_btn = QPushButton("Pas encore de compte ? Inscrivez-vous")
        signup_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #002347;
                font-size: 13px;
                text-decoration: underline;
            }
            QPushButton:hover {
                color: #003366;
            }
        """)
        signup_btn.setCursor(Qt.PointingHandCursor)
        signup_btn.clicked.connect(self.show_signup)
        right_layout.addWidget(signup_btn)
        
        right_layout.addStretch()
        
        main_layout.addWidget(left_widget)
        main_layout.addWidget(right_widget)

    def update_stats(self):
        # Cette méthode sera appelée par le timer pour mettre à jour les stats
        if hasattr(self, 'stats_container'):
            # Supprimer l'ancien widget de stats
            self.header_layout.removeWidget(self.stats_container)
            self.stats_container.deleteLater()
            
            # Recréer le widget avec les nouvelles données
            self.create_stats_widget(self.header_layout)
    
    def login(self):
        email = self.email_input.text().strip()
        password = self.password_input.text()
        
        if not email or not password:
            QMessageBox.warning(self, "Erreur", "Veuillez remplir tous les champs")
            return
            
        user = DatabaseManager.execute_query(
            "SELECT id, first_name, last_name, role, address FROM users WHERE email = %s AND password = %s",
            (email, password),
            fetch_one=True
        )
        
        if user:
            self.close()
            dashboard = DashboardWindow(
                user['id'],
                f"{user['first_name']} {user['last_name']}",
                user['role'],
                user['address']
            )
            dashboard.show()
        else:
            QMessageBox.warning(self, "Erreur", "Email ou mot de passe incorrect")
    
    def show_signup(self):
        self.signup_window = SignupWindow()
        self.signup_window.show()
        self.close()
class PlanningDialog(BaseDialog):
    def __init__(self, gym_id, parent=None):
        self.gym_id = gym_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("Gestion du Planning")
        self.setFixedSize(600, 500)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Default hours
        default_group = QGroupBox("Heures par défaut")
        default_layout = QHBoxLayout(default_group)
        
        self.default_start = QTimeEdit()
        self.default_start.setDisplayFormat("HH:mm")
        self.default_end = QTimeEdit()
        self.default_end.setDisplayFormat("HH:mm")
        
        apply_btn = QPushButton("Appliquer à tous")
        apply_btn.clicked.connect(self.apply_default_hours)
        
        default_layout.addWidget(QLabel("De:"))
        default_layout.addWidget(self.default_start)
        default_layout.addWidget(QLabel("à:"))
        default_layout.addWidget(self.default_end)
        default_layout.addWidget(apply_btn)
        
        layout.addWidget(default_group)
        
        # Days grid
        self.day_widgets = {}
        days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        
        grid = QGridLayout()
        grid.setSpacing(15)
        
        for i, day in enumerate(days):
            # Checkbox
            chk = QCheckBox(day)
            self.day_widgets[day] = {'checkbox': chk}
            grid.addWidget(chk, i, 0)
            
            # Time inputs
            start_time = QTimeEdit()
            start_time.setDisplayFormat("HH:mm")
            end_time = QTimeEdit()
            end_time.setDisplayFormat("HH:mm")
            
            self.day_widgets[day]['start'] = start_time
            self.day_widgets[day]['end'] = end_time
            
            grid.addWidget(QLabel("De:"), i, 1)
            grid.addWidget(start_time, i, 2)
            grid.addWidget(QLabel("à:"), i, 3)
            grid.addWidget(end_time, i, 4)
        
        layout.addLayout(grid)
        
        # Buttons
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Enregistrer")
        save_btn.clicked.connect(self.save_planning)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
        
        # Load existing planning
        self.load_planning()
    
    def load_planning(self):
        planning = DatabaseManager.get_planning(self.gym_id)
        if not planning:
            return
            
        for day, data in planning.items():
            if day in self.day_widgets:
                self.day_widgets[day]['checkbox'].setChecked(data['is_open'])
                if data['opening_time']:
                    time = QTime.fromString(data['opening_time'], "HH:mm:ss")
                    self.day_widgets[day]['start'].setTime(time)
                if data['closing_time']:
                    time = QTime.fromString(data['closing_time'], "HH:mm:ss")
                    self.day_widgets[day]['end'].setTime(time)
    
    def apply_default_hours(self):
        start = self.default_start.time()
        end = self.default_end.time()
        
        for day in self.day_widgets.values():
            day['start'].setTime(start)
            day['end'].setTime(end)
    
    def save_planning(self):
        planning_data = {}
        for day, widgets in self.day_widgets.items():
            planning_data[day] = {
                'is_open': widgets['checkbox'].isChecked(),
                'opening_time': widgets['start'].time().toString("HH:mm:ss"),
                'closing_time': widgets['end'].time().toString("HH:mm:ss")
            }
        
        if DatabaseManager.save_planning(self.gym_id, planning_data):
            QMessageBox.information(self, "Succès", "Planning enregistré avec succès!")
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de l'enregistrement du planning")

class WorkoutTrackingDialog(BaseDialog):
    def __init__(self, member_id, gym_id, parent=None):
        self.member_id = member_id
        self.gym_id = gym_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("Enregistrer mon entraînement")
        self.setFixedSize(500, 400)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Date
        date_label = QLabel(f"Date: {QDate.currentDate().toString('dddd d MMMM yyyy')}")
        layout.addWidget(date_label)
        
        # Description
        desc_label = QLabel("Description (optionnel):")
        layout.addWidget(desc_label)
        
        self.desc_input = QTextEdit()
        self.desc_input.setPlaceholderText("Ex: Séance pectoraux 3x10...")
        layout.addWidget(self.desc_input)
        
        # Photo
        photo_label = QLabel("Photo (optionnel):")
        layout.addWidget(photo_label)
        
        self.photo_path = QLineEdit()
        self.photo_path.setReadOnly(True)
        
        browse_btn = QPushButton("Parcourir...")
        browse_btn.clicked.connect(self.browse_photo)
        
        photo_layout = QHBoxLayout()
        photo_layout.addWidget(self.photo_path)
        photo_layout.addWidget(browse_btn)
        layout.addLayout(photo_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Enregistrer")
        save_btn.clicked.connect(self.save_workout)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
    
    def browse_photo(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner une photo", "", "Images (*.png *.jpg *.jpeg)"
        )
        if file_path:
            self.photo_path.setText(file_path)
    
    def save_workout(self):
        description = self.desc_input.toPlainText() or None
        photo_path = self.photo_path.text() or None
        
        print(f"Saving workout for member_id: {self.member_id}, gym_id: {self.gym_id}, date: {QDate.currentDate().toString('yyyy-MM-dd')}")
        
        if DatabaseManager.save_workout(
            self.member_id,
            self.gym_id,
            QDate.currentDate().toString("yyyy-MM-dd"),
            description,
            photo_path
        ):
            QMessageBox.information(self, "Succès", "Entraînement enregistré!")
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de l'enregistrement")

class WorkoutHistoryDialog(BaseDialog):
    def __init__(self, member_id, gym_id, parent=None):
        self.member_id = member_id
        self.gym_id = gym_id
        self.current_week = QDate.currentDate()
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle(f"💪 Historique d'entraînement")
        self.setFixedSize(900, 650)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)

        # Modern header
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 140, 0, 0.2),
                    stop:1 rgba(255, 107, 53, 0.2));
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(20, 15, 20, 15)

        # Gym info
        gym_info = DatabaseManager.execute_query(
            "SELECT name FROM gyms WHERE id = %s",
            (self.gym_id,),
            fetch_one=True
        )
        gym_name = gym_info['name'] if gym_info else "Inconnu"

        gym_label = QLabel(f"🏋️ Salle: {gym_name}")
        gym_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: 700;
                color: white;
                background: transparent;
                border: none;
            }
        """)
        header_layout.addWidget(gym_label)
        layout.addWidget(header_container)
        
        # Modern week navigation
        nav_container = QWidget()
        nav_container.setStyleSheet("""
            QWidget {
                background: rgba(45, 45, 45, 0.8);
                border-radius: 10px;
                border: 1px solid rgba(255, 140, 0, 0.2);
            }
        """)
        nav_layout = QHBoxLayout(nav_container)
        nav_layout.setContentsMargins(15, 10, 15, 10)

        self.prev_week_btn = QPushButton("◀ Précédente")
        self.prev_week_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff8c00, stop:1 #ff6b35);
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 8px 15px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff7700, stop:1 #ff5722);
            }
        """)
        self.prev_week_btn.clicked.connect(self.prev_week)

        self.week_label = QLabel()
        self.week_label.setAlignment(Qt.AlignCenter)
        self.week_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: 600;
                color: white;
                background: transparent;
                border: none;
            }
        """)

        self.next_week_btn = QPushButton("Suivante ▶")
        self.next_week_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff8c00, stop:1 #ff6b35);
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 8px 15px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff7700, stop:1 #ff5722);
            }
        """)
        self.next_week_btn.clicked.connect(self.next_week)

        nav_layout.addWidget(self.prev_week_btn)
        nav_layout.addStretch()
        nav_layout.addWidget(self.week_label)
        nav_layout.addStretch()
        nav_layout.addWidget(self.next_week_btn)
        layout.addWidget(nav_container)
        
        # Modern calendar grid
        calendar_container = QWidget()
        calendar_container.setStyleSheet("""
            QWidget {
                background: rgba(45, 45, 45, 0.6);
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.2);
            }
        """)
        calendar_layout = QVBoxLayout(calendar_container)
        calendar_layout.setContentsMargins(20, 15, 20, 15)

        self.grid = QGridLayout()
        self.grid.setSpacing(12)

        # Modern day headers
        days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        for i, day in enumerate(days):
            day_label = QLabel(day)
            day_label.setStyleSheet("""
                QLabel {
                    font-size: 12px;
                    font-weight: 600;
                    color: rgba(255, 140, 0, 0.9);
                    background: transparent;
                    border: none;
                    padding: 5px;
                }
            """)
            day_label.setAlignment(Qt.AlignCenter)
            self.grid.addWidget(day_label, 0, i, Qt.AlignCenter)

        # Modern day buttons
        self.day_buttons = []
        for i in range(7):
            btn = QPushButton()
            btn.setFixedSize(90, 90)
            btn.setStyleSheet("""
                QPushButton {
                    background: rgba(35, 35, 35, 0.8);
                    border: 2px solid rgba(255, 140, 0, 0.3);
                    border-radius: 12px;
                    color: white;
                    font-size: 11px;
                    font-weight: 500;
                    transition: all 0.3s ease-in-out;
                }
                QPushButton:hover {
                    background: rgba(255, 140, 0, 0.2);
                    border: 2px solid rgba(255, 140, 0, 0.6);
                    transform: scale(1.05);
                }
            """)
            btn.clicked.connect(lambda _, idx=i: self.show_workout_details(idx))
            self.day_buttons.append(btn)
            self.grid.addWidget(btn, 1, i, Qt.AlignCenter)

        calendar_layout.addLayout(self.grid)
        layout.addWidget(calendar_container)
        
        # Modern details section
        details_container = QWidget()
        details_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(45, 45, 45, 0.9),
                    stop:1 rgba(35, 35, 35, 0.9));
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        details_main_layout = QVBoxLayout(details_container)
        details_main_layout.setContentsMargins(20, 15, 20, 15)

        # Details header
        details_header = QLabel("📋 Détails de l'entraînement")
        details_header.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 700;
                color: rgba(255, 140, 0, 0.9);
                background: transparent;
                border: none;
                margin-bottom: 10px;
            }
        """)
        details_main_layout.addWidget(details_header)

        # Scrollable content
        details_scroll = QScrollArea()
        details_scroll.setWidgetResizable(True)
        details_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 140, 0, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 140, 0, 0.5);
                border-radius: 4px;
                min-height: 20px;
            }
        """)

        details_content = QWidget()
        details_content.setStyleSheet("background: transparent;")
        details_layout = QVBoxLayout(details_content)
        details_layout.setSpacing(15)

        # Text details
        self.details_label = QLabel()
        self.details_label.setWordWrap(True)
        self.details_label.setAlignment(Qt.AlignTop)
        self.details_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: white;
                background: rgba(35, 35, 35, 0.5);
                border-radius: 8px;
                padding: 15px;
                border: 1px solid rgba(255, 140, 0, 0.2);
                line-height: 1.4;
            }
        """)

        # Photo section
        self.photo_label = QLabel()
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setMinimumHeight(200)
        self.photo_label.setStyleSheet("""
            QLabel {
                background: rgba(35, 35, 35, 0.5);
                border-radius: 8px;
                border: 1px solid rgba(255, 140, 0, 0.2);
                color: rgba(255, 255, 255, 0.6);
            }
        """)

        details_layout.addWidget(self.details_label)
        details_layout.addWidget(self.photo_label)
        details_scroll.setWidget(details_content)
        details_main_layout.addWidget(details_scroll)

        layout.addWidget(details_container)
        
        # Update display
        self.update_week_display()

    def update_week_display(self):
        # Calculate start of week (Monday)
        start_date = self.current_week.addDays(1 - self.current_week.dayOfWeek())
        end_date = start_date.addDays(6)
        
        self.week_label.setText(f"Semaine du {start_date.toString('dd/MM')} au {end_date.toString('dd/MM')}")
        
        # Get workouts for this week
        workouts = DatabaseManager.get_workouts(
            self.member_id,
            self.gym_id,
            start_date.toString("yyyy-MM-dd")
        )
        
        workout_dates = {str(w['workout_date']) for w in workouts}  # Convert date to string for comparison
        
        # Update day buttons with modern styling
        for i in range(7):
            date = start_date.addDays(i)
            btn = self.day_buttons[i]

            if date.toString("yyyy-MM-dd") in workout_dates:
                # Day with workout - orange gradient
                btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 #ff8c00, stop:1 #ff6b35);
                        border: 2px solid rgba(255, 140, 0, 0.8);
                        border-radius: 12px;
                        color: white;
                        font-size: 12px;
                        font-weight: 700;
                        transition: all 0.3s ease-in-out;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 #ff7700, stop:1 #ff5722);
                        transform: scale(1.08);
                        box-shadow: 0 4px 15px rgba(255, 140, 0, 0.4);
                    }
                """)
                btn.setText(f"{date.toString('d')}\n💪")
                btn.setProperty("has_workout", True)
            else:
                # Day without workout - dark theme
                btn.setStyleSheet("""
                    QPushButton {
                        background: rgba(35, 35, 35, 0.8);
                        border: 2px solid rgba(255, 140, 0, 0.3);
                        border-radius: 12px;
                        color: rgba(255, 255, 255, 0.7);
                        font-size: 12px;
                        font-weight: 500;
                        transition: all 0.3s ease-in-out;
                    }
                    QPushButton:hover {
                        background: rgba(255, 140, 0, 0.2);
                        border: 2px solid rgba(255, 140, 0, 0.6);
                        color: white;
                        transform: scale(1.05);
                    }
                """)
                btn.setText(date.toString("d"))
                btn.setProperty("has_workout", False)

            btn.setProperty("date", date.toString("yyyy-MM-dd"))
    
    def prev_week(self):
        self.current_week = self.current_week.addDays(-7)
        self.update_week_display()
    
    def next_week(self):
        self.current_week = self.current_week.addDays(7)
        self.update_week_display()
    
    
    
    def show_workout_details(self, day_idx):
        btn = self.day_buttons[day_idx]
        if not btn.property("has_workout"):
            self.details_label.setText("Aucun entraînement enregistré ce jour")
            self.photo_label.clear()
            return
        
        date = btn.property("date")
        workout = DatabaseManager.execute_query("""
            SELECT description, photo FROM workout_tracking
            WHERE member_id = %s AND gym_id = %s AND workout_date = %s
            LIMIT 1
        """, (self.member_id, self.gym_id, date), fetch_one=True)
        
        if workout:
            details = f"Entraînement du {QDate.fromString(date, 'yyyy-MM-dd').toString('dddd d MMMM yyyy')}\n\n"
            details += workout['description'] if workout['description'] else "Pas de description"
            self.details_label.setText(details)
            
            if workout['photo']:
                pixmap = QPixmap()
                pixmap.loadFromData(workout['photo'])
                self.photo_label.setPixmap(pixmap.scaledToHeight(200, Qt.SmoothTransformation))
            else:
                self.photo_label.clear()

class WorkoutHistoryDialog(BaseDialog):
    def __init__(self, member_id, gym_id, parent=None):
        self.member_id = member_id
        self.gym_id = gym_id
        self.current_week = QDate.currentDate()
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("💪 Historique d'entraînement")
        self.setFixedSize(1000, 750)

        # Ultra-modern dark theme styling
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2c2c2c, stop:1 #1a1a1a);
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.3);
            }
            QLabel {
                color: white;
                background: transparent;
            }
            QPushButton {
                background: rgba(255, 255, 255, 0.1);
                border: 1px solid rgba(255, 107, 53, 0.3);
                border-radius: 10px;
                color: white;
                font-weight: 500;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background: rgba(255, 107, 53, 0.2);
                border: 1px solid rgba(255, 107, 53, 0.6);
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # Ultra-modern header section
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 140, 0, 0.2),
                    stop:1 rgba(255, 107, 53, 0.2));
                border-radius: 15px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(25, 20, 25, 20)

        # Main title
        title = QLabel("💪 HISTORIQUE D'ENTRAÎNEMENT")
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: 800;
            color: white;
            text-align: center;
            letter-spacing: 1px;
            margin-bottom: 15px;
        """)
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)

        # Info container
        info_container = QWidget()
        info_container.setStyleSheet("background: transparent;")
        info_layout = QVBoxLayout(info_container)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_layout.setSpacing(8)

        # Gym info with modern styling
        self.gym_label = QLabel()
        self.gym_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 700;
            color: #ff6b35;
            background: rgba(255, 107, 53, 0.1);
            padding: 10px 15px;
            border-radius: 8px;
            border: 1px solid rgba(255, 107, 53, 0.3);
        """)
        info_layout.addWidget(self.gym_label)

        # Member info with modern styling
        self.member_label = QLabel()
        self.member_label.setStyleSheet("""
            font-size: 14px;
            font-weight: 600;
            color: white;
            background: rgba(255, 255, 255, 0.05);
            padding: 8px 15px;
            border-radius: 8px;
            border: 1px solid rgba(255, 107, 53, 0.2);
        """)
        info_layout.addWidget(self.member_label)

        header_layout.addWidget(info_container)
        layout.addWidget(header_container)

        # Ultra-modern week navigation
        nav_container = QWidget()
        nav_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.8);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        nav_layout = QHBoxLayout(nav_container)
        nav_layout.setContentsMargins(20, 15, 20, 15)

        self.prev_week_btn = QPushButton("◀ PRÉCÉDENTE")
        self.prev_week_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a5568, stop:1 #2d3748);
                border: none;
                border-radius: 10px;
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 12px 20px;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a6578, stop:1 #3d4758);
                transform: translateY(-1px);
            }
        """)
        self.prev_week_btn.clicked.connect(self.prev_week)

        self.week_label = QLabel()
        self.week_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 700;
            color: white;
            background: transparent;
            padding: 10px;
        """)
        self.week_label.setAlignment(Qt.AlignCenter)

        self.next_week_btn = QPushButton("SUIVANTE ▶")
        self.next_week_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a5568, stop:1 #2d3748);
                border: none;
                border-radius: 10px;
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 12px 20px;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a6578, stop:1 #3d4758);
                transform: translateY(-1px);
            }
        """)
        self.next_week_btn.clicked.connect(self.next_week)
        
        nav_layout.addWidget(self.prev_week_btn)
        nav_layout.addStretch()
        nav_layout.addWidget(self.week_label)
        nav_layout.addStretch()
        nav_layout.addWidget(self.next_week_btn)
        layout.addWidget(nav_container)

        # Ultra-modern days grid
        days_container = QWidget()
        days_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.6);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        days_main_layout = QVBoxLayout(days_container)
        days_main_layout.setContentsMargins(20, 20, 20, 20)
        days_main_layout.setSpacing(15)

        self.grid = QGridLayout()
        self.grid.setSpacing(15)

        # Modern day headers
        days = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI", "DIMANCHE"]
        for i, day in enumerate(days):
            day_label = QLabel(day)
            day_label.setStyleSheet("""
                font-size: 12px;
                font-weight: 600;
                color: #ff6b35;
                background: transparent;
                padding: 5px;
                letter-spacing: 1px;
            """)
            day_label.setAlignment(Qt.AlignCenter)
            self.grid.addWidget(day_label, 0, i)

        # Modern day buttons
        self.day_buttons = []
        for i in range(7):
            btn = QPushButton()
            btn.setFixedSize(120, 90)
            btn.clicked.connect(lambda _, idx=i: self.show_workout_details(idx))
            self.day_buttons.append(btn)
            self.grid.addWidget(btn, 1, i, Qt.AlignCenter)

        days_main_layout.addLayout(self.grid)
        layout.addWidget(days_container)
        
        # Ultra-modern details section
        details_container = QWidget()
        details_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(45, 45, 45, 0.9),
                    stop:1 rgba(35, 35, 35, 0.9));
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.3);
            }
        """)
        details_main_layout = QVBoxLayout(details_container)
        details_main_layout.setContentsMargins(25, 20, 25, 20)

        # Details header
        details_header = QLabel("📋 Détails de l'entraînement")
        details_header.setStyleSheet("""
            font-size: 18px;
            font-weight: 700;
            color: #ff6b35;
            background: transparent;
            margin-bottom: 15px;
        """)
        details_main_layout.addWidget(details_header)

        # Scrollable content
        details_scroll = QScrollArea()
        details_scroll.setWidgetResizable(True)
        details_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 107, 53, 0.1);
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 107, 53, 0.5);
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(255, 107, 53, 0.7);
            }
        """)

        details_content = QWidget()
        details_content.setStyleSheet("background: transparent;")
        details_layout = QVBoxLayout(details_content)
        details_layout.setSpacing(15)

        # Text details
        self.details_label = QLabel()
        self.details_label.setWordWrap(True)
        self.details_label.setAlignment(Qt.AlignTop)
        self.details_label.setStyleSheet("""
            font-size: 14px;
            color: white;
            background: rgba(35, 35, 35, 0.5);
            border-radius: 8px;
            padding: 15px;
            border: 1px solid rgba(255, 107, 53, 0.2);
            line-height: 1.4;
        """)

        # Photo section
        self.photo_label = QLabel()
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setMinimumHeight(200)
        self.photo_label.setStyleSheet("""
            background: rgba(35, 35, 35, 0.5);
            border-radius: 8px;
            border: 1px solid rgba(255, 107, 53, 0.2);
            color: rgba(255, 255, 255, 0.6);
        """)

        details_layout.addWidget(self.details_label)
        details_layout.addWidget(self.photo_label)
        details_scroll.setWidget(details_content)
        details_main_layout.addWidget(details_scroll)

        layout.addWidget(details_container)

        # Update display
        self.update_info()
        self.update_week_display()
    
    def update_info(self):
        # Update gym info
        if self.gym_id:
            gym_info = DatabaseManager.execute_query(
                "SELECT name FROM gyms WHERE id = %s", 
                (self.gym_id,), 
                fetch_one=True
            )
            gym_name = gym_info['name'] if gym_info else "Inconnu"
            self.gym_label.setText(f"Salle: {gym_name}")
        
        # Update member info
        if self.member_id:
            member_info = DatabaseManager.execute_query(
                "SELECT first_name, last_name FROM users WHERE id = %s", 
                (self.member_id,), 
                fetch_one=True
            )
            if member_info:
                self.member_label.setText(f"Membre: {member_info['first_name']} {member_info['last_name']}")
            else:
                self.member_label.setText("Membre: Inconnu")
    
    def update_week_display(self):
        print(f"🔄 update_week_display called - member_id: {self.member_id}, gym_id: {self.gym_id}")
        if not self.member_id or not self.gym_id:
            print("❌ Missing member_id or gym_id")
            return

        # Calculate start of week (Monday)
        start_date = self.current_week.addDays(1 - self.current_week.dayOfWeek())
        end_date = start_date.addDays(6)

        self.week_label.setText(f"Semaine du {start_date.toString('dd/MM')} au {end_date.toString('dd/MM')}")
        print(f"📅 Week range: {start_date.toString('dd/MM')} to {end_date.toString('dd/MM')}")

        # Get workouts for this week
        workouts = DatabaseManager.get_workouts(
            self.member_id,
            self.gym_id,
            start_date.toString("yyyy-MM-dd")
        )

        print(f"💪 Found {len(workouts) if workouts else 0} workouts for this week")
        if workouts:
            for workout in workouts:
                print(f"  - {workout['workout_date']}: {workout.get('description', 'No description')}")

        workout_dates = {str(w['workout_date']) for w in workouts} if workouts else set()
        
        # Update day buttons
        for i in range(7):
            date = start_date.addDays(i)
            btn = self.day_buttons[i]
            
            if date.toString("yyyy-MM-dd") in workout_dates:
                btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 #ff6b35, stop:1 #ff8c00);
                        border: 2px solid rgba(255, 107, 53, 0.8);
                        border-radius: 12px;
                        color: white;
                        font-size: 14px;
                        font-weight: 700;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 #ff5722, stop:1 #ff7700);
                        transform: scale(1.05);
                    }
                """)
                btn.setText(f"{date.toString('d')}\n💪")
                btn.setProperty("has_workout", True)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: rgba(35, 35, 35, 0.8);
                        border: 2px solid rgba(255, 107, 53, 0.3);
                        border-radius: 12px;
                        color: white;
                        font-size: 14px;
                        font-weight: 500;
                    }
                    QPushButton:hover {
                        background: rgba(255, 107, 53, 0.2);
                        border: 2px solid rgba(255, 107, 53, 0.6);
                    }
                """)
                btn.setText(date.toString("d"))
                btn.setProperty("has_workout", False)
            
            btn.setProperty("date", date.toString("yyyy-MM-dd"))
    
    def prev_week(self):
        self.current_week = self.current_week.addDays(-7)
        self.update_week_display()
    
    def next_week(self):
        self.current_week = self.current_week.addDays(7)
        self.update_week_display()
    
    def show_workout_details(self, day_idx):
        btn = self.day_buttons[day_idx]
        if not btn.property("has_workout"):
            self.details_label.setText("Aucun entraînement enregistré ce jour")
            self.photo_label.clear()
            return
        
        date = btn.property("date")
        workout = DatabaseManager.execute_query("""
            SELECT description, photo FROM workout_tracking
            WHERE member_id = %s AND gym_id = %s AND workout_date = %s
            LIMIT 1
        """, (self.member_id, self.gym_id, date), fetch_one=True)
        
        if workout:
            details = f"Entraînement du {QDate.fromString(date, 'yyyy-MM-dd').toString('dddd d MMMM yyyy')}\n\n"
            details += workout['description'] if workout['description'] else "Pas de description"
            self.details_label.setText(details)
            
            if workout['photo']:
                pixmap = QPixmap()
                pixmap.loadFromData(workout['photo'])
                self.photo_label.setPixmap(pixmap.scaledToHeight(200, Qt.SmoothTransformation))
            else:
                self.photo_label.clear()

class AddReviewDialog(BaseDialog):
    def __init__(self, gym_id, member_id, parent=None):
        self.gym_id = gym_id
        self.member_id = member_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("Ajouter une évaluation")
        self.setFixedSize(450, 400)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Modern title
        title_label = QLabel("⭐ Évaluez cette salle")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: 700;
                color: white;
                background: transparent;
                border: none;
                letter-spacing: 1px;
                margin-bottom: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Modern rating section
        rating_label = QLabel("🌟 Note (1-5 étoiles):")
        rating_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 600;
                color: white;
                background: transparent;
                border: none;
                margin-bottom: 5px;
            }
        """)
        layout.addWidget(rating_label)

        self.rating_combo = QComboBox()
        self.rating_combo.addItems(["⭐ 1 étoile", "⭐⭐ 2 étoiles", "⭐⭐⭐ 3 étoiles", "⭐⭐⭐⭐ 4 étoiles", "⭐⭐⭐⭐⭐ 5 étoiles"])
        self.rating_combo.setCurrentIndex(4)  # Default to 5 stars
        self.rating_combo.setMinimumHeight(45)
        layout.addWidget(self.rating_combo)

        # Modern comment section
        comment_label = QLabel("💬 Commentaire (optionnel):")
        comment_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 600;
                color: white;
                background: transparent;
                border: none;
                margin-bottom: 5px;
            }
        """)
        layout.addWidget(comment_label)

        self.comment_input = QTextEdit()
        self.comment_input.setPlaceholderText("Décrivez votre expérience dans cette salle...")
        self.comment_input.setMinimumHeight(100)
        layout.addWidget(self.comment_input)

        # Modern buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)

        cancel_btn = QPushButton("❌ Annuler")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255, 255, 255, 0.1);
                color: rgba(255, 255, 255, 0.8);
                font-size: 14px;
                font-weight: 600;
                padding: 12px 20px;
                border-radius: 10px;
                border: 1px solid rgba(255, 255, 255, 0.3);
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: rgba(255, 255, 255, 0.2);
                color: white;
            }
        """)
        cancel_btn.clicked.connect(self.reject)

        submit_btn = QPushButton("✨ Envoyer l'évaluation")
        submit_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #f093fb, stop:1 #f5576c);
                color: white;
                font-size: 14px;
                font-weight: 700;
                padding: 12px 25px;
                border-radius: 10px;
                border: none;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e91e63, stop:1 #f44336);
            }
        """)
        submit_btn.clicked.connect(self.submit_review)

        btn_layout.addWidget(cancel_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(submit_btn)
        layout.addLayout(btn_layout)
    
    def submit_review(self):
        # Extract rating number from the selected text (e.g., "⭐⭐⭐⭐⭐ 5 étoiles" -> 5)
        rating = self.rating_combo.currentIndex() + 1  # Since index starts at 0, add 1
        comment = self.comment_input.toPlainText() or None

        if DatabaseManager.add_gym_review(self.gym_id, self.member_id, rating, comment):
            QMessageBox.information(self, "Succès", "Merci pour votre évaluation!")
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Impossible d'ajouter l'évaluation")

class ReviewsDialog(BaseDialog):
    def __init__(self, gym_id, parent=None):
        self.gym_id = gym_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("⭐ Avis sur la salle")
        self.setFixedSize(800, 650)

        # Modern dark theme styling
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2c2c2c, stop:1 #1a1a1a);
                border-radius: 20px;
                border: 2px solid rgba(255, 107, 53, 0.3);
            }
            QLabel {
                color: white;
                background: transparent;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # Modern header with gym info
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 140, 0, 0.2),
                    stop:1 rgba(255, 107, 53, 0.2));
                border-radius: 15px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        header_layout = QVBoxLayout(header_container)
        header_layout.setContentsMargins(25, 20, 25, 20)

        # Get gym name
        gym_info = DatabaseManager.execute_query(
            "SELECT name FROM gyms WHERE id = %s",
            (self.gym_id,),
            fetch_one=True
        )
        gym_name = gym_info['name'] if gym_info else "Salle inconnue"

        gym_title = QLabel(f"🏋️ {gym_name}")
        gym_title.setStyleSheet("""
            font-size: 20px;
            font-weight: 800;
            color: white;
            margin-bottom: 10px;
        """)
        header_layout.addWidget(gym_title)

        # Average rating with modern styling
        avg_rating = DatabaseManager.get_gym_avg_rating(self.gym_id)
        rating_text = f"⭐ Note moyenne: {avg_rating['avg']}/5"
        if avg_rating['count'] > 0:
            rating_text += f" • {avg_rating['count']} avis"
        else:
            rating_text += " • Aucun avis"

        rating_label = QLabel(rating_text)
        rating_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #ff6b35;
            background: rgba(255, 107, 53, 0.1);
            padding: 10px 15px;
            border-radius: 8px;
            border: 1px solid rgba(255, 107, 53, 0.3);
        """)
        header_layout.addWidget(rating_label)
        layout.addWidget(header_container)

        # Modern reviews list with dark scrollable area
        reviews_container = QWidget()
        reviews_container.setStyleSheet("""
            QWidget {
                background: rgba(35, 35, 35, 0.8);
                border-radius: 15px;
                border: 1px solid rgba(255, 107, 53, 0.2);
            }
        """)
        reviews_main_layout = QVBoxLayout(reviews_container)
        reviews_main_layout.setContentsMargins(20, 20, 20, 20)

        reviews_header = QLabel("💬 Commentaires des membres")
        reviews_header.setStyleSheet("""
            font-size: 18px;
            font-weight: 700;
            color: #ff6b35;
            margin-bottom: 15px;
        """)
        reviews_main_layout.addWidget(reviews_header)

        # Scrollable reviews area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(255, 107, 53, 0.1);
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 107, 53, 0.5);
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(255, 107, 53, 0.7);
            }
        """)

        content = QWidget()
        content.setStyleSheet("background: transparent;")
        self.reviews_layout = QVBoxLayout(content)
        self.reviews_layout.setSpacing(15)
        scroll.setWidget(content)
        reviews_main_layout.addWidget(scroll)
        layout.addWidget(reviews_container)

        self.load_reviews()
    
    def load_reviews(self):
        # Clear existing reviews
        for i in reversed(range(self.reviews_layout.count())):
            widget = self.reviews_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        reviews = DatabaseManager.get_gym_reviews(self.gym_id)

        if not reviews:
            # Modern empty state
            empty_container = QWidget()
            empty_container.setStyleSheet("""
                background: rgba(255, 255, 255, 0.05);
                border-radius: 12px;
                border: 2px dashed rgba(255, 107, 53, 0.3);
            """)
            empty_layout = QVBoxLayout(empty_container)
            empty_layout.setContentsMargins(40, 30, 40, 30)

            empty_icon = QLabel("💬")
            empty_icon.setStyleSheet("""
                font-size: 48px;
                color: rgba(255, 107, 53, 0.6);
                background: transparent;
            """)
            empty_icon.setAlignment(Qt.AlignCenter)

            empty_text = QLabel("Aucun avis pour cette salle")
            empty_text.setStyleSheet("""
                font-size: 16px;
                color: rgba(255, 255, 255, 0.7);
                background: transparent;
                font-weight: 500;
            """)
            empty_text.setAlignment(Qt.AlignCenter)

            empty_subtext = QLabel("Soyez le premier à laisser un avis!")
            empty_subtext.setStyleSheet("""
                font-size: 14px;
                color: rgba(255, 255, 255, 0.5);
                background: transparent;
                margin-top: 5px;
            """)
            empty_subtext.setAlignment(Qt.AlignCenter)

            empty_layout.addWidget(empty_icon)
            empty_layout.addWidget(empty_text)
            empty_layout.addWidget(empty_subtext)
            self.reviews_layout.addWidget(empty_container)
        else:
            for review in reviews:
                self.add_review_card(review)

            # Add some spacing at the end
            self.reviews_layout.addStretch()
    
    def add_review_card(self, review):
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(45, 45, 45, 0.9),
                    stop:1 rgba(35, 35, 35, 0.9));
                border-radius: 12px;
                border: 1px solid rgba(255, 107, 53, 0.2);
                padding: 0px;
            }
            QFrame:hover {
                border: 1px solid rgba(255, 107, 53, 0.4);
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(50, 50, 50, 0.9),
                    stop:1 rgba(40, 40, 40, 0.9));
            }
        """)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(12)

        # Modern header with name and rating
        header = QWidget()
        header.setStyleSheet("background: transparent;")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        # User info section
        user_info = QWidget()
        user_info.setStyleSheet("background: transparent;")
        user_info_layout = QVBoxLayout(user_info)
        user_info_layout.setContentsMargins(0, 0, 0, 0)
        user_info_layout.setSpacing(2)

        name = QLabel(f"👤 {review['first_name']} {review['last_name']}")
        name.setStyleSheet("""
            font-size: 14px;
            font-weight: 700;
            color: white;
            background: transparent;
        """)

        # Date if available
        date_label = QLabel("📅 Récemment")
        date_label.setStyleSheet("""
            font-size: 11px;
            color: rgba(255, 255, 255, 0.6);
            background: transparent;
        """)

        user_info_layout.addWidget(name)
        user_info_layout.addWidget(date_label)

        # Rating section with modern stars
        rating_container = QWidget()
        rating_container.setStyleSheet("background: transparent;")
        rating_layout = QVBoxLayout(rating_container)
        rating_layout.setContentsMargins(0, 0, 0, 0)
        rating_layout.setAlignment(Qt.AlignRight | Qt.AlignTop)

        stars = QLabel("★" * review['rating'] + "☆" * (5 - review['rating']))
        stars.setStyleSheet("""
            color: #ff6b35;
            font-size: 18px;
            background: transparent;
            font-weight: bold;
        """)

        rating_text = QLabel(f"{review['rating']}/5")
        rating_text.setStyleSheet("""
            font-size: 12px;
            color: rgba(255, 255, 255, 0.7);
            background: transparent;
            text-align: right;
        """)

        rating_layout.addWidget(stars)
        rating_layout.addWidget(rating_text)

        header_layout.addWidget(user_info)
        header_layout.addStretch()
        header_layout.addWidget(rating_container)

        layout.addWidget(header)

        # Comment section with better styling
        if review['comment']:
            comment_container = QWidget()
            comment_container.setStyleSheet("""
                background: rgba(255, 255, 255, 0.05);
                border-radius: 8px;
                border: 1px solid rgba(255, 107, 53, 0.1);
            """)
            comment_layout = QVBoxLayout(comment_container)
            comment_layout.setContentsMargins(15, 12, 15, 12)

            comment = QLabel(review['comment'])
            comment.setWordWrap(True)
            comment.setStyleSheet("""
                font-size: 13px;
                color: rgba(255, 255, 255, 0.9);
                background: transparent;
                line-height: 1.4;
            """)
            comment_layout.addWidget(comment)
            layout.addWidget(comment_container)
        else:
            # No comment placeholder
            no_comment = QLabel("💭 Aucun commentaire")
            no_comment.setStyleSheet("""
                font-size: 12px;
                color: rgba(255, 255, 255, 0.4);
                background: transparent;
                font-style: italic;
            """)
            layout.addWidget(no_comment)

        self.reviews_layout.addWidget(card)

class GymChatDialog(BaseDialog):
    def __init__(self, gym_id, user_id, parent=None):
        self.gym_id = gym_id
        self.user_id = user_id
        self.timer = QTimer()
        super().__init__(parent)


        
        
    def setup_ui(self):
        self.setWindowTitle("💬 Chat de la salle")
        self.setFixedSize(600, 700)

        # Override base dialog styling for chat
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a1a, stop:0.5 #2d2d2d, stop:1 #1a1a1a);
                border-radius: 15px;
                border: 2px solid rgba(255, 140, 0, 0.4);
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Modern chat header
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 140, 0, 0.3),
                    stop:1 rgba(255, 107, 53, 0.3));
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.4);
            }
        """)
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(15, 10, 15, 10)

        chat_title = QLabel("💬 Chat de la salle")
        chat_title.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: 700;
                color: white;
                background: transparent;
                border: none;
            }
        """)

        online_indicator = QLabel("🟢 En ligne")
        online_indicator.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: rgba(255, 255, 255, 0.8);
                background: transparent;
                border: none;
            }
        """)

        header_layout.addWidget(chat_title)
        header_layout.addStretch()
        header_layout.addWidget(online_indicator)
        layout.addWidget(header_container)

        # Modern messages area
        self.messages_area = QScrollArea()
        self.messages_area.setWidgetResizable(True)
        self.messages_area.setStyleSheet("""
            QScrollArea {
                background: rgba(35, 35, 35, 0.8);
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.2);
            }
            QScrollBar:vertical {
                background: rgba(255, 140, 0, 0.1);
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255, 140, 0, 0.5);
                border-radius: 4px;
                min-height: 20px;
            }
        """)

        self.messages_content = QWidget()
        self.messages_content.setStyleSheet("background: transparent;")
        self.messages_layout = QVBoxLayout(self.messages_content)
        self.messages_layout.setContentsMargins(15, 15, 15, 15)
        self.messages_layout.setSpacing(12)
        self.messages_layout.addStretch()
        self.messages_area.setWidget(self.messages_content)

        layout.addWidget(self.messages_area)
        
        # Modern input area
        input_container = QWidget()
        input_container.setStyleSheet("""
            QWidget {
                background: rgba(45, 45, 45, 0.9);
                border-radius: 12px;
                border: 1px solid rgba(255, 140, 0, 0.3);
            }
        """)
        input_layout = QVBoxLayout(input_container)
        input_layout.setContentsMargins(15, 12, 15, 12)
        input_layout.setSpacing(10)

        # Modern message input
        self.message_input = QTextEdit()
        self.message_input.setPlaceholderText("💭 Écrivez votre message...")
        self.message_input.setMaximumHeight(80)
        self.message_input.setStyleSheet("""
            QTextEdit {
                background: rgba(35, 35, 35, 0.8);
                border: 1px solid rgba(255, 140, 0, 0.3);
                border-radius: 8px;
                padding: 10px;
                font-size: 14px;
                color: white;
                font-weight: 500;
            }
            QTextEdit:focus {
                border: 2px solid rgba(255, 140, 0, 0.6);
                background: rgba(35, 35, 35, 0.9);
            }
        """)
        input_layout.addWidget(self.message_input)

        # Modern buttons row
        buttons_row = QHBoxLayout()
        buttons_row.setSpacing(10)

        # Modern image button
        image_btn = QPushButton("📷 Image")
        image_btn.setToolTip("Envoyer une image")
        image_btn.setStyleSheet("""
            QPushButton {
                background: rgba(74, 144, 226, 0.8);
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 8px 12px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: rgba(74, 144, 226, 1.0);
                transform: translateY(-1px);
            }
        """)
        image_btn.clicked.connect(self.send_image)
        buttons_row.addWidget(image_btn)

        # Modern audio button
        audio_btn = QPushButton("🎤 Audio")
        audio_btn.setToolTip("Envoyer un audio")
        audio_btn.setStyleSheet("""
            QPushButton {
                background: rgba(40, 167, 69, 0.8);
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding: 8px 12px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: rgba(40, 167, 69, 1.0);
                transform: translateY(-1px);
            }
        """)
        audio_btn.clicked.connect(self.send_audio)
        buttons_row.addWidget(audio_btn)

        buttons_row.addStretch()

        # Modern send button
        send_btn = QPushButton("🚀 Envoyer")
        send_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff8c00, stop:1 #ff6b35);
                color: white;
                font-size: 14px;
                font-weight: 700;
                padding: 10px 20px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #ff7700, stop:1 #ff5722);
                transform: translateY(-2px);
            }
        """)
        send_btn.clicked.connect(self.send_text_message)
        buttons_row.addWidget(send_btn)

        input_layout.addLayout(buttons_row)
        layout.addWidget(input_container)
        
        # Load initial messages
        self.load_messages()
        
        # Set up timer to refresh messages
        self.timer.timeout.connect(self.load_messages)
        self.timer.start(3000)  # Refresh every 3 seconds
    
    def load_messages(self):
        # Clear existing messages (except the stretch)
        for i in reversed(range(self.messages_layout.count())):
            widget = self.messages_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        messages = DatabaseManager.get_messages(self.gym_id)
        
        if messages:
            for msg in reversed(messages):  # Show newest at bottom
                self.add_message(msg)
    
    def add_message(self, message):
        is_me = message['sender_id'] == self.user_id

        # Container for alignment
        container = QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        container_layout = QHBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(10)

        # Profile picture for the sender (only for others' messages)
        if not is_me:
            profile_pic_container = QWidget()
            profile_pic_container.setFixedSize(40, 40)
            profile_pic_container.setStyleSheet("""
                QWidget {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #ff8c00, stop:1 #ff6b35);
                    border-radius: 20px;
                    border: 1px solid rgba(255, 255, 255, 0.3);
                }
            """)

            profile_pic_layout = QVBoxLayout(profile_pic_container)
            profile_pic_layout.setContentsMargins(0, 0, 0, 0)
            profile_pic_layout.setAlignment(Qt.AlignCenter)

            try:
                # Get sender's profile picture
                sender_data = DatabaseManager.execute_query("""
                    SELECT profile_picture FROM users WHERE id = %s
                """, (message['sender_id'],), fetch_one=True)

                if sender_data and sender_data['profile_picture'] is not None:
                    # Load profile picture
                    profile_pic_label = QLabel()
                    pixmap = QPixmap()

                    # Check if profile_picture is bytes or string path
                    if isinstance(sender_data['profile_picture'], bytes):
                        # Load from binary data
                        success = pixmap.loadFromData(sender_data['profile_picture'])
                    elif isinstance(sender_data['profile_picture'], str) and sender_data['profile_picture'].strip():
                        # Load from file path
                        success = pixmap.load(sender_data['profile_picture'])
                    else:
                        success = False

                    if success and not pixmap.isNull():
                        scaled_pixmap = pixmap.scaled(36, 36, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        profile_pic_label.setPixmap(scaled_pixmap)
                        profile_pic_label.setStyleSheet("border-radius: 18px; background: transparent;")
                        profile_pic_layout.addWidget(profile_pic_label)
                    else:
                        # Default user icon
                        default_icon = QLabel("👤")
                        default_icon.setStyleSheet("""
                            font-size: 16px;
                            color: white;
                            background: transparent;
                            border: none;
                        """)
                        default_icon.setAlignment(Qt.AlignCenter)
                        profile_pic_layout.addWidget(default_icon)
                else:
                    # Default user icon
                    default_icon = QLabel("👤")
                    default_icon.setStyleSheet("""
                        font-size: 16px;
                        color: white;
                        background: transparent;
                        border: none;
                    """)
                    default_icon.setAlignment(Qt.AlignCenter)
                    profile_pic_layout.addWidget(default_icon)
            except Exception as e:
                print(f"Error loading sender profile picture: {e}")
                # Default user icon
                default_icon = QLabel("👤")
                default_icon.setStyleSheet("""
                    font-size: 16px;
                    color: white;
                    background: transparent;
                    border: none;
                """)
                default_icon.setAlignment(Qt.AlignCenter)
                profile_pic_layout.addWidget(default_icon)

            container_layout.addWidget(profile_pic_container)

        if is_me:
            container_layout.addStretch()  # Push to right

        # Modern message bubble
        message_widget = QWidget()
        message_widget.setMaximumWidth(400)

        if is_me:
            # My messages - orange gradient
            message_widget.setStyleSheet("""
                QWidget {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #ff8c00, stop:1 #ff6b35);
                    border-radius: 15px;
                    border-top-right-radius: 5px;
                    padding: 12px;
                    margin: 2px;
                }
            """)
        else:
            # Others' messages - dark theme
            message_widget.setStyleSheet("""
                QWidget {
                    background: rgba(55, 55, 55, 0.9);
                    border-radius: 15px;
                    border-top-left-radius: 5px;
                    border: 1px solid rgba(255, 140, 0, 0.2);
                    padding: 12px;
                    margin: 2px;
                }
            """)

        layout = QVBoxLayout(message_widget)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(4)

        # Modern header with name and time
        if not is_me:  # Only show name for others
            name = QLabel(f"{message['first_name']} {message['last_name']}")
            name.setStyleSheet("""
                QLabel {
                    font-weight: 600;
                    font-size: 12px;
                    color: rgba(255, 140, 0, 0.9);
                    background: transparent;
                    border: none;
                    margin-bottom: 2px;
                }
            """)
            layout.addWidget(name)

        # Time stamp
        time = QLabel(message['created_at'].strftime("%H:%M"))
        time.setStyleSheet(f"""
            QLabel {{
                color: {'rgba(255, 255, 255, 0.7)' if is_me else 'rgba(255, 255, 255, 0.6)'};
                font-size: 10px;
                background: transparent;
                border: none;
                font-weight: 500;
            }}
        """)
        time.setAlignment(Qt.AlignRight if is_me else Qt.AlignLeft)
        
        # Modern message content
        if message['message_type'] == 'text':
            content = QLabel(message['content'])
            content.setWordWrap(True)
            content.setStyleSheet(f"""
                QLabel {{
                    color: {'white' if is_me else 'white'};
                    font-size: 14px;
                    font-weight: 500;
                    background: transparent;
                    border: none;
                    line-height: 1.4;
                }}
            """)
            layout.addWidget(content)
        elif message['message_type'] == 'image':
            btn = QPushButton("📷 Voir l'image")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {'rgba(255, 255, 255, 0.2)' if is_me else 'rgba(255, 140, 0, 0.3)'};
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-size: 12px;
                    font-weight: 600;
                }}
                QPushButton:hover {{
                    background: {'rgba(255, 255, 255, 0.3)' if is_me else 'rgba(255, 140, 0, 0.5)'};
                }}
            """)
            btn.clicked.connect(lambda _, mid=message['id']: self.show_image(mid))
            layout.addWidget(btn)
        elif message['message_type'] == 'audio':
            btn = QPushButton("🎤 Écouter l'audio")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {'rgba(255, 255, 255, 0.2)' if is_me else 'rgba(255, 140, 0, 0.3)'};
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-size: 12px;
                    font-weight: 600;
                }}
                QPushButton:hover {{
                    background: {'rgba(255, 255, 255, 0.3)' if is_me else 'rgba(255, 140, 0, 0.5)'};
                }}
            """)
            btn.clicked.connect(lambda _, mid=message['id']: self.play_audio(mid))
            layout.addWidget(btn)

        layout.addWidget(time)

        container_layout.addWidget(message_widget)

        if not is_me:
            container_layout.addStretch()  # Push to left

        self.messages_layout.insertWidget(0, container)
    
    def send_text_message(self):
        text = self.message_input.toPlainText().strip()
        if text:
            if DatabaseManager.send_message(
                self.gym_id, 
                self.user_id, 
                'text', 
                content=text
            ):
                self.message_input.clear()
                self.load_messages()
    
    def send_image(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner une image", "", "Images (*.png *.jpg *.jpeg)"
        )
        if file_path:
            if DatabaseManager.send_message(
                self.gym_id,
                self.user_id,
                'image',
                media_path=file_path
            ):
                self.load_messages()
    
    def send_audio(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner un fichier audio", "", "Audio (*.mp3 *.wav)"
        )
        if file_path:
            if DatabaseManager.send_message(
                self.gym_id,
                self.user_id,
                'audio',
                media_path=file_path
            ):
                self.load_messages()
    
    def show_image(self, message_id):
        media = DatabaseManager.get_message_media(message_id)
        if media and media['media']:
            dialog = QDialog(self)
            dialog.setWindowTitle("Image")
            layout = QVBoxLayout(dialog)
            
            pixmap = QPixmap()
            pixmap.loadFromData(media['media'])
            label = QLabel()
            label.setPixmap(pixmap.scaled(400, 400, Qt.KeepAspectRatio))
            
            layout.addWidget(label)
            dialog.exec()
    
    def play_audio(self, message_id):
        media = DatabaseManager.get_message_media(message_id)
        if media and media['media']:
            # Créer un fichier temporaire pour jouer l'audio
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as tmp:
                tmp.write(media['media'])
                tmp_path = tmp.name
            
            # Jouer l'audio (nécessite des bibliothèques supplémentaires comme pygame)
            try:
                import pygame
                pygame.mixer.init()
                pygame.mixer.music.load(tmp_path)
                pygame.mixer.music.play()
            except ImportError:
                QMessageBox.warning(self, "Erreur", "La lecture audio nécessite pygame. Installez-le avec: pip install pygame")
    
    def closeEvent(self, event):
        self.timer.stop()
        super().closeEvent(event)
class ModifyGymDialog(BaseDialog):
    def __init__(self, gym_id, parent=None):
        self.gym_id = gym_id
        super().__init__(parent)
        
    def setup_ui(self):
        self.setWindowTitle("Modifier la salle")
        self.setFixedSize(500, 400)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Charger les infos actuelles
        self.gym_info = DatabaseManager.execute_query(
            "SELECT * FROM gyms WHERE id = %s", 
            (self.gym_id,), 
            fetch_one=True
        )
        
        # Nom de la salle
        name_label = QLabel("Nom:")
        self.name_input = QLineEdit(self.gym_info['name'])
        layout.addWidget(name_label)
        layout.addWidget(self.name_input)
        
        # Adresse
        address_label = QLabel("Adresse:")
        self.address_input = QLineEdit(self.gym_info['address'])
        layout.addWidget(address_label)
        layout.addWidget(self.address_input)
        
        # Options d'abonnement
        subs_label = QLabel("Options d'abonnement:")
        layout.addWidget(subs_label)
        
        try:
            self.subscriptions = json.loads(self.gym_info['subscriptions']) if self.gym_info['subscriptions'] else {}
        except:
            self.subscriptions = {}
            
        # Mensuel
        monthly_group = QGroupBox("Mensuel")
        monthly_layout = QHBoxLayout(monthly_group)
        self.monthly_price = QLineEdit(str(self.subscriptions.get('monthly', {}).get('price', '')))
        self.monthly_price.setValidator(QDoubleValidator(0, 999, 2))
        monthly_layout.addWidget(QLabel("Prix:"))
        monthly_layout.addWidget(self.monthly_price)
        layout.addWidget(monthly_group)
        
        # Trimestriel
        quarterly_group = QGroupBox("Trimestriel")
        quarterly_layout = QHBoxLayout(quarterly_group)
        self.quarterly_price = QLineEdit(str(self.subscriptions.get('quarterly', {}).get('price', '')))
        self.quarterly_price.setValidator(QDoubleValidator(0, 999, 2))
        quarterly_layout.addWidget(QLabel("Prix:"))
        quarterly_layout.addWidget(self.quarterly_price)
        layout.addWidget(quarterly_group)
        
        # Annuel (nouveau)
        annual_group = QGroupBox("Annuel")
        annual_layout = QHBoxLayout(annual_group)
        self.annual_price = QLineEdit(str(self.subscriptions.get('annual', {}).get('price', '')))
        self.annual_price.setValidator(QDoubleValidator(0, 999, 2))
        annual_layout.addWidget(QLabel("Prix:"))
        annual_layout.addWidget(self.annual_price)
        layout.addWidget(annual_group)
        
        # Boutons
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Enregistrer")
        save_btn.clicked.connect(self.save_changes)
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
        
    def save_changes(self):
        # Préparer les données d'abonnement
        subscriptions = {
            "monthly": {
                "duration": "1 mois",
                "price": float(self.monthly_price.text()) if self.monthly_price.text() else 0
            },
            "quarterly": {
                "duration": "3 mois", 
                "price": float(self.quarterly_price.text()) if self.quarterly_price.text() else 0
            },
            "annual": {
                "duration": "12 mois",
                "price": float(self.annual_price.text()) if self.annual_price.text() else 0
            }
        }
        
        # Mettre à jour en base de données
        success = DatabaseManager.execute_query("""
            UPDATE gyms 
            SET name = %s, address = %s, subscriptions = %s
            WHERE id = %s
        """, (
            self.name_input.text(),
            self.address_input.text(),
            json.dumps(subscriptions),
            self.gym_id
        ))
        
        if success:
            QMessageBox.information(self, "Succès", "Modifications enregistrées!")
            self.accept()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de la mise à jour")
def main():
    app = QApplication(sys.argv)
        # Après avoir créé votre QApplication
    
    DatabaseManager.initialize_database()
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()